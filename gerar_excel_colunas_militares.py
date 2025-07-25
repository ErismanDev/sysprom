#!/usr/bin/env python
"""
Script para gerar um arquivo Excel com os nomes das colunas da tabela de militares
"""

import os
import sys
import django
from datetime import datetime

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sepromcbmepi.settings')
django.setup()

from militares.models import Militar

def gerar_excel_colunas_militares():
    """Gera um arquivo Excel com os nomes das colunas da tabela de militares"""
    
    print("📊 Gerando arquivo Excel com colunas da tabela de militares...")
    
    # Obter informações dos campos do modelo Militar
    campos = []
    
    for field in Militar._meta.fields:
        campo_info = {
            'nome_campo': field.name,
            'tipo_campo': field.get_internal_type(),
            'verbose_name': getattr(field, 'verbose_name', field.name),
            'max_length': getattr(field, 'max_length', ''),
            'null': field.null,
            'blank': field.blank,
            'default': field.default,
            'help_text': getattr(field, 'help_text', ''),
            'choices': getattr(field, 'choices', []),
        }
        campos.append(campo_info)
    
    # Criar arquivo Excel usando openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Colunas Tabela Militares"
        
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos
        headers = [
            'Nome do Campo',
            'Tipo do Campo',
            'Nome de Exibição',
            'Tamanho Máximo',
            'Permite Nulo',
            'Permite Vazio',
            'Valor Padrão',
            'Texto de Ajuda',
            'Opções de Escolha'
        ]
        
        # Aplicar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Dados dos campos
        for row, campo in enumerate(campos, 2):
            # Nome do campo
            ws.cell(row=row, column=1, value=campo['nome_campo']).border = border
            
            # Tipo do campo
            ws.cell(row=row, column=2, value=campo['tipo_campo']).border = border
            
            # Nome de exibição
            ws.cell(row=row, column=3, value=campo['verbose_name']).border = border
            
            # Tamanho máximo
            max_length = campo['max_length'] if campo['max_length'] else ''
            ws.cell(row=row, column=4, value=max_length).border = border
            
            # Permite nulo
            ws.cell(row=row, column=5, value="Sim" if campo['null'] else "Não").border = border
            
            # Permite vazio
            ws.cell(row=row, column=6, value="Sim" if campo['blank'] else "Não").border = border
            
            # Valor padrão
            default = campo['default']
            if default is not None:
                if callable(default):
                    default_str = "Função"
                else:
                    default_str = str(default)
            else:
                default_str = ""
            ws.cell(row=row, column=7, value=default_str).border = border
            
            # Texto de ajuda
            ws.cell(row=row, column=8, value=campo['help_text']).border = border
            
            # Opções de escolha
            choices = campo['choices']
            if choices:
                choices_str = "; ".join([f"{choice[0]}={choice[1]}" for choice in choices])
            else:
                choices_str = ""
            ws.cell(row=row, column=9, value=choices_str).border = border
        
        # Ajustar largura das colunas
        column_widths = [20, 15, 25, 15, 12, 12, 15, 30, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Criar segunda aba com informações gerais
        ws2 = wb.create_sheet("Informações Gerais")
        
        # Informações gerais
        info_geral = [
            ["Informação", "Valor"],
            ["Nome da Tabela", "militares_militar"],
            ["Nome do Modelo", "Militar"],
            ["Total de Campos", len(campos)],
            ["Data de Geração", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
            ["Sistema", "SEPROMCBMEPI"],
            ["", ""],
            ["Campos Obrigatórios", ""],
        ]
        
        # Adicionar campos obrigatórios
        campos_obrigatorios = [campo for campo in campos if not campo['null'] and not campo['blank']]
        for campo in campos_obrigatorios:
            info_geral.append([campo['nome_campo'], campo['verbose_name']])
        
        # Aplicar informações gerais
        for row, (info, valor) in enumerate(info_geral, 1):
            ws2.cell(row=row, column=1, value=info).font = Font(bold=True)
            ws2.cell(row=row, column=2, value=valor)
        
        # Ajustar largura das colunas da segunda aba
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 40
        
        # Salvar arquivo
        nome_arquivo = f"colunas_tabela_militares_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(nome_arquivo)
        
        print(f"✅ Arquivo Excel gerado com sucesso: {nome_arquivo}")
        print(f"📋 Total de campos: {len(campos)}")
        print(f"🔗 Campos obrigatórios: {len(campos_obrigatorios)}")
        
        return nome_arquivo
        
    except ImportError:
        print("❌ Erro: Biblioteca openpyxl não encontrada")
        print("💡 Instale com: pip install openpyxl")
        return None
    except Exception as e:
        print(f"❌ Erro ao gerar arquivo Excel: {e}")
        return None

def gerar_csv_colunas_militares():
    """Gera um arquivo CSV com os nomes das colunas da tabela de militares (alternativa)"""
    
    print("📊 Gerando arquivo CSV com colunas da tabela de militares...")
    
    # Obter informações dos campos do modelo Militar
    campos = []
    
    for field in Militar._meta.fields:
        campo_info = {
            'nome_campo': field.name,
            'tipo_campo': field.get_internal_type(),
            'verbose_name': getattr(field, 'verbose_name', field.name),
            'max_length': getattr(field, 'max_length', ''),
            'null': field.null,
            'blank': field.blank,
            'default': field.default,
            'help_text': getattr(field, 'help_text', ''),
            'choices': getattr(field, 'choices', []),
        }
        campos.append(campo_info)
    
    # Criar arquivo CSV
    nome_arquivo = f"colunas_tabela_militares_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    try:
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            # Cabeçalhos
            f.write("Nome do Campo,Tipo do Campo,Nome de Exibição,Tamanho Máximo,Permite Nulo,Permite Vazio,Valor Padrão,Texto de Ajuda,Opções de Escolha\n")
            
            # Dados dos campos
            for campo in campos:
                # Opções de escolha
                choices = campo['choices']
                if choices:
                    choices_str = "; ".join([f"{choice[0]}={choice[1]}" for choice in choices])
                else:
                    choices_str = ""
                
                # Valor padrão
                default = campo['default']
                if default is not None:
                    if callable(default):
                        default_str = "Função"
                    else:
                        default_str = str(default)
                else:
                    default_str = ""
                
                # Escrever linha
                linha = [
                    campo['nome_campo'],
                    campo['tipo_campo'],
                    campo['verbose_name'],
                    str(campo['max_length']) if campo['max_length'] else '',
                    "Sim" if campo['null'] else "Não",
                    "Sim" if campo['blank'] else "Não",
                    default_str,
                    campo['help_text'],
                    choices_str
                ]
                
                # Escapar vírgulas e aspas
                linha_escaped = []
                for item in linha:
                    if ',' in str(item) or '"' in str(item):
                        item_escaped = '"' + str(item).replace('"', '""') + '"'
                    else:
                        item_escaped = str(item)
                    linha_escaped.append(item_escaped)
                
                f.write(','.join(linha_escaped) + '\n')
        
        print(f"✅ Arquivo CSV gerado com sucesso: {nome_arquivo}")
        print(f"📋 Total de campos: {len(campos)}")
        
        return nome_arquivo
        
    except Exception as e:
        print(f"❌ Erro ao gerar arquivo CSV: {e}")
        return None

if __name__ == "__main__":
    print("🔧 Gerador de Colunas da Tabela de Militares")
    print("=" * 50)
    
    # Tentar gerar Excel primeiro
    arquivo_excel = gerar_excel_colunas_militares()
    
    if not arquivo_excel:
        print("🔄 Tentando gerar arquivo CSV como alternativa...")
        arquivo_csv = gerar_csv_colunas_militares()
        
        if arquivo_csv:
            print("✅ Arquivo CSV gerado como alternativa")
        else:
            print("❌ Não foi possível gerar nenhum arquivo")
    else:
        print("✅ Processo concluído com sucesso!") 