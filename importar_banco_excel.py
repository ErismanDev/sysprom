#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para importar dados do Excel para o banco no servidor
"""
import os
import sys
import django
import glob
import time
import random
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sepromcbmepi.settings')
django.setup()

from django.contrib.auth.models import User
from militares.models import Militar, QuadroAcesso, Promocao
from django.db import transaction

def importar_usuarios(workbook):
    """Importar usuários"""
    print("📥 Importando usuários...")
    
    try:
        ws = workbook["Usuários"]
    except KeyError:
        print("   ⚠️ Planilha 'Usuários' não encontrada")
        return
    
    # Ler dados
    data = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = row
            continue
        if row[0] is None:
            break
        data.append(dict(zip(headers, row)))
    
    if not data:
        print("   ⚠️ Nenhum usuário para importar")
        return
    
    # Importar
    importados = 0
    atualizados = 0
    
    with transaction.atomic():
        for item in data:
            username = item.get('username')
            if not username:
                continue
            
            # Garantir que first_name e last_name não sejam None ou vazios
            first_name = item.get('first_name') or 'Usuário'
            last_name = item.get('last_name') or 'Sistema'
            email = item.get('email') or ''
            
            user, created = User.objects.update_or_create(
                username=username,
                defaults={
                    'email': email,
                    'first_name': first_name,
                    'last_name': last_name,
                    'is_active': bool(item.get('is_active', True)),
                    'is_staff': bool(item.get('is_staff', False)),
                    'is_superuser': bool(item.get('is_superuser', False)),
                }
            )
            
            if created:
                importados += 1
            else:
                atualizados += 1
    
    print(f"   ✅ {importados} usuários importados, {atualizados} atualizados")

def importar_militares(workbook):
    """Importar militares"""
    print("📥 Importando militares...")
    
    try:
        ws = workbook["Militares"]
    except KeyError:
        print("   ⚠️ Planilha 'Militares' não encontrada")
        return
    
    # Ler dados
    data = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = row
            continue
        if row[0] is None:
            break
        data.append(dict(zip(headers, row)))
    
    if not data:
        print("   ⚠️ Nenhum militar para importar")
        return
    
    # Importar
    importados = 0
    atualizados = 0
    
    with transaction.atomic():
        for item in data:
            matricula = item.get('matricula')
            if not matricula:
                continue
            
            # Preparar dados (garantir valores não-nulos para campos obrigatórios)
            nome_completo = item.get('nome_completo') or 'Militar'
            nome_guerra = item.get('nome_guerra') or nome_completo.split()[0] if nome_completo else 'Militar'
            
            # Gerar CPF único se não houver (baseado na matrícula)
            cpf = item.get('cpf')
            if not cpf or cpf == '000.000.000-00' or cpf == '' or cpf is None:
                # Gerar CPF único baseado na matrícula + índice para garantir unicidade
                matricula_limpa = matricula.replace('-', '').replace('.', '').replace('/', '')[:9]
                # Adicionar índice do loop para garantir unicidade
                cpf_base = f"{matricula_limpa.zfill(7)}{importados % 100:02d}-{random.randint(10, 99)}"
                # Verificar se CPF já existe, se sim, gerar outro
                tentativas = 0
                while Militar.objects.filter(cpf=cpf_base).exists() and tentativas < 10:
                    cpf_base = f"{matricula_limpa.zfill(7)}{(importados + tentativas) % 100:02d}-{random.randint(10, 99)}"
                    tentativas += 1
                cpf = cpf_base
            
            # Gerar RG único se não houver
            rg = item.get('rg')
            if not rg or rg == '0000000' or rg == '':
                rg = f"RG{matricula.replace('-', '').replace('.', '').replace('/', '')[:7]}"
            
            militar_data = {
                'nome_completo': nome_completo,
                'nome_guerra': nome_guerra,
                'cpf': cpf,
                'rg': rg,
                'data_nascimento': item.get('data_nascimento'),
                'sexo': item.get('sexo') or 'M',
                'quadro': item.get('quadro') or 'COMB',
                'posto_graduacao': item.get('posto_graduacao') or 'SD',
                'data_ingresso': item.get('data_ingresso'),
                'data_promocao_atual': item.get('data_promocao_atual'),
                'classificacao': item.get('classificacao') or 'ATIVO',
                'email': item.get('email') or '',
                'telefone': item.get('telefone') or '',
                'celular': item.get('celular') or '',
            }
            
            # Remover apenas campos None (manter strings vazias onde permitido)
            militar_data = {k: v for k, v in militar_data.items() if v is not None}
            
            # Vincular usuário se existir
            user_id = item.get('user_id')
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                    militar_data['user'] = user
                except User.DoesNotExist:
                    pass
            
            # Tentar atualizar ou criar por matrícula primeiro
            try:
                militar = Militar.objects.get(matricula=matricula)
                # Atualizar
                for key, value in militar_data.items():
                    setattr(militar, key, value)
                militar.save()
                atualizados += 1
                if atualizados <= 5:
                    print(f"   🔄 Atualizado: {matricula}")
            except Militar.DoesNotExist:
                # Verificar se CPF já existe antes de criar
                if Militar.objects.filter(cpf=cpf).exists():
                    # Se CPF já existe, gerar novo CPF único
                    matricula_limpa = matricula.replace('-', '').replace('.', '').replace('/', '')[:9]
                    tentativas = 0
                    while Militar.objects.filter(cpf=cpf).exists() and tentativas < 20:
                        cpf = f"{matricula_limpa.zfill(7)}{(importados + tentativas) % 100:02d}-{random.randint(10, 99)}"
                        tentativas += 1
                    militar_data['cpf'] = cpf
                    if tentativas >= 20:
                        # Se ainda não conseguir, usar timestamp
                        cpf = f"{matricula_limpa.zfill(5)}{int(time.time()) % 100000:05d}-00"
                        militar_data['cpf'] = cpf
                
                # Criar novo
                try:
                    militar = Militar.objects.create(**militar_data)
                    importados += 1
                    if importados <= 5:
                        print(f"   ✅ Criado: {matricula} - {nome_completo}")
                except Exception as e:
                    # Se der erro de CPF duplicado, gerar CPF único
                    if 'cpf' in str(e).lower() or 'unique' in str(e).lower():
                        matricula_limpa = matricula.replace('-', '').replace('.', '').replace('/', '')[:9]
                        cpf_novo = f"{matricula_limpa.zfill(7)}{(importados + 1) % 100:02d}-{random.randint(10, 99)}"
                        tentativas = 0
                        while Militar.objects.filter(cpf=cpf_novo).exists() and tentativas < 10:
                            cpf_novo = f"{matricula_limpa.zfill(7)}{(importados + tentativas + 1) % 100:02d}-{random.randint(10, 99)}"
                            tentativas += 1
                        militar_data['cpf'] = cpf_novo
                        try:
                            militar = Militar.objects.create(**militar_data)
                            importados += 1
                            if importados <= 5:
                                print(f"   ✅ Criado com CPF único: {matricula} - {nome_completo}")
                        except:
                            # Se ainda der erro, usar timestamp
                            cpf_novo = f"{matricula_limpa.zfill(5)}{int(time.time()) % 100000:05d}-00"
                            militar_data['cpf'] = cpf_novo
                            militar = Militar.objects.create(**militar_data)
                            importados += 1
                            if importados <= 5:
                                print(f"   ✅ Criado com CPF timestamp: {matricula} - {nome_completo}")
                    else:
                        raise
    
    print(f"   ✅ {importados} militares importados, {atualizados} atualizados")

def importar_quadro_acesso(workbook):
    """Importar quadro de acesso"""
    print("📥 Importando quadro de acesso...")
    
    try:
        ws = workbook["QuadroAcesso"]
    except KeyError:
        print("   ⚠️ Planilha 'QuadroAcesso' não encontrada")
        return
    
    # Ler dados
    data = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = row
            continue
        if row[0] is None:
            break
        data.append(dict(zip(headers, row)))
    
    if not data:
        print("   ⚠️ Nenhum quadro de acesso para importar")
        return
    
    # Importar
    importados = 0
    atualizados = 0
    
    with transaction.atomic():
        for item in data:
            militar_id = item.get('militar_id')
            if not militar_id:
                continue
            
            try:
                militar = Militar.objects.get(id=militar_id)
            except Militar.DoesNotExist:
                print(f"   ⚠️ Militar ID {militar_id} não encontrado, pulando...")
                continue
            
            quadro_data = {
                'militar': militar,
                'quadro': item.get('quadro', 'COMB'),
                'posto_graduacao': item.get('posto_graduacao', 'SD'),
                'categoria': item.get('categoria', ''),
                'numero': item.get('numero'),
                'data_documento': item.get('data_documento'),
                'homologado': item.get('homologado', False),
                'data_homologacao': item.get('data_homologacao'),
            }
            
            # Remover campos None
            quadro_data = {k: v for k, v in quadro_data.items() if v is not None}
            
            quadro, created = QuadroAcesso.objects.update_or_create(
                id=item.get('id'),
                defaults=quadro_data
            )
            
            if created:
                importados += 1
            else:
                atualizados += 1
    
    print(f"   ✅ {importados} quadros importados, {atualizados} atualizados")

def main():
    # Procurar arquivo Excel mais recente
    arquivos = glob.glob("/home/seprom/sepromcbmepi/backup_sepromcbmepi_*.xlsx")
    
    if not arquivos:
        print("❌ Nenhum arquivo Excel encontrado em /home/seprom/sepromcbmepi/")
        print("💡 Envie o arquivo via WinSCP primeiro")
        sys.exit(1)
    
    arquivo_excel = sorted(arquivos, reverse=True)[0]
    
    print(f"📦 Importando dados do Excel...")
    print(f"📁 Arquivo: {arquivo_excel}")
    print("")
    
    try:
        # Carregar workbook
        wb = load_workbook(arquivo_excel, data_only=True)
        
        # Importar cada planilha
        importar_usuarios(wb)
        importar_militares(wb)
        importar_quadro_acesso(wb)
        
        print("")
        print("✅ Importação concluída com sucesso!")
        
    except Exception as e:
        print(f"\n❌ Erro ao importar: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

