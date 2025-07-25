import openpyxl
from django.core.management.base import BaseCommand
from militares.models import Militar
from django.db import transaction

class Command(BaseCommand):
    help = 'Atualiza a antiguidade dos militares conforme o arquivo Excel Efetivo CBMEPI Atito SI Promoção.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='Efetivo CBMEPI Atito SI Promoção.xlsx',
            help='Caminho para o arquivo Excel'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Executa sem salvar no banco'
        )

    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        
        try:
            # Abrir o arquivo Excel
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            self.stdout.write(f'Arquivo Excel aberto: {file_path}')
            
            # Encontrar as colunas necessárias
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            self.stdout.write(f'Colunas encontradas: {headers}')
            
            # Procurar pelas colunas necessárias
            matricula_col = None
            ord_col = None
            
            for i, header in enumerate(headers, 1):
                if header and 'matr' in str(header).lower():
                    matricula_col = i
                elif header and 'ord' in str(header).lower():
                    ord_col = i
            
            if not matricula_col or not ord_col:
                self.stdout.write(
                    self.style.ERROR('Colunas "Matrícula" e "ord" não encontradas no arquivo')
                )
                return
            
            self.stdout.write(f'Coluna Matrícula: {matricula_col} ({headers[matricula_col-1]})')
            self.stdout.write(f'Coluna ord: {ord_col} ({headers[ord_col-1]})')
            
            atualizados = 0
            nao_encontrados = 0
            
            # Processar cada linha do Excel
            for row_num in range(2, sheet.max_row + 1):
                matricula = sheet.cell(row=row_num, column=matricula_col).value
                ord_value = sheet.cell(row=row_num, column=ord_col).value
                
                if not matricula or not ord_value:
                    continue
                
                # Limpar a matrícula
                matricula = str(matricula).strip()
                
                try:
                    militar = Militar.objects.get(matricula=matricula)
                    if str(militar.antiguidade) != str(ord_value):
                        self.stdout.write(f'{militar.nome_guerra} ({matricula}): {militar.antiguidade} -> {ord_value}')
                        if not dry_run:
                            militar.antiguidade = str(ord_value)
                            militar.save()
                        atualizados += 1
                except Militar.DoesNotExist:
                    nao_encontrados += 1
                    self.stdout.write(f'Militar não encontrado: {matricula}')
            
            workbook.close()
            
            self.stdout.write(self.style.SUCCESS(f'{atualizados} militares atualizados.'))
            if nao_encontrados:
                self.stdout.write(self.style.WARNING(f'{nao_encontrados} matrículas não encontradas no banco.'))
                
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Erro ao processar arquivo: {str(e)}')
            ) 