#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para importar dados do Excel para o banco no servidor (com debug)
"""
import os
import sys
import django
import glob
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'sepromcbmepi.settings')
django.setup()

from django.contrib.auth.models import User
from militares.models import Militar, QuadroAcesso, Promocao
from django.db import transaction

def importar_usuarios(workbook):
    """Importar usuários"""
    print("📥 Importando usuários...")
    
    try:
        ws = workbook["Usuários"]
    except KeyError:
        print("   ⚠️ Planilha 'Usuários' não encontrada")
        return
    
    # Ler dados
    data = []
    headers = None
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if headers is None:
            headers = row
            print(f"   📋 Cabeçalhos: {headers}")
            continue
        if row[0] is None:
            print(f"   ⚠️ Linha {row_count} vazia, parando")
            break
        data.append(dict(zip(headers, row)))
        if len(data) <= 3:
            print(f"   📝 Linha {row_count}: username={row[1] if len(row) > 1 else 'N/A'}")
    
    print(f"   📊 Total de linhas lidas: {len(data)}")
    
    if not data:
        print("   ⚠️ Nenhum usuário para importar")
        return
    
    # Importar
    importados = 0
    atualizados = 0
    erros = 0
    
    with transaction.atomic():
        for idx, item in enumerate(data):
            try:
                username = item.get('username')
                if not username:
                    print(f"   ⚠️ Linha {idx+2}: username vazio, pulando")
                    continue
                
                user, created = User.objects.update_or_create(
                    username=username,
                    defaults={
                        'email': item.get('email', '') or '',
                        'first_name': item.get('first_name', '') or '',
                        'last_name': item.get('last_name', '') or '',
                        'is_active': bool(item.get('is_active', True)),
                        'is_staff': bool(item.get('is_staff', False)),
                        'is_superuser': bool(item.get('is_superuser', False)),
                    }
                )
                
                if created:
                    importados += 1
                    if importados <= 5:
                        print(f"   ✅ Criado: {username}")
                else:
                    atualizados += 1
                    if atualizados <= 5:
                        print(f"   🔄 Atualizado: {username}")
            except Exception as e:
                erros += 1
                print(f"   ❌ Erro na linha {idx+2}: {e}")
                if erros > 10:
                    print("   ⚠️ Muitos erros, parando...")
                    break
    
    print(f"   ✅ {importados} usuários importados, {atualizados} atualizados, {erros} erros")

def importar_militares(workbook):
    """Importar militares"""
    print("📥 Importando militares...")
    
    try:
        ws = workbook["Militares"]
    except KeyError:
        print("   ⚠️ Planilha 'Militares' não encontrada")
        return
    
    # Ler dados
    data = []
    headers = None
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if headers is None:
            headers = row
            print(f"   📋 Cabeçalhos: {headers[:5]}...")
            continue
        if row[0] is None:
            print(f"   ⚠️ Linha {row_count} vazia, parando")
            break
        data.append(dict(zip(headers, row)))
        if len(data) <= 3:
            print(f"   📝 Linha {row_count}: matricula={row[1] if len(row) > 1 else 'N/A'}")
    
    print(f"   📊 Total de linhas lidas: {len(data)}")
    
    if not data:
        print("   ⚠️ Nenhum militar para importar")
        return
    
    # Importar
    importados = 0
    atualizados = 0
    erros = 0
    
    with transaction.atomic():
        for idx, item in enumerate(data):
            try:
                matricula = item.get('matricula')
                if not matricula:
                    print(f"   ⚠️ Linha {idx+2}: matricula vazia, pulando")
                    continue
                
                # Preparar dados
                militar_data = {
                    'nome_completo': item.get('nome_completo', '') or '',
                    'nome_guerra': item.get('nome_guerra', '') or '',
                    'cpf': item.get('cpf', '') or '',
                    'rg': item.get('rg', '') or '',
                    'data_nascimento': item.get('data_nascimento'),
                    'sexo': item.get('sexo', 'M') or 'M',
                    'quadro': item.get('quadro', 'COMB') or 'COMB',
                    'posto_graduacao': item.get('posto_graduacao', 'SD') or 'SD',
                    'data_ingresso': item.get('data_ingresso'),
                    'data_promocao_atual': item.get('data_promocao_atual'),
                    'classificacao': item.get('classificacao', 'ATIVO') or 'ATIVO',
                    'email': item.get('email', '') or '',
                    'telefone': item.get('telefone', '') or '',
                    'celular': item.get('celular', '') or '',
                }
                
                # Remover campos None
                militar_data = {k: v for k, v in militar_data.items() if v is not None and v != ''}
                
                # Vincular usuário se existir
                user_id = item.get('user_id')
                if user_id:
                    try:
                        user = User.objects.get(id=user_id)
                        militar_data['user'] = user
                    except User.DoesNotExist:
                        pass
                
                militar, created = Militar.objects.update_or_create(
                    matricula=matricula,
                    defaults=militar_data
                )
                
                if created:
                    importados += 1
                    if importados <= 5:
                        print(f"   ✅ Criado: {matricula} - {militar_data.get('nome_completo', 'N/A')}")
                else:
                    atualizados += 1
                    if atualizados <= 5:
                        print(f"   🔄 Atualizado: {matricula}")
            except Exception as e:
                erros += 1
                print(f"   ❌ Erro na linha {idx+2}: {e}")
                if erros > 10:
                    print("   ⚠️ Muitos erros, parando...")
                    break
    
    print(f"   ✅ {importados} militares importados, {atualizados} atualizados, {erros} erros")

def main():
    # Procurar arquivo Excel mais recente
    arquivos = glob.glob("/home/seprom/sepromcbmepi/backup_sepromcbmepi_*.xlsx")
    
    if not arquivos:
        print("❌ Nenhum arquivo Excel encontrado em /home/seprom/sepromcbmepi/")
        print("💡 Envie o arquivo via WinSCP primeiro")
        sys.exit(1)
    
    arquivo_excel = sorted(arquivos, reverse=True)[0]
    
    print(f"📦 Importando dados do Excel...")
    print(f"📁 Arquivo: {arquivo_excel}")
    print("")
    
    try:
        # Carregar workbook
        wb = load_workbook(arquivo_excel, data_only=True)
        
        # Importar cada planilha
        importar_usuarios(wb)
        print("")
        importar_militares(wb)
        
        print("")
        print("✅ Importação concluída!")
        
    except Exception as e:
        print(f"\n❌ Erro ao importar: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()

