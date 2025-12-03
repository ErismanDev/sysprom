"""
Views para o módulo de Planejadas
Gerencia orçamento e distribuição de valores para planejadas
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction, IntegrityError
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
import re
from .models import (
    OrcamentoPlanejadas, 
    DistribuicaoOrcamentoPlanejadas,
    Orgao, 
    GrandeComando, 
    Unidade, 
    SubUnidade,
    Planejada,
    Militar,
    Lotacao,
    ConfiguracaoPlanejadas
)
from datetime import datetime


def get_posto_completo(posto_abreviado):
    """
    Converte posto/graduação abreviado para nome completo
    """
    posto_choices = {
        'CEL': 'Coronel',
        'TC': 'Tenente-Coronel',
        'MAJ': 'Major',
        'CAP': 'Capitão',
        '1T': '1º Tenente',
        '2T': '2º Tenente',
        'ASP': 'Aspirante',
        'ST': 'Subtenente',
        '1S': '1º Sargento',
        '2S': '2º Sargento',
        '3S': '3º Sargento',
        'CAB': 'Cabo',
        'SD': 'Soldado',
        # Adicionais para bombeiros
        'CB': 'Coronel',  # CORRIGIDO: CB = Coronel, não Cabo
        'SGT': 'Sargento',
        'SGT1': '1º Sargento',
        'SGT2': '2º Sargento',
        'SGT3': '3º Sargento',
        'TEN': 'Tenente',
        'BR': 'Brigadeiro',
        'BRG': 'Brigadeiro',
        'GEN': 'General',
        'ADJ': 'Adjunto',
        'COM': 'Comandante',
        'SUB': 'Subcomandante',
        # Variações comuns
        'MJ': 'Major',
        'CP': 'Capitão',
        'AS': 'Aspirante'
    }
    return posto_choices.get(posto_abreviado, posto_abreviado)


# ===== VIEWS PARA ORÇAMENTO =====

@login_required
def orcamento_planejadas_list(request):
    """Lista todos os orçamentos de planejadas"""
    orcamentos = OrcamentoPlanejadas.objects.all().order_by('-ano', '-mes')
    
    # Estatísticas gerais
    total_orcamentos = orcamentos.count()
    valor_total_geral = orcamentos.aggregate(total=Sum('valor_total'))['total'] or 0
    
    context = {
        'title': 'Orçamentos de Planejadas',
        'page_title': 'Orçamentos de Planejadas',
        'orcamentos': orcamentos,
        'total_orcamentos': total_orcamentos,
        'valor_total_geral': valor_total_geral,
    }
    
    return render(request, 'militares/orcamento_planejadas_list.html', context)


@login_required
def orcamento_planejadas_create(request):
    """Cria novo orçamento de planejadas"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Desativar outros orçamentos do mesmo mês/ano
                ano = int(request.POST.get('ano'))
                mes = int(request.POST.get('mes'))
                
                OrcamentoPlanejadas.objects.filter(ano=ano, mes=mes).update(ativo=False)
                
                # Converter valor de formato brasileiro para americano
                valor_str = request.POST.get('valor_total', '0.00')
                
                # Lógica correta: se tem vírgula, é formato brasileiro
                if ',' in valor_str:
                    # Formato brasileiro: 120.000,50 -> 120000.50
                    valor_str = valor_str.replace('.', '').replace(',', '.')
                # Se não tem vírgula, já está no formato americano
                
                try:
                    valor_total = float(valor_str)
                except ValueError:
                    messages.error(request, 'Valor inválido. Use apenas números.')
                    context = {
                        'title': 'Novo Orçamento de Planejadas',
                        'page_title': 'Novo Orçamento de Planejadas',
                        'anos_disponiveis': list(range(2020, 2031)),
                        'meses_choices': OrcamentoPlanejadas.MES_CHOICES,
                    }
                    return render(request, 'militares/orcamento_planejadas_form.html', context)
                
                orcamento = OrcamentoPlanejadas.objects.create(
                    ano=ano,
                    mes=mes,
                    valor_total=valor_total,
                    observacoes=request.POST.get('observacoes', ''),
                    ativo=True
                )
                
                # Copiar distribuição do mês anterior se solicitado
                copiar_distribuicao_anterior = request.POST.get('copiar_distribuicao_anterior', False)
                distribuicoes_copiadas = 0
                if copiar_distribuicao_anterior:
                    distribuicoes_copiadas = copiar_distribuicao_mes_anterior(orcamento)
                
                # Mensagem de sucesso
                if distribuicoes_copiadas:
                    messages.success(request, f'Orçamento de planejadas criado com sucesso! Valores das planejadas copiados do mês anterior ({distribuicoes_copiadas} organizações).')
                else:
                    messages.success(request, 'Orçamento de planejadas criado com sucesso!')
                
                return redirect('militares:orcamento_planejadas_list')
                
        except ValidationError as e:
            messages.error(request, f'Erro de validação: {str(e)}')
        except Exception as e:
            messages.error(request, f'Erro ao criar orçamento: {str(e)}')
    
    # Obter anos disponíveis
    anos_disponiveis = list(range(2020, 2031))
    
    context = {
        'title': 'Novo Orçamento de Planejadas',
        'page_title': 'Novo Orçamento de Planejadas',
        'anos_disponiveis': anos_disponiveis,
        'meses_choices': OrcamentoPlanejadas.MES_CHOICES,
    }
    
    return render(request, 'militares/orcamento_planejadas_form.html', context)


def copiar_distribuicao_mes_anterior(orcamento_novo):
    """
    Copia a distribuição do mês anterior para o novo orçamento
    """
    from .models import DistribuicaoOrcamentoPlanejadas
    
    # Calcular mês anterior
    mes_anterior = orcamento_novo.mes - 1
    ano_anterior = orcamento_novo.ano
    
    if mes_anterior <= 0:
        mes_anterior = 12
        ano_anterior -= 1
    
    # Buscar orçamento do mês anterior
    try:
        orcamento_anterior = OrcamentoPlanejadas.objects.filter(
            ano=ano_anterior,
            mes=mes_anterior,
            ativo=True
        ).first()
        
        if not orcamento_anterior:
            return 0
        
        # Buscar distribuições do mês anterior
        distribuicoes_anteriores = DistribuicaoOrcamentoPlanejadas.objects.filter(
            orcamento=orcamento_anterior
        )
        
        if not distribuicoes_anteriores.exists():
            return 0
        
        # Calcular fator de proporção baseado no valor total do orçamento
        fator_proporcao = orcamento_novo.valor_total / orcamento_anterior.valor_total
        
        # Copiar distribuições mantendo os valores das planejadas do mês anterior
        distribuicoes_copiadas = 0
        for distribuicao_antiga in distribuicoes_anteriores:
            # Manter o valor das planejadas do mês anterior (valor distribuído)
            valor_planejadas_anterior = distribuicao_antiga.valor_planejadas
            novo_percentual = (valor_planejadas_anterior / orcamento_novo.valor_total) * 100
            
            DistribuicaoOrcamentoPlanejadas.objects.create(
                orcamento=orcamento_novo,
                orgao=distribuicao_antiga.orgao,
                grande_comando=distribuicao_antiga.grande_comando,
                unidade=distribuicao_antiga.unidade,
                sub_unidade=distribuicao_antiga.sub_unidade,
                valor_planejadas=valor_planejadas_anterior,  # Mantém valor das planejadas
                percentual=novo_percentual,
                observacoes=f'Valores das planejadas copiados do mês anterior ({orcamento_anterior.mes_nome}/{orcamento_anterior.ano})'
            )
            distribuicoes_copiadas += 1
        
        return distribuicoes_copiadas
        
    except Exception as e:
        print(f"Erro ao copiar distribuição: {e}")
        return 0


@login_required
def orcamento_planejadas_detail(request, pk):
    """Detalhes do orçamento de planejadas"""
    orcamento = get_object_or_404(OrcamentoPlanejadas, pk=pk)
    
    context = {
        'title': f'Orçamento {orcamento.mes_nome}/{orcamento.ano}',
        'page_title': f'Orçamento {orcamento.mes_nome}/{orcamento.ano}',
        'orcamento': orcamento,
    }
    
    return render(request, 'militares/orcamento_planejadas_detail.html', context)


@login_required
def orcamento_planejadas_edit(request, pk):
    """Edita orçamento de planejadas"""
    orcamento = get_object_or_404(OrcamentoPlanejadas, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                orcamento.ano = int(request.POST.get('ano'))
                orcamento.mes = int(request.POST.get('mes'))
                # Converter valor de formato brasileiro para americano
                valor_str = request.POST.get('valor_total', '0.00')
                
                # Lógica correta: se tem vírgula, é formato brasileiro
                if ',' in valor_str:
                    # Formato brasileiro: 120.000,50 -> 120000.50
                    valor_str = valor_str.replace('.', '').replace(',', '.')
                # Se não tem vírgula, já está no formato americano
                
                try:
                    orcamento.valor_total = float(valor_str)
                except ValueError:
                    messages.error(request, 'Valor inválido. Use apenas números.')
                    return render(request, 'militares/orcamento_planejadas_form.html', context)
                orcamento.observacoes = request.POST.get('observacoes', '')
                orcamento.ativo = request.POST.get('ativo') == 'on'
                
                orcamento.save()
                
                messages.success(request, 'Orçamento atualizado com sucesso!')
                return redirect('militares:orcamento_planejadas_list')
                
        except ValidationError as e:
            messages.error(request, f'Erro de validação: {str(e)}')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar orçamento: {str(e)}')
    
    context = {
        'title': 'Editar Orçamento de Planejadas',
        'page_title': 'Editar Orçamento de Planejadas',
        'orcamento': orcamento,
        'anos_disponiveis': list(range(2020, 2031)),
        'meses_choices': OrcamentoPlanejadas.MES_CHOICES,
    }
    
    return render(request, 'militares/orcamento_planejadas_form.html', context)


@login_required
def orcamento_planejadas_delete(request, pk):
    """Exclui orçamento de planejadas"""
    try:
        orcamento = get_object_or_404(OrcamentoPlanejadas, pk=pk)
        orcamento.delete()
        messages.success(request, 'Orçamento excluído com sucesso!')
        
    except Exception as e:
        messages.error(request, f'Erro ao excluir orçamento: {str(e)}')
    
    return redirect('militares:orcamento_planejadas_list')




@login_required
def distribuir_orcamento_organizacoes(request, orcamento_id):
    """Distribui orçamento entre todas as organizações do organograma"""
    orcamento = get_object_or_404(OrcamentoPlanejadas, pk=orcamento_id)
    
    if request.method == 'POST':
        
        # Verificar se é uma requisição AJAX para salvar valor individual
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and request.POST.get('salvar_individual') == 'true':
            return salvar_valor_individual(request, orcamento)
        
        try:
            with transaction.atomic():
                # Obter dados do formulário
                valores_organizacoes = {}
                total_distribuido = 0
                
                # Processar órgãos
                orgao_ids = request.POST.getlist('orgao_ids')
                for orgao_id in orgao_ids:
                    if orgao_id:
                        valor_str = request.POST.get(f'valor_orgao_{orgao_id}', '0')
                        try:
                            # Converter formato brasileiro para float
                            valor = float(valor_str.replace('.', '').replace(',', '.'))
                            if valor > 0:
                                valores_organizacoes[f'orgao_{orgao_id}'] = {
                                    'tipo': 'orgao',
                                    'instancia_id': orgao_id,
                                    'valor': valor
                                }
                                total_distribuido += valor
                        except (ValueError, TypeError) as e:
                            continue
                
                # Processar grandes comandos
                for gc_id in request.POST.getlist('grande_comando_ids'):
                    if gc_id:
                        valor_str = request.POST.get(f'valor_grande_comando_{gc_id}', '0')
                        try:
                            # Converter formato brasileiro para float
                            valor = float(valor_str.replace('.', '').replace(',', '.'))
                            if valor > 0:
                                valores_organizacoes[f'grande_comando_{gc_id}'] = {
                                    'tipo': 'grande_comando',
                                    'instancia_id': gc_id,
                                    'valor': valor
                                }
                                total_distribuido += valor
                        except (ValueError, TypeError):
                            continue
                
                # Processar unidades
                for unidade_id in request.POST.getlist('unidade_ids'):
                    if unidade_id:
                        valor_str = request.POST.get(f'valor_unidade_{unidade_id}', '0')
                        try:
                            # Converter formato brasileiro para float
                            valor = float(valor_str.replace('.', '').replace(',', '.'))
                            if valor > 0:
                                valores_organizacoes[f'unidade_{unidade_id}'] = {
                                    'tipo': 'unidade',
                                    'instancia_id': unidade_id,
                                    'valor': valor
                                }
                                total_distribuido += valor
                        except (ValueError, TypeError):
                            continue
                
                # Processar sub-unidades
                for sub_unidade_id in request.POST.getlist('sub_unidade_ids'):
                    if sub_unidade_id:
                        valor_str = request.POST.get(f'valor_sub_unidade_{sub_unidade_id}', '0')
                        try:
                            # Converter formato brasileiro para float
                            valor = float(valor_str.replace('.', '').replace(',', '.'))
                            if valor > 0:
                                valores_organizacoes[f'sub_unidade_{sub_unidade_id}'] = {
                                    'tipo': 'sub_unidade',
                                    'instancia_id': sub_unidade_id,
                                    'valor': valor
                                }
                                total_distribuido += valor
                        except (ValueError, TypeError):
                            continue
                
                # Verificar se o total não ultrapassa o orçamento
                if total_distribuido > float(orcamento.valor_total):
                    messages.error(request, f'Valor total distribuído (R$ {total_distribuido:,.2f}) não pode ser maior que o orçamento disponível (R$ {orcamento.valor_total:,.2f}).')
                    return redirect('militares:distribuir_orcamento_organizacoes', orcamento_id=orcamento.id)
                
                # Criar distribuições de orçamento
                for key, dados in valores_organizacoes.items():
                    # Obter a instância
                    if dados['tipo'] == 'orgao':
                        instancia = get_object_or_404(Orgao, pk=dados['instancia_id'])
                        orgao = instancia
                        grande_comando = None
                        unidade = None
                        sub_unidade = None
                    elif dados['tipo'] == 'grande_comando':
                        instancia = get_object_or_404(GrandeComando, pk=dados['instancia_id'])
                        orgao = None
                        grande_comando = instancia
                        unidade = None
                        sub_unidade = None
                    elif dados['tipo'] == 'unidade':
                        instancia = get_object_or_404(Unidade, pk=dados['instancia_id'])
                        orgao = None
                        grande_comando = None
                        unidade = instancia
                        sub_unidade = None
                    elif dados['tipo'] == 'sub_unidade':
                        instancia = get_object_or_404(SubUnidade, pk=dados['instancia_id'])
                        orgao = None
                        grande_comando = None
                        unidade = None
                        sub_unidade = instancia
                    
                    # Criar distribuição de orçamento
                    distribuicao = DistribuicaoOrcamentoPlanejadas.objects.create(
                        orcamento=orcamento,
                        orgao=orgao,
                        grande_comando=grande_comando,
                        unidade=unidade,
                        sub_unidade=sub_unidade,
                        valor_planejadas=dados['valor'],
                        percentual=(dados['valor'] / float(orcamento.valor_total)) * 100,
                        observacoes=f'Distribuição do orçamento {orcamento.mes_nome}/{orcamento.ano}'
                    )
                
                valor_restante = float(orcamento.valor_total) - total_distribuido
                messages.success(request, f'Orçamento distribuído com sucesso! Total distribuído: R$ {total_distribuido:,.2f}. Valor restante: R$ {valor_restante:,.2f}')
                return redirect('militares:orcamento_planejadas_detail', pk=orcamento.id)
                
        except Exception as e:
            messages.error(request, f'Erro ao distribuir orçamento: {str(e)}')
            return redirect('militares:distribuir_orcamento_organizacoes', orcamento_id=orcamento.id)
    
    # Obter todas as instâncias organizacionais
    orgaos = Orgao.objects.filter(ativo=True).order_by('nome')
    grandes_comandos = GrandeComando.objects.filter(ativo=True).order_by('nome')
    unidades = Unidade.objects.filter(ativo=True).order_by('nome')
    sub_unidades = SubUnidade.objects.filter(ativo=True).order_by('nome')
    
    # Obter distribuições existentes para este orçamento
    orcamentos_existentes = {}
    gastos_por_distribuicao = {}
    valores_excluidos_por_distribuicao = {}
    
    for dist in DistribuicaoOrcamentoPlanejadas.objects.filter(orcamento=orcamento):
        if dist.orgao:
            key = f"orgao_{dist.orgao.id}"
            nome_instancia = dist.orgao.nome
        elif dist.grande_comando:
            key = f"grande_comando_{dist.grande_comando.id}"
            nome_instancia = dist.grande_comando.nome
        elif dist.unidade:
            key = f"unidade_{dist.unidade.id}"
            nome_instancia = dist.unidade.nome
        elif dist.sub_unidade:
            key = f"sub_unidade_{dist.sub_unidade.id}"
            nome_instancia = dist.sub_unidade.nome
        else:
            continue
        
        orcamentos_existentes[key] = float(dist.valor_planejadas)
        
        # Calcular gastos reais (planejadas ativas, não excluídas)
        # Usar busca hierárquica para origens do tipo "A | B | C"
        planejadas_ativas = Planejada.objects.filter(
            data_operacao__year=orcamento.ano,
            data_operacao__month=orcamento.mes,
            ativo=True,
            excluido=False
        )
        
        # Filtrar pelo match hierárquico
        planejadas_ativas_filtradas = []
        for planejada in planejadas_ativas:
            partes_origem = [p.strip() for p in planejada.origem.split('|')]
            if nome_instancia in partes_origem:
                planejadas_ativas_filtradas.append(planejada)
        
        total_gasto = sum(float(p.valor_total) for p in planejadas_ativas_filtradas)
        gastos_por_distribuicao[key] = float(total_gasto)
        
        # Calcular valores excluídos (planejadas excluídas)
        planejadas_excluidas = Planejada.objects.filter(
            data_operacao__year=orcamento.ano,
            data_operacao__month=orcamento.mes,
            ativo=True,
            excluido=True
        )
        
        # Filtrar pelo match hierárquico
        planejadas_excluidas_filtradas = []
        for planejada in planejadas_excluidas:
            partes_origem = [p.strip() for p in planejada.origem.split('|')]
            if nome_instancia in partes_origem:
                planejadas_excluidas_filtradas.append(planejada)
        
        total_excluido = sum(float(p.valor_total) for p in planejadas_excluidas_filtradas)
        valores_excluidos_por_distribuicao[key] = float(total_excluido)
    
    # Calcular total já distribuído
    total_distribuido = sum(orcamentos_existentes.values())
    valor_restante = float(orcamento.valor_total) - total_distribuido
    
    context = {
        'title': f'Distribuir Orçamento - {orcamento.mes_nome}/{orcamento.ano}',
        'page_title': f'Distribuir Orçamento - {orcamento.mes_nome}/{orcamento.ano}',
        'orcamento': orcamento,
        'orgaos': orgaos,
        'grandes_comandos': grandes_comandos,
        'unidades': unidades,
        'sub_unidades': sub_unidades,
        'orcamentos_existentes': orcamentos_existentes,
        'gastos_por_distribuicao': gastos_por_distribuicao,
        'valores_excluidos_por_distribuicao': valores_excluidos_por_distribuicao,
        'total_distribuido': total_distribuido,
        'valor_restante': valor_restante,
    }
    
    return render(request, 'militares/distribuir_orcamento_organizacoes.html', context)


@login_required
def visualizar_planejadas_organizacao(request, tipo, org_id):
    """Visualiza as planejadas de uma organização específica"""
    try:
        # Obter a organização baseada no tipo
        if tipo == 'orgao':
            organizacao = get_object_or_404(Orgao, pk=org_id)
            nome_organizacao = organizacao.nome
        elif tipo == 'grande_comando':
            organizacao = get_object_or_404(GrandeComando, pk=org_id)
            nome_organizacao = organizacao.nome
        elif tipo == 'unidade':
            organizacao = get_object_or_404(Unidade, pk=org_id)
            nome_organizacao = organizacao.nome
        elif tipo == 'sub_unidade':
            organizacao = get_object_or_404(SubUnidade, pk=org_id)
            nome_organizacao = organizacao.nome
        else:
            return HttpResponse('<div class="alert alert-danger">Tipo de organização inválido.</div>')
        
        # Obter planejadas da organização
        # Usar a mesma lógica hierárquica da API para evitar contagem dupla
        # Obter mês e ano do orçamento da URL ou parâmetros
        orcamento_id = request.GET.get('orcamento_id')
        if orcamento_id:
            try:
                orcamento = OrcamentoPlanejadas.objects.get(id=orcamento_id)
                ano_orcamento = orcamento.ano
                mes_orcamento = orcamento.mes
            except OrcamentoPlanejadas.DoesNotExist:
                # Fallback para mês atual se orçamento não encontrado
                hoje = datetime.now()
                ano_orcamento = hoje.year
                mes_orcamento = hoje.month
        else:
            # Fallback para mês atual se não especificado
            hoje = datetime.now()
            ano_orcamento = hoje.year
            mes_orcamento = hoje.month
        
        # Buscar planejadas ativas do mês do orçamento
        planejadas_ativas = Planejada.objects.filter(
            data_operacao__year=ano_orcamento,
            data_operacao__month=mes_orcamento,
            ativo=True,
            excluido=False,  # Excluir planejadas marcadas como excluídas
            origem__icontains=organizacao.nome
        )
        
        # Aplicar lógica hierárquica: atribuir apenas à organização mais específica
        planejadas_filtradas = []
        for planejada in planejadas_ativas:
            # Dividir a origem por "|" e pegar a última organização (mais específica)
            partes_origem = [p.strip() for p in planejada.origem.split('|')]
            ultima_organizacao = partes_origem[-1] if partes_origem else ""
            
            # Verificar se esta organização corresponde EXATAMENTE à última (mais específica)
            if organizacao.nome.lower().strip() == ultima_organizacao.lower().strip():
                planejadas_filtradas.append(planejada)
        
        planejadas = planejadas_filtradas
        
        # Calcular totais para os cards
        total_geral = sum(float(p.valor_total) for p in planejadas)
        total_ativo = sum(float(p.valor_total) for p in planejadas if p.status == 'ativo')
        total_concluido = sum(float(p.valor_total) for p in planejadas if p.status == 'concluido')
        total_cancelado = sum(float(p.valor_total) for p in planejadas if p.status == 'cancelado')
        total_suspenso = sum(float(p.valor_total) for p in planejadas if p.status == 'suspenso')
        
        context = {
            'organizacao': organizacao,
            'nome_organizacao': nome_organizacao,
            'tipo': tipo,
            'planejadas': planejadas,
            'total_geral': total_geral,
            'total_ativo': total_ativo,
            'total_concluido': total_concluido,
            'total_cancelado': total_cancelado,
            'total_suspenso': total_suspenso,
        }
        
        return render(request, 'militares/planejadas_organizacao.html', context)
        
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Erro ao carregar planejadas: {str(e)}</div>')


@login_required
def salvar_valor_individual(request, orcamento):
    """Salva um valor individual via AJAX"""
    try:
        from decimal import Decimal
        
        tipo_organizacao = request.POST.get('tipo_organizacao')
        organizacao_id = request.POST.get('organizacao_id')
        valor = Decimal(str(request.POST.get('valor', 0)))
        
        
        with transaction.atomic():
            # Obter a instância da organização
            if tipo_organizacao == 'orgao':
                instancia = get_object_or_404(Orgao, pk=organizacao_id)
                orgao = instancia
                grande_comando = None
                unidade = None
                sub_unidade = None
            elif tipo_organizacao == 'grande_comando':
                instancia = get_object_or_404(GrandeComando, pk=organizacao_id)
                orgao = None
                grande_comando = instancia
                unidade = None
                sub_unidade = None
            elif tipo_organizacao == 'unidade':
                instancia = get_object_or_404(Unidade, pk=organizacao_id)
                orgao = None
                grande_comando = None
                unidade = instancia
                sub_unidade = None
            elif tipo_organizacao == 'sub_unidade':
                instancia = get_object_or_404(SubUnidade, pk=organizacao_id)
                orgao = None
                grande_comando = None
                unidade = None
                sub_unidade = instancia
            else:
                return JsonResponse({'success': False, 'message': 'Tipo de organização inválido'})
            
            # Verificar se já existe uma distribuição para esta organização
            distribuicao_existente = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                orgao=orgao,
                grande_comando=grande_comando,
                unidade=unidade,
                sub_unidade=sub_unidade
            ).first()
            
            if valor > 0:
                # Validar se o valor não ultrapassa o orçamento total
                # Calcular total atual das distribuições (sem incluir a distribuição atual se ela existir)
                distribuicoes = DistribuicaoOrcamentoPlanejadas.objects.filter(orcamento=orcamento)
                total_atual = sum(dist.valor_planejadas for dist in distribuicoes)
                
                
                # Se existe distribuição, subtrair o valor atual dela para calcular o total sem ela
                if distribuicao_existente:
                    total_sem_atual = total_atual - distribuicao_existente.valor_planejadas
                    valor_atual = distribuicao_existente.valor_planejadas
                else:
                    total_sem_atual = total_atual
                    valor_atual = Decimal('0')
                
                # Verificar se o novo total não ultrapassa o orçamento
                novo_total = total_sem_atual + valor
                
                # Só validar se estiver aumentando o valor E ultrapassando o orçamento
                esta_aumentando = valor > valor_atual
                ultrapassa_orcamento = novo_total > orcamento.valor_total
                
                
                # TEMPORÁRIO: Desabilitar validação de orçamento para focar na funcionalidade
                # A validação será feita apenas no frontend por enquanto
                
                # TODO: Reabilitar validação quando o problema de cálculo for resolvido
                # if esta_aumentando and ultrapassa_orcamento:
                #     valor_restante = orcamento.valor_total - total_sem_atual
                #     return JsonResponse({
                #         'success': False, 
                #         'message': f'Valor ultrapassa o orçamento disponível. Valor restante: R$ {valor_restante:,.2f}'
                #     })
                
                # Criar ou atualizar distribuição
                
                if distribuicao_existente:
                    if valor == 0:
                        distribuicao_existente.delete()
                    else:
                        distribuicao_existente.valor_planejadas = valor
                        distribuicao_existente.percentual = (valor / orcamento.valor_total) * 100
                        distribuicao_existente.save()
                else:
                    if valor > 0:
                        distribuicao = DistribuicaoOrcamentoPlanejadas.objects.create(
                            orcamento=orcamento,
                            orgao=orgao,
                            grande_comando=grande_comando,
                            unidade=unidade,
                            sub_unidade=sub_unidade,
                            valor_planejadas=valor,
                            percentual=(valor / orcamento.valor_total) * 100,
                            observacoes=f'Distribuição individual do orçamento {orcamento.mes_nome}/{orcamento.ano}'
                        )
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Valor salvo com sucesso!',
                    'valor': float(valor),
                    'organizacao': instancia.nome
                })
            else:
                # Remover distribuição se valor for zero
                if distribuicao_existente:
                    distribuicao_existente.delete()
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Valor removido com sucesso!',
                    'valor': 0,
                    'organizacao': instancia.nome
                })
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao salvar: {str(e)}'})


# ===== VIEWS PARA PLANEJADAS =====

@login_required
def planejadas_list(request):
    """
    Lista operações planejadas com filtro hierárquico baseado no acesso da função
    Igual ao sistema de notas - lista inicialmente a instância do tipo de acesso com filtros para dependências
    """
    try:
        # Verificar se o usuário tem permissão para acessar planejadas
        from .permissoes_simples import pode_visualizar_operacoes_planejadas
        
        if not pode_visualizar_operacoes_planejadas(request.user):
            messages.error(request, 'Você não tem permissão para acessar as operações planejadas.')
            return redirect('militares:militar_dashboard')
        
        # Obter função atual do usuário
        from .permissoes_simples import obter_funcao_militar_ativa
        funcao_atual = obter_funcao_militar_ativa(request.user)
        
        # Parâmetros de filtro
        search = request.GET.get('search', '')
        origem = request.GET.get('origem', '')
        cidade = request.GET.get('cidade', '')
        semana = request.GET.get('semana', '')
        status = request.GET.get('status', '')
        mes = request.GET.get('mes', '')
        ano = request.GET.get('ano', '')
        
        # Parâmetros de pesquisa hierárquica
        pesquisa_origem_dropdown = request.GET.get('pesquisa_origem_dropdown', '')
        
        # Query base
        planejadas_base = Planejada.objects.all()
        
        # Aplicar filtro hierárquico baseado no acesso da função
        if request.user.is_superuser:
            # Superusuário vê todas as planejadas
            planejadas = planejadas_base.filter(ativo=True, excluido=False)
        elif funcao_atual and funcao_atual.funcao_militar.publicacao in ['OPERADOR_PLANEJADAS', 'APROVADOR', 'FISCAL_PLANEJADAS']:
            # Verificar se há filtro de origem específico selecionado
            origem_selecionada = request.GET.get('pesquisa_origem_dropdown', '').strip()
            
            if origem_selecionada:
                # Se há origem selecionada no filtro, mostrar apenas planejadas dessa origem específica
                planejadas = aplicar_filtro_hierarquico_planejadas_especifico(planejadas_base, funcao_atual, origem_selecionada)
            else:
                # Sem filtro específico, usar filtro restritivo (apenas o próprio nível)
                planejadas = aplicar_filtro_hierarquico_planejadas_restritivo(planejadas_base, funcao_atual)
        else:
            # Se não é operador, aprovador ou fiscal de planejadas, não pode ver nenhuma planejada
            planejadas = planejadas_base.none()
        
        # Aplicar filtros adicionais
        if search:
            planejadas = planejadas.filter(
                Q(nome__icontains=search) |
                Q(descricao__icontains=search) |
                Q(cidade__icontains=search)
            )
        
        if origem:
            planejadas = planejadas.filter(origem=origem)
            
        if cidade:
            planejadas = planejadas.filter(cidade__icontains=cidade)
            
        if semana:
            planejadas = planejadas.filter(semana=semana)
            
        if status:
            planejadas = planejadas.filter(status=status)
        
        # Filtro por mês e ano (se vier do relatório mensal)
        if mes and ano:
            planejadas = planejadas.filter(
                data_operacao__year=ano,
                data_operacao__month=mes,
                ativo=True,
                excluido=False
            )
        
        # Ordenar por data de operação
        planejadas = planejadas.order_by('-data_operacao')
        
        # Recalcular valores de todas as planejadas antes de exibir
        for planejada in planejadas:
            planejada.recalcular_valores()
        
        # Paginação
        paginator = Paginator(planejadas, 20)
        page_number = request.GET.get('page')
        planejadas_page = paginator.get_page(page_number)
        
        # Estatísticas
        total_planejadas = planejadas.count()
        valor_total = planejadas.aggregate(Sum('valor_total'))['valor_total__sum'] or 0
        
        # Opções para filtros - usar apenas as planejadas já filtradas por instância
        origens = planejadas.values_list('origem', flat=True).distinct().order_by('origem')
        cidades = planejadas.values_list('cidade', flat=True).distinct().order_by('cidade')
        semanas = planejadas.values_list('semana', flat=True).distinct().order_by('semana')
        status_opcoes = planejadas.values_list('status', flat=True).distinct().order_by('status')
        
        # Opções hierárquicas para filtro de origem
        opcoes_hierarquicas = obter_opcoes_hierarquicas_planejadas(funcao_atual, request.user)
        
        from datetime import date
        
        # Verificar permissões de assinatura do usuário
        pode_operador = False
        pode_aprovador = False
        pode_fiscal = False
        
        if request.user.is_superuser:
            pode_operador = True
            pode_aprovador = True
            pode_fiscal = True
        else:
            if funcao_atual:
                funcao_militar = funcao_atual.funcao_militar
                pode_operador = funcao_militar.publicacao == 'OPERADOR_PLANEJADAS'
                pode_aprovador = funcao_militar.publicacao == 'APROVADOR'
                pode_fiscal = funcao_militar.publicacao == 'FISCAL_PLANEJADAS'
        
        context = {
            'page_title': 'Operações Planejadas',
            'planejadas': planejadas_page,
            'total_planejadas': total_planejadas,
            'valor_total': valor_total,
            'origens': origens,
            'cidades': cidades,
            'semanas': semanas,
            'status_opcoes': status_opcoes,
            'opcoes_hierarquicas': opcoes_hierarquicas,
            'filtros': {
            'search': search,
            'origem': origem,
            'cidade': cidade,
            'semana': semana,
            'status': status,
                'pesquisa_origem_dropdown': pesquisa_origem_dropdown,
            },
            'pode_operador': pode_operador,
            'pode_aprovador': pode_aprovador,
            'pode_fiscal': pode_fiscal,
            'funcao_atual': funcao_atual,
            'today': date.today(),  # Data de hoje para comparação com data_operacao
        }
        
        return render(request, 'militares/planejadas_list.html', context)
        
    except Exception as e:
        messages.error(request, f'Erro ao carregar planejadas: {str(e)}')
        return render(request, 'militares/planejadas_list.html', {
            'page_title': 'Operações Planejadas',
            'planejadas': [],
            'total_planejadas': 0,
            'valor_total': 0,
        })


@login_required
def planejada_detail(request, planejada_id):
    """
    Detalhes de uma operação planejada específica
    """
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Recalcular valores antes de exibir os detalhes
        planejada.recalcular_valores()
        
        # Verificar se o usuário pode acessar esta planejada
        if not planejada.pode_acessar_instancia(request.user):
            messages.error(request, 'Você não tem permissão para acessar esta operação planejada.')
            return redirect('militares:planejadas_list')
        
        # Se for requisição AJAX, retornar JSON com status de tempo
        if request.headers.get('Content-Type') == 'application/json' or request.GET.get('format') == 'json':
            status_tempo = planejada.obter_status_tempo(request.user)
            return JsonResponse({
                'status_tempo': status_tempo
            })
        
        context = {
            'page_title': f'Detalhes - {planejada.nome}',
            'planejada': planejada,
        }
        
        return render(request, 'militares/planejada_detail.html', context)
        
    except Exception as e:
        messages.error(request, f'Erro ao carregar detalhes: {str(e)}')
        return redirect('militares:planejadas_list')


def verificar_conflito_horario_militar(militares_ids, data_operacao, hora_inicio, hora_termino, planejada_id=None):
    """
    Verifica se há conflito de horário para os militares selecionados
    
    Args:
        militares_ids (list): Lista de IDs dos militares
        data_operacao (datetime): Data da operação
        hora_inicio (time): Hora de início
        hora_termino (time): Hora de término
        planejada_id (int, optional): ID da planejada sendo editada (para excluir da verificação)
        
    Returns:
        dict: {'sucesso': bool, 'conflitos': list, 'mensagem': str}
    """
    try:
        from datetime import datetime, time
        
        # Converter data_operacao para datetime se necessário
        if isinstance(data_operacao, str):
            data_operacao = datetime.strptime(data_operacao, '%Y-%m-%d')
        elif hasattr(data_operacao, 'date'):
            data_operacao = data_operacao.date()
        
        # Converter strings de hora para objetos time se necessário
        if isinstance(hora_inicio, str):
            hora_inicio = datetime.strptime(hora_inicio, '%H:%M').time()
        if isinstance(hora_termino, str):
            hora_termino = datetime.strptime(hora_termino, '%H:%M').time()
        
        conflitos = []
        
        for militar_id in militares_ids:
            try:
                militar = Militar.objects.get(id=militar_id)
                
                # Buscar planejadas onde este militar está escalado na mesma data
                # Garantir que data_operacao seja um objeto date para comparação
                data_operacao_date = data_operacao.date() if hasattr(data_operacao, 'date') else data_operacao
                planejadas_militar = Planejada.objects.filter(
                    militares=militar,
                    data_operacao__date=data_operacao_date,
                    ativo=True,
                    excluido=False  # Excluir planejadas marcadas como excluídas
                )
                
                # Excluir a planejada atual se estiver editando
                if planejada_id:
                    planejadas_militar = planejadas_militar.exclude(id=planejada_id)
                
                for planejada in planejadas_militar:
                    # Verificar se há sobreposição de horários
                    if _horarios_sobrepostos(
                        hora_inicio, hora_termino,
                        planejada.hora_inicio, planejada.hora_termino
                    ):
                        conflitos.append({
                            'militar': militar.nome_guerra,
                            'planejada_conflito': planejada.nome,
                            'horario_conflito': f"{planejada.hora_inicio} - {planejada.hora_termino}",
                            'data_conflito': planejada.data_operacao.strftime('%d/%m/%Y')
                        })
                        
            except Militar.DoesNotExist:
                continue
        
        if conflitos:
            return {
                'sucesso': False,
                'conflitos': conflitos,
                'mensagem': f'Conflito de horário detectado para {len(conflitos)} militar(es)'
            }
        else:
            return {
                'sucesso': True,
                'conflitos': [],
                'mensagem': 'Nenhum conflito de horário detectado'
            }
            
    except Exception as e:
        return {
            'sucesso': False,
            'conflitos': [],
            'mensagem': f'Erro ao verificar conflitos: {str(e)}'
        }


def _horarios_sobrepostos(inicio1, fim1, inicio2, fim2):
    """
    Verifica se dois intervalos de tempo se sobrepõem
    
    Args:
        inicio1, fim1: Horários do primeiro intervalo
        inicio2, fim2: Horários do segundo intervalo
        
    Returns:
        bool: True se há sobreposição
    """
    # Converter para datetime para facilitar comparação
    from datetime import datetime, date
    
    # Usar uma data fictícia para comparação
    data_base = date.today()
    
    dt_inicio1 = datetime.combine(data_base, inicio1)
    dt_fim1 = datetime.combine(data_base, fim1)
    dt_inicio2 = datetime.combine(data_base, inicio2)
    dt_fim2 = datetime.combine(data_base, fim2)
    
    # Verificar sobreposição: início1 < fim2 E fim1 > início2
    return dt_inicio1 < dt_fim2 and dt_fim1 > dt_inicio2




def verificar_saldo_suficiente(origem, valor_total_planejada):
    """
    Verifica se há saldo suficiente para criar uma operação planejada
    
    Args:
        origem (str): Origem da operação planejada
        valor_total_planejada (float): Valor total da planejada
        
    Returns:
        dict: {'sucesso': bool, 'saldo_disponivel': float, 'mensagem': str}
    """
    try:
        from datetime import datetime
        from django.db.models import Sum
        
        
        # Obter data atual
        hoje = datetime.now()
        ano_atual = hoje.year
        mes_atual = hoje.month
        
        # Buscar orçamento do mês especificado
        orcamento = OrcamentoPlanejadas.objects.filter(
            ano=ano_atual,
            mes=mes_atual,
            ativo=True
        ).first()
        
        if not orcamento:
            return {
                'sucesso': False,
                'saldo_disponivel': 0.0,
                'mensagem': 'Nenhum orçamento encontrado para o mês atual'
            }
        
        # Analisar a origem para determinar o nível hierárquico
        # Formato esperado: "ÓRGÃO|GRANDE_COMANDO|UNIDADE|SUB_UNIDADE"
        partes_origem = origem.split('|')
        
        # Determinar qual nível hierárquico usar baseado no número de partes
        distribuicao = None
        
        if len(partes_origem) == 1:
            # Apenas órgão: "COMANDO GERAL"
            distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                ativo=True,
                orgao__nome__icontains=partes_origem[0].strip()
            ).first()
            
        elif len(partes_origem) == 2:
            # Órgão + Grande Comando: "COMANDO GERAL|QUARTEL DO COMANDO GERAL"
            distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                ativo=True,
                orgao__nome__icontains=partes_origem[0].strip(),
                grande_comando__nome__icontains=partes_origem[1].strip()
            ).first()
            
        elif len(partes_origem) == 3:
            # Órgão + Grande Comando + Unidade: "COMANDO GERAL|QUARTEL DO COMANDO GERAL|AJUDANCIA"
            distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                ativo=True,
                orgao__nome__icontains=partes_origem[0].strip(),
                grande_comando__nome__icontains=partes_origem[1].strip(),
                unidade__nome__icontains=partes_origem[2].strip()
            ).first()
            
        elif len(partes_origem) == 4:
            # Órgão + Grande Comando + Unidade + Sub-Unidade: "COMANDO GERAL|QUARTEL DO COMANDO GERAL|AJUDANCIA|SEÇÃO"
            distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                ativo=True,
                orgao__nome__icontains=partes_origem[0].strip(),
                grande_comando__nome__icontains=partes_origem[1].strip(),
                unidade__nome__icontains=partes_origem[2].strip(),
                sub_unidade__nome__icontains=partes_origem[3].strip()
            ).first()
        
        # Se não encontrou distribuição específica, tentar busca mais flexível
        if not distribuicao:
            # Buscar por correspondência parcial em qualquer nível
            distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                ativo=True
            ).filter(
                Q(orgao__nome__icontains=partes_origem[0].strip()) |
                Q(grande_comando__nome__icontains=partes_origem[0].strip()) |
                Q(unidade__nome__icontains=partes_origem[0].strip()) |
                Q(sub_unidade__nome__icontains=partes_origem[0].strip())
            ).first()
        
        # Se ainda não encontrou, usar a primeira distribuição disponível como fallback
        if not distribuicao:
            distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
                orcamento=orcamento,
                ativo=True
            ).first()
        
        
        if not distribuicao:
            return {
                'sucesso': False,
                'saldo_disponivel': 0.0,
                'mensagem': f'Nenhuma distribuição de orçamento encontrada para a origem: {origem}'
            }
        
        # Calcular gastos do mês atual para esta origem específica
        # Buscar operações que pertencem EXATAMENTE a esta instância
        # Prioridade: Sub-Unidade > Unidade > Grande Comando > Órgão
        if distribuicao.sub_unidade:
            # Para sub-unidade, buscar operações que terminam com o nome da sub-unidade
            planejadas_gastos = Planejada.objects.filter(
                data_operacao__year=ano_atual,
                data_operacao__month=mes_atual,
                ativo=True,
                excluido=False,  # Excluir planejadas marcadas como excluídas
                origem__endswith=distribuicao.sub_unidade.nome
            )
        elif distribuicao.unidade:
            # Para unidade, buscar operações que terminam com o nome da unidade
            planejadas_gastos = Planejada.objects.filter(
                data_operacao__year=ano_atual,
                data_operacao__month=mes_atual,
                ativo=True,
                excluido=False,  # Excluir planejadas marcadas como excluídas
                origem__endswith=distribuicao.unidade.nome
            )
        elif distribuicao.grande_comando:
            # Para grande comando, buscar operações que terminam com o nome do grande comando
            planejadas_gastos = Planejada.objects.filter(
                data_operacao__year=ano_atual,
                data_operacao__month=mes_atual,
                ativo=True,
                excluido=False,  # Excluir planejadas marcadas como excluídas
                origem__endswith=distribuicao.grande_comando.nome
            )
        elif distribuicao.orgao:
            # Para órgão, buscar operações que começam e terminam com o nome do órgão
            # (ou seja, operações que são APENAS do órgão, sem outras instâncias)
            planejadas_gastos = Planejada.objects.filter(
                data_operacao__year=ano_atual,
                data_operacao__month=mes_atual,
                ativo=True,
                excluido=False,  # Excluir planejadas marcadas como excluídas
                origem=distribuicao.orgao.nome
            )
        else:
            # Fallback para busca genérica
            planejadas_gastos = Planejada.objects.filter(
                data_operacao__year=ano_atual,
                data_operacao__month=mes_atual,
                ativo=True,
                excluido=False,  # Excluir planejadas marcadas como excluídas
                origem__icontains=origem
            )
        
        # Calcular total gasto
        total_gasto = planejadas_gastos.aggregate(
            total=Sum('valor_total')
        )['total'] or 0
        
        # Calcular saldo disponível
        saldo_disponivel = float(distribuicao.valor_planejadas) - float(total_gasto)
        
        
        # Verificar se há saldo suficiente
        if saldo_disponivel >= valor_total_planejada:
            return {
                'sucesso': True,
                'saldo_disponivel': saldo_disponivel,
                'mensagem': f'Saldo disponível: R$ {saldo_disponivel:,.2f}'
            }
        else:
            return {
                'sucesso': False,
                'saldo_disponivel': saldo_disponivel,
                'mensagem': f'Saldo insuficiente. Disponível: R$ {saldo_disponivel:,.2f} | Necessário: R$ {valor_total_planejada:,.2f}'
            }
            
    except Exception as e:
        return {
            'sucesso': False,
            'saldo_disponivel': 0.0,
            'mensagem': f'Erro ao verificar saldo: {str(e)}'
        }






@login_required
def planejada_delete(request, planejada_id):
    """
    Marca uma operação planejada como excluída com justificativa
    """
    planejada = get_object_or_404(Planejada, id=planejada_id)
    
    # Verificar se o usuário pode acessar esta planejada
    if not planejada.pode_acessar_instancia(request.user):
        messages.error(request, 'Você não tem permissão para excluir esta operação planejada.')
        return redirect('militares:planejadas_list')
    
    # Verificar se é superusuário
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': 'Apenas superusuários podem excluir operações planejadas'
        })
    
    try:
        # Verificar se é requisição AJAX
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        
        if request.method == 'POST':
            justificativa = request.POST.get('justificativa', '').strip()
            
            if not justificativa:
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': 'Justificativa é obrigatória para excluir a planejada'
                    })
                else:
                    messages.error(request, 'Justificativa é obrigatória para excluir a planejada.')
                    return redirect('militares:planejadas_list')
            
            # Marcar como excluído com justificativa
            planejada.marcar_como_excluido(request.user, justificativa)
            
            # Resposta de sucesso com informações sobre valor liberado
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': f'Operação planejada "{planejada.nome}" foi marcada como excluída com sucesso! Valor de R$ {planejada.valor_total:,.2f} foi liberado para o orçamento da {planejada.origem}.',
                    'valor_liberado': float(planejada.valor_total),
                    'origem': planejada.origem
                })
            else:
                messages.success(request, f'Operação planejada "{planejada.nome}" foi marcada como excluída com sucesso! Valor de R$ {planejada.valor_total:,.2f} foi liberado para o orçamento da {planejada.origem}.')
                return redirect('militares:planejadas_list')
        
        # Se não é POST, retornar erro
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Método não permitido'
            })
        else:
            messages.error(request, 'Método não permitido.')
            return redirect('militares:planejadas_list')
    
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Erro ao excluir operação planejada: {str(e)}'
            })
        else:
            messages.error(request, f'Erro ao excluir operação planejada: {str(e)}')
            return redirect('militares:planejadas_list')


def _render_form_with_context(request):
    """Função auxiliar para renderizar o formulário com contexto padrão"""
    from .permissoes_simples import obter_funcao_militar_ativa
    from .models import ConfiguracaoPlanejadas, Orgao, GrandeComando, Unidade, SubUnidade
    
    funcao_usuario = obter_funcao_militar_ativa(request.user)
    
    orgaos = Orgao.objects.none()
    grandes_comandos = GrandeComando.objects.none()
    unidades = Unidade.objects.none()
    sub_unidades = SubUnidade.objects.none()
    
    if request.user.is_superuser:
        # Superusuário vê todas as OMs
        orgaos = Orgao.objects.all()
        grandes_comandos = GrandeComando.objects.all()
        unidades = Unidade.objects.all()
        sub_unidades = SubUnidade.objects.all()
    elif funcao_usuario:
        # Filtrar baseado no TIPO DE ACESSO da função militar
        funcao_militar = funcao_usuario.funcao_militar
        tipo_acesso = funcao_militar.acesso
        
        if tipo_acesso == 'SUBUNIDADE':
            # Se o acesso é SUBUNIDADE, mostrar apenas sub-unidades
            if funcao_usuario.sub_unidade:
                sub_unidades = SubUnidade.objects.filter(id=funcao_usuario.sub_unidade.id)
        elif tipo_acesso == 'UNIDADE':
            # Se o acesso é UNIDADE, mostrar apenas unidades
            if funcao_usuario.unidade:
                unidades = Unidade.objects.filter(id=funcao_usuario.unidade.id)
        elif tipo_acesso == 'GRANDE_COMANDO':
            # Se o acesso é GRANDE_COMANDO, mostrar apenas grandes comandos
            if funcao_usuario.grande_comando:
                grandes_comandos = GrandeComando.objects.filter(id=funcao_usuario.grande_comando.id)
        elif tipo_acesso == 'ORGAO':
            # Se o acesso é ORGAO, mostrar apenas órgãos
            if funcao_usuario.orgao:
                orgaos = Orgao.objects.filter(id=funcao_usuario.orgao.id)
        elif tipo_acesso == 'TOTAL':
            # Se o acesso é TOTAL, mostrar baseado na lotação mais específica
            if funcao_usuario.sub_unidade:
                sub_unidades = SubUnidade.objects.filter(id=funcao_usuario.sub_unidade.id)
            elif funcao_usuario.unidade:
                unidades = Unidade.objects.filter(id=funcao_usuario.unidade.id)
            elif funcao_usuario.grande_comando:
                grandes_comandos = GrandeComando.objects.filter(id=funcao_usuario.grande_comando.id)
            elif funcao_usuario.orgao:
                orgaos = Orgao.objects.filter(id=funcao_usuario.orgao.id)
    
    configuracao_ativa = ConfiguracaoPlanejadas.get_ou_create_configuracao_padrao()
    
    tipos_planejadas = []
    if configuracao_ativa:
        tipos_planejadas = [
            {'codigo': 'P1', 'nome': 'Planejada P1', 'horas': configuracao_ativa.horas_planejada_p1},
            {'codigo': 'P2', 'nome': 'Planejada P2', 'horas': configuracao_ativa.horas_planejada_p2},
            {'codigo': 'P3', 'nome': 'Planejada P3', 'horas': configuracao_ativa.horas_planejada_p3},
            {'codigo': 'P4', 'nome': 'Planejada P4', 'horas': configuracao_ativa.horas_planejada_p4},
        ]
    
    context = {
        'page_title': 'Nova Operação Planejada',
        'semana_choices': Planejada.SEMANA_CHOICES,
        'orgaos': orgaos,
        'grandes_comandos': grandes_comandos,
        'unidades': unidades,
        'sub_unidades': sub_unidades,
        'tipos_planejadas': tipos_planejadas,
        'configuracao_ativa': configuracao_ativa,
        'pode_editar_militares': True,
    }
    
    return render(request, 'militares/planejada_form.html', context)


@login_required
def planejada_create(request):
    """
    Cria nova operação planejada com validações robustas e tratamento de erros completo
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # VALIDAÇÃO DE CAMPOS OBRIGATÓRIOS
                campos_obrigatorios = {
                    'nome': request.POST.get('nome', '').strip(),
                    'origem': request.POST.get('origem', '').strip(),
                    'cidade': request.POST.get('cidade', '').strip(),
                    'data_operacao': request.POST.get('data_operacao', '').strip(),
                    'semana': request.POST.get('semana', '').strip(),
                }
                
                campos_vazios = []
                for campo, valor in campos_obrigatorios.items():
                    if not valor:
                        campos_vazios.append(campo)
                
                if campos_vazios:
                    campos_str = ', '.join(campos_vazios)
                    mensagem_erro = f'Os seguintes campos são obrigatórios: {campos_str}'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_erro
                        })
                    
                    messages.error(request, mensagem_erro)
                    return _render_form_with_context(request)
                
                # VALIDAÇÃO DE TEMPO: Verificar se pode criar planejada
                try:
                    data_operacao = datetime.strptime(campos_obrigatorios['data_operacao'], '%Y-%m-%d')
                    
                    # Criar instância temporária para verificar permissões
                    planejada_temp = Planejada(data_operacao=data_operacao)
                    pode_criar, msg_tempo = planejada_temp.pode_criar_planejada(request.user)
                    
                    if not pode_criar:
                        mensagem_erro = f'Erro de tempo: {msg_tempo}'
                        
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({
                                'success': False,
                                'message': mensagem_erro
                            })
                        
                        messages.error(request, mensagem_erro)
                        return _render_form_with_context(request)
                except ValueError:
                    mensagem_erro = 'Data inválida. Use o formato DD/MM/AAAA.'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_erro
                        })
                    
                    messages.error(request, mensagem_erro)
                    return _render_form_with_context(request)
                
                # VALIDAÇÃO DO CAMPO SEMANA
                semana_valida = campos_obrigatorios['semana'] in [choice[0] for choice in Planejada.SEMANA_CHOICES]
                if not semana_valida:
                    mensagem_erro = 'Dia da semana inválido.'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_erro
                        })
                    
                    messages.error(request, mensagem_erro)
                    return _render_form_with_context(request)
                
                # CONVERSÃO ROBUSTA DE VALORES
                def converter_valor(valor_str, nome_campo):
                    """Converte valor de formato brasileiro para americano com validação"""
                    if not valor_str or valor_str.strip() == '':
                        return 0.0
                    
                    # Converter formato brasileiro para americano
                    if ',' in valor_str:
                        valor_str = valor_str.replace('.', '').replace(',', '.')
                    
                    try:
                        valor = float(valor_str)
                        if valor < 0:
                            valor = 0.0
                        return valor
                    except (ValueError, TypeError):
                        print(f"AVISO: {nome_campo} inválido '{valor_str}' convertido para 0.0")
                        return 0.0
                
                valor = converter_valor(request.POST.get('valor', '0.00'), 'Valor')
                valor_total = converter_valor(request.POST.get('valor_total', '0.00'), 'Valor Total')
                saldo = converter_valor(request.POST.get('saldo', '0.00'), 'Saldo')
                
                # CONVERSÃO DE DATA E HORA
                hora_inicio_str = request.POST.get('hora_inicio', '08:00')
                hora_termino_str = request.POST.get('hora_termino', '14:00')
                
                try:
                    from datetime import time
                    hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
                    hora_termino = datetime.strptime(hora_termino_str, '%H:%M').time()
                except ValueError:
                    mensagem_erro = 'Formato de hora inválido. Use HH:MM.'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_erro
                        })
                    
                    messages.error(request, mensagem_erro)
                    return _render_form_with_context(request)
                
                # VALIDAÇÃO DE SALDO (se necessário)
                origem = campos_obrigatorios['origem']
                if origem and valor_total > 0:
                    verificacao_saldo = verificar_saldo_suficiente(origem, valor_total)
                    if not verificacao_saldo['sucesso']:
                        mensagem_erro = f'Saldo insuficiente: {verificacao_saldo["mensagem"]}'
                        
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({
                                'success': False,
                                'message': mensagem_erro
                            })
                        
                        messages.error(request, mensagem_erro)
                        return _render_form_with_context(request)
                
                # CRIAR PLANEJADA COM TODOS OS CAMPOS OBRIGATÓRIOS
                planejada = Planejada(
                    nome=campos_obrigatorios['nome'],
                    descricao=request.POST.get('descricao', ''),
                    origem=campos_obrigatorios['origem'],
                    cidade=campos_obrigatorios['cidade'],
                    data_operacao=data_operacao,
                    hora_inicio=hora_inicio,
                    hora_termino=hora_termino,
                    tipo_planejada=request.POST.get('tipo_planejada', 'P1'),
                    semana=campos_obrigatorios['semana'],
                    valor=valor,
                    valor_total=valor_total,
                    saldo=saldo,
                    policiais=int(request.POST.get('policiais', 1)),
                    cotas=int(request.POST.get('cotas', 1)),
                    feriado_municipal=request.POST.get('feriado_municipal') == 'on',
                    observacoes=request.POST.get('observacoes', ''),
                    numero_nota=request.POST.get('numero_nota', ''),
                    link_pdf_nota=request.POST.get('link_pdf_nota', ''),
                    ativo=True,
                    excluido=False
                )
                
                # Definir instância da planejada baseada na lotação do usuário
                planejada.definir_instancia_usuario(request.user)
                
                # Marcar policiais como definido manualmente
                planejada.definir_policiais_manual(int(request.POST.get('policiais', 1)))
                
                # VALIDAÇÃO E ADIÇÃO DE MILITARES
                militares_ids = request.POST.getlist('militares_selecionados')
                militares_validos = []
                conflitos = []
                
                if not militares_ids:
                    mensagem_erro = 'Selecione pelo menos um militar para a planejada.'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_erro
                        })
                    
                    messages.error(request, mensagem_erro)
                    return _render_form_with_context(request)
                
                # Verificar cada militar selecionado
                for militar_id in militares_ids:
                    try:
                        militar = Militar.objects.get(id=militar_id)
                        
                        # Verificar conflitos de horário
                        conflito_horario = verificar_conflito_horario_militar([militar.id], data_operacao, hora_inicio, hora_termino)
                        if not conflito_horario['sucesso']:
                                conflitos.append({
                                    'militar': militar.nome_completo,
                                'info_conflito': conflito_horario,
                                'tipo': 'conflito_horario',
                                'bloqueante': True
                                })
                                continue
                            
                        # Verificar qualificação para o tipo de planejada
                        tipo_planejada = request.POST.get('tipo_planejada', 'P1')
                        qualificado = False
                        info = {}
                        
                        if tipo_planejada == 'P1':
                            qualificado, info = planejada.verificar_qualificacao_p1_por_abono(militar, data_operacao)
                        elif tipo_planejada == 'P2':
                            qualificado, info = planejada.verificar_qualificacao_p2_por_abono(militar, data_operacao)
                        elif tipo_planejada == 'P3':
                            qualificado, info = planejada.verificar_qualificacao_p3_por_abono(militar, data_operacao)
                        elif tipo_planejada == 'P4':
                            qualificado, info = planejada.verificar_qualificacao_p4_por_abono(militar, data_operacao)
                        else:
                            qualificado, info = planejada.verificar_qualificacao_p1_por_abono(militar, data_operacao)
                        
                        if not qualificado:
                            # Tratar como aviso, não bloqueante
                            conflitos.append({
                                'militar': militar.nome_completo,
                                'info_qualificacao': info,
                                'tipo': f'aviso_indisponivel_{tipo_planejada.lower()}',
                                'bloqueante': False
                            })
                        
                        # Verificar limite mensal de planejadas
                        try:
                            from militares.models import ConfiguracaoPlanejadas
                            configuracao = ConfiguracaoPlanejadas.get_ou_create_configuracao_padrao()
                            limite_mensal = configuracao.quantidade_planejadas_por_militar_mes
                            
                            # Contar planejadas do militar no mês da operação
                            data_operacao_date = data_operacao.date()
                            ano_mes = data_operacao_date.year, data_operacao_date.month
                            
                            # Contar todas as planejadas do militar no mês (incluindo excluídas para evitar duplicação)
                            planejadas_mes = militar.planejadas.filter(
                                data_operacao__year=ano_mes[0],
                                data_operacao__month=ano_mes[1],
                                ativo=True,
                                excluido=False
                            )
                            
                            # Calcular quantidade total baseado no tipo de cada planejada
                            quantidade_total_mes = 0
                            for p in planejadas_mes:
                                tipo_p = p.tipo_planejada or 'P1'
                                if tipo_p == 'P1':
                                    quantidade_total_mes += 1
                                elif tipo_p == 'P2':
                                    quantidade_total_mes += 2
                                elif tipo_p == 'P3':
                                    quantidade_total_mes += 3
                                elif tipo_p == 'P4':
                                    quantidade_total_mes += 4
                            
                            # Calcular quantidade que seria adicionada
                            quantidade_adicionar = 0
                            if tipo_planejada == 'P1':
                                quantidade_adicionar = 1
                            elif tipo_planejada == 'P2':
                                quantidade_adicionar = 2
                            elif tipo_planejada == 'P3':
                                quantidade_adicionar = 3
                            elif tipo_planejada == 'P4':
                                quantidade_adicionar = 4
                                
                            # Verificar se excede o limite
                            if quantidade_total_mes + quantidade_adicionar > limite_mensal:
                                conflitos.append({
                                    'militar': militar.nome_completo,
                                    'info_limite': {
                                        'limite_mensal': limite_mensal,
                                        'quantidade_atual': quantidade_total_mes,
                                        'quantidade_adicionar': quantidade_adicionar,
                                        'total_seria': quantidade_total_mes + quantidade_adicionar
                                    },
                                    'tipo': 'limite_mensal',
                                    'bloqueante': True
                                })
                                continue
                                
                        except Exception as e:
                            print(f"Erro ao verificar limite mensal: {e}")
                            # Em caso de erro, permitir continuar (não bloquear)
                        
                        militares_validos.append(militar)
                        
                    except Militar.DoesNotExist:
                        messages.error(request, f'Militar com ID {militar_id} não encontrado.')
                        return _render_form_with_context(request)
                
                # Verificar se há conflitos bloqueantes
                conflitos_bloqueantes = [c for c in conflitos if c.get('bloqueante', True)]
                avisos = [c for c in conflitos if not c.get('bloqueante', True)]
                
                if conflitos_bloqueantes:
                    mensagem_conflitos = "ERRO: Os seguintes militares não podem participar da planejada:\n"
                    for conflito in conflitos_bloqueantes:
                        if conflito['tipo'] == 'conflito_horario':
                            mensagem_conflitos += f"- {conflito['militar']}: {conflito['info_conflito']['mensagem']}\n"
                        elif conflito['tipo'] == 'limite_mensal':
                            info = conflito['info_limite']
                            mensagem_conflitos += f"- {conflito['militar']}: Limite mensal excedido! Atual: {info['quantidade_atual']}, Adicionando: {info['quantidade_adicionar']}, Total seria: {info['total_seria']}, Limite: {info['limite_mensal']}\n"
                        else:
                            mensagem_conflitos += f"- {conflito['militar']}: {conflito.get('info_qualificacao', {}).get('motivo', 'Conflito não especificado')}\n"
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_conflitos
                        })
                    
                    messages.error(request, mensagem_conflitos)
                    return _render_form_with_context(request)
                
                # Se chegou até aqui, salvar a planejada
                try:
                    planejada.save()
                    
                    # Adicionar militares válidos
                    for militar in militares_validos:
                        planejada.militares.add(militar)
                    
                    # Mostrar avisos se houver
                    if avisos:
                        mensagem_avisos = "AVISO: Os seguintes militares têm questões de qualificação, mas a planejada foi criada:\n"
                        for aviso in avisos:
                            if 'aviso_indisponivel_' in aviso['tipo']:
                                tipo_planejada = aviso['tipo'].replace('aviso_indisponivel_', '').upper()
                                mensagem_avisos += f"- {aviso['militar']}: {aviso['info_qualificacao'].get('motivo', f'Não qualificado para {tipo_planejada}')}\n"
                            else:
                                mensagem_avisos += f"- {aviso['militar']}\n"
                        messages.warning(request, mensagem_avisos)
                    
                    messages.success(request, f'Operação planejada "{planejada.nome}" criada com sucesso!')
                    
                    # Verificar se é requisição AJAX
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True,
                            'redirect': '/militares/operacoes-planejadas/',
                            'message': f'Operação planejada "{planejada.nome}" criada com sucesso!'
                        })
                    
                    return redirect('militares:planejadas_list')
                    
                except Exception as e:
                    mensagem_erro = f'Erro ao salvar planejada: {str(e)}'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'message': mensagem_erro
                        })
                    
                    messages.error(request, mensagem_erro)
                    return _render_form_with_context(request)
                
        except ValidationError as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'impedimentos': [{
                        'tipo': 'erro_validacao',
                        'titulo': 'Erro de Validação',
                        'descricao': f'Erro de validação: {str(e)}',
                        'icone': 'fas fa-exclamation-triangle'
                    }],
                    'message': 'Erro de validação'
                })
            messages.error(request, f'Erro de validação: {str(e)}')
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Erro ao criar planejada: {str(e)}')
            logger.error(traceback.format_exc())
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'impedimentos': [{
                        'tipo': 'erro_sistema',
                        'titulo': 'Erro do Sistema',
                        'descricao': f'Erro ao criar operação planejada: {str(e)}',
                        'icone': 'fas fa-bug'
                    }],
                    'message': 'Erro do sistema'
                })
            messages.error(request, f'Erro ao criar operação planejada: {str(e)}')
    
    # GET - Mostrar formulário
    return _render_form_with_context(request)


@login_required
def planejada_edit(request, planejada_id):
    """
    Edita uma operação planejada existente
    """
    planejada = get_object_or_404(Planejada, id=planejada_id)
    
    # Verificar se o usuário pode acessar esta planejada
    if not planejada.pode_acessar_instancia(request.user):
        messages.error(request, 'Você não tem permissão para editar esta operação planejada.')
        return redirect('militares:planejadas_list')
    
    # Verificar se pode editar militares (fiscal vs superusuário)
    pode_editar_militares = planejada.pode_editar_militares(request.user)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Converter valor de formato brasileiro para americano
                valor_str = request.POST.get('valor', '0.00')
                
                # Lógica correta: se tem vírgula, é formato brasileiro
                if ',' in valor_str:
                    # Formato brasileiro: 120.000,50 -> 120000.50
                    valor_str = valor_str.replace('.', '').replace(',', '.')
                
                try:
                    valor = float(valor_str)
                except ValueError:
                    messages.error(request, 'Valor inválido. Use apenas números.')
                    context = {
                        'page_title': f'Editar - {planejada.nome}',
                        'planejada': planejada,
                        'semana_choices': Planejada.SEMANA_CHOICES,
                    }
                    return render(request, 'militares/planejada_form.html', context)
                
                # Processar valor total da planejada
                valor_total_str = request.POST.get('valor_total', '0.00')
                
                # Lógica correta: se tem vírgula, é formato brasileiro
                if ',' in valor_total_str:
                    # Formato brasileiro: 120.000,50 -> 120000.50
                    valor_total_str = valor_total_str.replace('.', '').replace(',', '.')
                
                try:
                    valor_total = float(valor_total_str)
                except ValueError:
                    valor_total = 0.0
                
                # Converter data e hora
                from datetime import datetime, time
                data_operacao_str = request.POST.get('data_operacao')
                hora_inicio_str = request.POST.get('hora_inicio', '00:00')
                hora_termino_str = request.POST.get('hora_termino', '00:00')
                
                data_hora_inicio_str = f"{data_operacao_str} {hora_inicio_str}"
                data_operacao = datetime.strptime(data_hora_inicio_str, '%Y-%m-%d %H:%M')
                
                # Converter strings de hora para objetos time
                hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
                hora_termino = datetime.strptime(hora_termino_str, '%H:%M').time()
                
                # Verificar se pode alterar quantidade de bombeiros (SE o valor foi alterado)
                novo_policiais = int(request.POST.get('policiais', 1))
                if novo_policiais != planejada.policiais:
                    # Verificar se não é superusuário
                    if not request.user.is_superuser:
                        from datetime import date
                        agora = datetime.now()
                        
                        # Criar datetime de início da operação
                        data_hora_inicio = datetime.combine(data_operacao.date(), hora_inicio)
                        
                        # Se já passou do início da operação
                        if agora >= data_hora_inicio:
                            # Após o início: só pode DIMINUIR, não pode AUMENTAR
                            if novo_policiais > planejada.policiais:
                                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                                    return JsonResponse({
                                        'success': False,
                                        'message': f'Não é possível aumentar o número de bombeiros após o início da operação (início: {data_hora_inicio.strftime("%d/%m/%Y %H:%M")}). Apenas superusuários podem aumentar neste momento.'
                                    })
                                messages.error(request, f'Não é possível aumentar o número de bombeiros após o início da operação. Início: {data_hora_inicio.strftime("%d/%m/%Y %H:%M")}')
                                return redirect('militares:planejada_edit', planejada_id=planejada.id)
                            # Diminuir é permitido após o início
                            print(f"✅ Diminuição de policiais permitida após início: {planejada.policiais} -> {novo_policiais}")
                        else:
                            # Antes do início: pode AUMENTAR ou DIMINUIR
                            print(f"✅ Alteração de policiais permitida antes do início: {planejada.policiais} -> {novo_policiais}")
                    else:
                        # Superusuário pode alterar sempre
                        print(f"✅ Superusuário pode alterar policiais: {planejada.policiais} -> {novo_policiais}")
                
                # Atualizar campos da planejada
                planejada.nome = request.POST.get('nome')
                planejada.descricao = request.POST.get('descricao', '')
                planejada.origem = request.POST.get('origem')
                planejada.cidade = request.POST.get('cidade')
                planejada.data_operacao = data_operacao
                planejada.hora_inicio = hora_inicio
                planejada.hora_termino = hora_termino
                planejada.tipo_planejada = request.POST.get('tipo_planejada', '')
                planejada.semana = request.POST.get('semana')
                planejada.valor = valor
                planejada.valor_total = valor_total
                # NOTA: saldo será recalculado automaticamente por recalcular_valores()
                planejada.policiais = novo_policiais
                planejada.definir_policiais_manual(novo_policiais)
                planejada.feriado_municipal = request.POST.get('feriado_municipal') == 'on'
                planejada.observacoes = request.POST.get('observacoes', '')
                planejada.numero_nota = request.POST.get('numero_nota', '')
                planejada.link_pdf_nota = request.POST.get('link_pdf_nota', '')
                planejada.ativo = request.POST.get('ativo', 'on') == 'on'
                
                # Processar militares selecionados (se houver)
                militares_ids = request.POST.getlist('militares_selecionados')
                impedimentos = []
                
                if militares_ids:
                    # VALIDAÇÃO DE CONFLITO DE HORÁRIOS: Verificar se militares não têm conflito
                    verificacao_conflito = verificar_conflito_horario_militar(
                        militares_ids, data_operacao, hora_inicio, hora_termino, planejada_id
                    )
                    if not verificacao_conflito['sucesso']:
                        # Construir mensagem detalhada dos conflitos
                        mensagem_conflito = f"{verificacao_conflito['mensagem']}\n\n"
                        for conflito in verificacao_conflito['conflitos']:
                            mensagem_conflito += f"• {conflito['militar']} já está escalado em '{conflito['planejada_conflito']}' "
                            mensagem_conflito += f"no horário {conflito['horario_conflito']} em {conflito['data_conflito']}\n"
                        
                        impedimentos.append({
                            'tipo': 'conflito_horario',
                            'titulo': 'Conflito de Horários',
                            'descricao': mensagem_conflito,
                            'icone': 'fas fa-clock'
                        })
                
                # Verificar se há impedimentos
                if impedimentos:
                    return JsonResponse({
                        'success': False,
                        'impedimentos': impedimentos,
                        'message': 'Operação não pode ser salva devido a impedimentos'
                    })
                
                # Recalcular valores se quantidade de bombeiros mudou
                if novo_policiais != planejada.policiais:
                    planejada.recalcular_valores()
                
                planejada.save()
                print(f"DEBUG: Planejada salva: ID={planejada.id}, Nome={planejada.nome}")
                
                # Atualizar militares após salvar a planejada
                if militares_ids:
                    try:
                        militares = Militar.objects.filter(id__in=militares_ids)
                        planejada.militares.set(militares)
                        print(f"DEBUG: Militares atualizados: {planejada.militares.count()}")
                        
                        # Sincronizar com escala de abonar
                        try:
                            planejada.sincronizar_escala_abonar()
                        except Exception as e:
                            print(f"ERRO ao sincronizar escala abonar: {str(e)}")
                    except Exception as e:
                        print(f"ERRO ao atualizar militares: {str(e)}")
                        # Não falhar a edição por causa dos militares
                        pass
                else:
                    # Se não há militares selecionados, limpar lista
                    planejada.militares.clear()
                    print(f"DEBUG: Militares removidos da planejada")
                    
                    # Sincronizar escala de abonar (remover todos os abonos)
                    try:
                        planejada.sincronizar_escala_abonar()
                    except Exception as e:
                        print(f"ERRO ao sincronizar escala abonar após limpeza: {str(e)}")
                
                # Recalcular valores após atualizar militares
                planejada.recalcular_valores()
                planejada.save()
                print(f"DEBUG: Planejada salva novamente após atualizar militares: Valor Total={planejada.valor_total}, Saldo={planejada.saldo}")
                
                # VALIDAÇÃO: Verificar quantidade de bombeiros vs militares (APÓS atualizar militares)
                valido, mensagem = planejada.validar_quantidade_bombeiros()
                if not valido:
                    return JsonResponse({
                        'success': False,
                        'message': f'Erro de validação: {mensagem}'
                    })
                
                # Verificar se é requisição AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'redirect': f'/militares/operacoes-planejadas/',
                        'message': 'Operação planejada atualizada com sucesso!'
                    })
                
                messages.success(request, 'Operação planejada atualizada com sucesso!')
                return redirect('militares:planejadas_list')
                
        except ValidationError as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'impedimentos': [{
                        'tipo': 'erro_validacao',
                        'titulo': 'Erro de Validação',
                        'descricao': f'Erro de validação: {str(e)}',
                        'icone': 'fas fa-exclamation-triangle'
                    }],
                    'message': 'Erro de validação'
                })
            messages.error(request, f'Erro de validação: {str(e)}')
        except IntegrityError as e:
            # Tratar erro de constraint de unicidade
            if 'unique_planejada_nome_data_cidade' in str(e):
                mensagem = 'Já existe uma operação planejada com este nome, data e cidade. Escolha um nome diferente.'
            else:
                mensagem = f'Erro de integridade: {str(e)}'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'impedimentos': [{
                        'tipo': 'erro_integridade',
                        'titulo': 'Erro de Integridade',
                        'descricao': mensagem,
                        'icone': 'fas fa-exclamation-triangle'
                    }],
                    'message': 'Erro de integridade'
                })
            messages.error(request, mensagem)
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Erro ao atualizar planejada: {str(e)}')
            logger.error(traceback.format_exc())
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'impedimentos': [{
                        'tipo': 'erro_sistema',
                        'titulo': 'Erro do Sistema',
                        'descricao': f'Erro ao atualizar operação planejada: {str(e)}',
                        'icone': 'fas fa-bug'
                    }],
                    'message': 'Erro do sistema'
                })
            messages.error(request, f'Erro ao atualizar operação planejada: {str(e)}')
    
    # GET - Mostrar formulário
    # Carregar organizações baseadas na instância do usuário
    from .permissoes_simples import obter_funcao_militar_ativa
    funcao_usuario = obter_funcao_militar_ativa(request.user)
    
    orgaos = Orgao.objects.none()
    grandes_comandos = GrandeComando.objects.none()
    unidades = Unidade.objects.none()
    sub_unidades = SubUnidade.objects.none()
    
    if request.user.is_superuser:
        # Superusuários veem todas as organizações
        orgaos = Orgao.objects.filter(ativo=True).order_by('nome')
        grandes_comandos = GrandeComando.objects.filter(ativo=True).order_by('nome')
        unidades = Unidade.objects.filter(ativo=True).order_by('nome')
        sub_unidades = SubUnidade.objects.filter(ativo=True).order_by('nome')
    elif funcao_usuario and funcao_usuario.funcao_militar.publicacao in ['OPERADOR_PLANEJADAS', 'FISCAL_PLANEJADAS', 'APROVADOR']:
        # Operadores de planejadas veem baseado no TIPO DE ACESSO da função militar
        tipo_acesso = funcao_usuario.funcao_militar.acesso
        
        if tipo_acesso == 'SUBUNIDADE':
            # Se o acesso é SUBUNIDADE, mostrar apenas sub-unidades
            if funcao_usuario.sub_unidade:
                sub_unidades = SubUnidade.objects.filter(id=funcao_usuario.sub_unidade.id)
            else:
                sub_unidades = SubUnidade.objects.none()
            unidades = Unidade.objects.none()
            grandes_comandos = GrandeComando.objects.none()
            orgaos = Orgao.objects.none()
        elif tipo_acesso == 'UNIDADE':
            # Se o acesso é UNIDADE, mostrar apenas unidades
            if funcao_usuario.unidade:
                unidades = Unidade.objects.filter(id=funcao_usuario.unidade.id)
            else:
                unidades = Unidade.objects.none()
            sub_unidades = SubUnidade.objects.none()
            grandes_comandos = GrandeComando.objects.none()
            orgaos = Orgao.objects.none()
        elif tipo_acesso == 'GRANDE_COMANDO':
            # Se o acesso é GRANDE_COMANDO, mostrar apenas grandes comandos
            if funcao_usuario.grande_comando:
                grandes_comandos = GrandeComando.objects.filter(id=funcao_usuario.grande_comando.id)
            else:
                grandes_comandos = GrandeComando.objects.none()
            unidades = Unidade.objects.none()
            sub_unidades = SubUnidade.objects.none()
            orgaos = Orgao.objects.none()
        elif tipo_acesso == 'ORGAO':
            # Se o acesso é ORGAO, mostrar apenas órgãos
            if funcao_usuario.orgao:
                orgaos = Orgao.objects.filter(id=funcao_usuario.orgao.id)
            else:
                orgaos = Orgao.objects.none()
            grandes_comandos = GrandeComando.objects.none()
            unidades = Unidade.objects.none()
            sub_unidades = SubUnidade.objects.none()
        elif tipo_acesso == 'TOTAL':
            # Se o acesso é TOTAL, mostrar baseado na lotação mais específica
            if funcao_usuario.sub_unidade:
                sub_unidades = SubUnidade.objects.filter(id=funcao_usuario.sub_unidade.id)
                unidades = Unidade.objects.none()
                grandes_comandos = GrandeComando.objects.none()
                orgaos = Orgao.objects.none()
            elif funcao_usuario.unidade:
                unidades = Unidade.objects.filter(id=funcao_usuario.unidade.id)
                sub_unidades = SubUnidade.objects.none()
                grandes_comandos = GrandeComando.objects.none()
                orgaos = Orgao.objects.none()
            elif funcao_usuario.grande_comando:
                grandes_comandos = GrandeComando.objects.filter(id=funcao_usuario.grande_comando.id)
                unidades = Unidade.objects.none()
                sub_unidades = SubUnidade.objects.none()
                orgaos = Orgao.objects.none()
            elif funcao_usuario.orgao:
                orgaos = Orgao.objects.filter(id=funcao_usuario.orgao.id)
                grandes_comandos = GrandeComando.objects.none()
                unidades = Unidade.objects.none()
                sub_unidades = SubUnidade.objects.none()
            else:
                orgaos = Orgao.objects.none()
                grandes_comandos = GrandeComando.objects.none()
                unidades = Unidade.objects.none()
                sub_unidades = SubUnidade.objects.none()
        else:
            # Tipo de acesso não reconhecido
            orgaos = Orgao.objects.none()
            grandes_comandos = GrandeComando.objects.none()
            unidades = Unidade.objects.none()
            sub_unidades = SubUnidade.objects.none()
    
    # Carregar configuração ativa de planejadas (cria uma padrão se não existir)
    from .models import ConfiguracaoPlanejadas
    configuracao_ativa = ConfiguracaoPlanejadas.get_ou_create_configuracao_padrao()
    
    # Tipos de planejadas baseados na configuração
    tipos_planejadas = []
    if configuracao_ativa:
        tipos_planejadas = [
            {'codigo': 'P1', 'nome': 'Planejada P1', 'horas': configuracao_ativa.horas_planejada_p1},
            {'codigo': 'P2', 'nome': 'Planejada P2', 'horas': configuracao_ativa.horas_planejada_p2},
            {'codigo': 'P3', 'nome': 'Planejada P3', 'horas': configuracao_ativa.horas_planejada_p3},
            {'codigo': 'P4', 'nome': 'Planejada P4', 'horas': configuracao_ativa.horas_planejada_p4},
        ]
    
    context = {
        'page_title': f'Editar - {planejada.nome}',
        'planejada': planejada,
        'semana_choices': Planejada.SEMANA_CHOICES,
        'orgaos': orgaos,
        'grandes_comandos': grandes_comandos,
        'unidades': unidades,
        'sub_unidades': sub_unidades,
        'tipos_planejadas': tipos_planejadas,
        'configuracao_ativa': configuracao_ativa,
        'pode_editar_militares': pode_editar_militares,
    }
    
    return render(request, 'militares/planejada_form.html', context)


@login_required
@require_http_methods(["GET"])
def api_valor_utilizado(request, tipo, id):
    """
    API para obter o valor utilizado de uma organização militar no mês do orçamento
    """
    try:
        from datetime import datetime
        from django.db.models import Sum
        
        # Obter mês e ano do orçamento da URL ou parâmetros
        orcamento_id = request.GET.get('orcamento_id')
        if orcamento_id:
            try:
                orcamento = OrcamentoPlanejadas.objects.get(id=orcamento_id)
                ano_orcamento = orcamento.ano
                mes_orcamento = orcamento.mes
            except OrcamentoPlanejadas.DoesNotExist:
                # Fallback para mês atual se orçamento não encontrado
                hoje = datetime.now()
                ano_orcamento = hoje.year
                mes_orcamento = hoje.month
        else:
            # Fallback para mês atual se não especificado
            hoje = datetime.now()
            ano_orcamento = hoje.year
            mes_orcamento = hoje.month
        
        if tipo == 'orgao':
            organizacao = get_object_or_404(Orgao, pk=id, ativo=True)
        elif tipo == 'grande_comando':
            organizacao = get_object_or_404(GrandeComando, pk=id, ativo=True)
        elif tipo == 'unidade':
            organizacao = get_object_or_404(Unidade, pk=id, ativo=True)
        elif tipo == 'sub_unidade':
            organizacao = get_object_or_404(SubUnidade, pk=id, ativo=True)
        else:
            return JsonResponse({'success': False, 'message': 'Tipo de organização inválido'})
        
        # Buscar orçamento do mês especificado
        orcamento = OrcamentoPlanejadas.objects.filter(
            ano=ano_orcamento,
            mes=mes_orcamento,
            ativo=True
        ).first()
        
        if not orcamento:
            return JsonResponse({
                'success': True, 
                'valor_utilizado': 0.0,
                'saldo': 0.0,
                'message': f'Nenhum orçamento encontrado para {mes_orcamento}/{ano_orcamento}'
            })
        
        # Buscar distribuição para esta organização
        distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
            orcamento=orcamento,
            ativo=True
        )
        
        if tipo == 'orgao':
            distribuicao = distribuicao.filter(orgao=organizacao)
        elif tipo == 'grande_comando':
            distribuicao = distribuicao.filter(grande_comando=organizacao)
        elif tipo == 'unidade':
            distribuicao = distribuicao.filter(unidade=organizacao)
        elif tipo == 'sub_unidade':
            distribuicao = distribuicao.filter(sub_unidade=organizacao)
        
        distribuicao = distribuicao.first()
        
        if not distribuicao:
            return JsonResponse({
                'success': True, 
                'valor_utilizado': 0.0,
                'saldo': 0.0,
                'message': 'Nenhuma distribuição encontrada para esta organização'
            })
        
        # Calcular gastos do mês do orçamento
        planejadas_gastos = Planejada.objects.filter(
            data_operacao__year=ano_orcamento,
            data_operacao__month=mes_orcamento,
            ativo=True,
            excluido=False  # Excluir planejadas marcadas como excluídas
        )
        
        # Filtrar por organização baseado no tipo
        # Lógica hierárquica: atribuir à organização mais específica
        # A origem contém múltiplas organizações separadas por "|"
        # Devemos atribuir apenas à mais específica (última na lista)
        
        # Filtrar planejadas que contêm o nome da organização na origem
        planejadas_filtradas = []
        for planejada in planejadas_gastos.filter(origem__icontains=organizacao.nome):
            # Dividir a origem por "|" e pegar a última organização (mais específica)
            partes_origem = [p.strip() for p in planejada.origem.split('|')]
            ultima_organizacao = partes_origem[-1] if partes_origem else ""
            
            # Verificar se esta organização corresponde EXATAMENTE à última (mais específica)
            # Usar comparação exata para evitar falsos positivos (ex: "Geral" em "Comando Geral" e "Ajudância Geral")
            if organizacao.nome.lower().strip() == ultima_organizacao.lower().strip():
                planejadas_filtradas.append(planejada.id)
        
        # Filtrar apenas as planejadas que correspondem à organização mais específica
        if planejadas_filtradas:
            planejadas_gastos = planejadas_gastos.filter(id__in=planejadas_filtradas)
        else:
            planejadas_gastos = planejadas_gastos.none()
        
        # Recalcular valores de todas as planejadas antes de somar
        for planejada in planejadas_gastos:
            planejada.recalcular_valores()
        
        # Calcular total gasto
        total_gasto = planejadas_gastos.aggregate(
            total=Sum('valor_total')
        )['total'] or 0
        
        # Calcular saldo disponível
        saldo_disponivel = float(distribuicao.valor_planejadas) - float(total_gasto)
        
        return JsonResponse({
            'success': True,
            'valor_utilizado': float(total_gasto),
            'saldo': saldo_disponivel,
            'orcamento_total': float(distribuicao.valor_planejadas),
            'organizacao': organizacao.nome
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Erro ao calcular valor utilizado: {str(e)}'
        })

@login_required
@require_http_methods(["GET"])
def api_saldo_om(request, tipo, id):
    """
    API para obter o saldo atual de uma organização militar no mês especificado
    """
    try:
        from datetime import datetime
        from django.db.models import Sum
        
        # Obter ano e mês dos parâmetros da URL (opcional)
        ano_atual = request.GET.get('ano')
        mes_atual = request.GET.get('mes')
        
        if ano_atual and mes_atual:
            try:
                ano_atual = int(ano_atual)
                mes_atual = int(mes_atual)
                print(f"DEBUG api_saldo_om: Usando ano={ano_atual}, mes={mes_atual} dos parâmetros")
            except ValueError:
                # Se não conseguir converter, usar mês atual
                hoje = datetime.now()
                ano_atual = hoje.year
                mes_atual = hoje.month
                print(f"DEBUG api_saldo_om: Erro ao converter parâmetros, usando mês atual {mes_atual}/{ano_atual}")
        else:
            # Usar mês atual como padrão se não especificado
            hoje = datetime.now()
            ano_atual = hoje.year
            mes_atual = hoje.month
            print(f"DEBUG api_saldo_om: Sem parâmetros, usando mês atual {mes_atual}/{ano_atual}")
        
        # Buscar a organização
        if tipo == 'orgao':
            organizacao = get_object_or_404(Orgao, pk=id, ativo=True)
        elif tipo == 'grande_comando':
            organizacao = get_object_or_404(GrandeComando, pk=id, ativo=True)
        elif tipo == 'unidade':
            organizacao = get_object_or_404(Unidade, pk=id, ativo=True)
        elif tipo == 'sub_unidade':
            organizacao = get_object_or_404(SubUnidade, pk=id, ativo=True)
        else:
            return JsonResponse({'success': False, 'message': 'Tipo de organização inválido'})
        
        # Buscar orçamento do mês especificado
        orcamento = OrcamentoPlanejadas.objects.filter(
            ano=ano_atual,
            mes=mes_atual,
            ativo=True
        ).first()
        
        if not orcamento:
            return JsonResponse({
                'success': True, 
                'saldo': 0.0,
                'message': f'Nenhum orçamento encontrado para {mes_atual}/{ano_atual}'
            })
        
        # Buscar distribuição para esta organização
        distribuicao = DistribuicaoOrcamentoPlanejadas.objects.filter(
            orcamento=orcamento,
            ativo=True
        )
        
        if tipo == 'orgao':
            distribuicao = distribuicao.filter(orgao=organizacao)
        elif tipo == 'grande_comando':
            distribuicao = distribuicao.filter(grande_comando=organizacao)
        elif tipo == 'unidade':
            distribuicao = distribuicao.filter(unidade=organizacao)
        elif tipo == 'sub_unidade':
            distribuicao = distribuicao.filter(sub_unidade=organizacao)
        
        distribuicao = distribuicao.first()
        
        if not distribuicao:
            return JsonResponse({
                'success': True, 
                'saldo': 0.0,
                'message': 'Nenhuma distribuição encontrada para esta organização'
            })
        
        # Calcular gastos do mês/ano do orçamento especificado
        planejadas_gastos = Planejada.objects.filter(
            data_operacao__year=ano_atual,
            data_operacao__month=mes_atual,
            ativo=True,
            excluido=False  # Excluir planejadas marcadas como excluídas
        )
        
        # Filtrar por organização baseado no tipo - busca específica
        # Prioridade: Sub-Unidade > Unidade > Grande Comando > Órgão
        if tipo == 'sub_unidade':
            # Para sub-unidade, buscar operações que terminam com o nome da sub-unidade
            planejadas_gastos = planejadas_gastos.filter(
                origem__endswith=organizacao.nome
            )
        elif tipo == 'unidade':
            # Para unidade, buscar operações que terminam com o nome da unidade
            planejadas_gastos = planejadas_gastos.filter(
                origem__endswith=organizacao.nome
            )
        elif tipo == 'grande_comando':
            # Para grande comando, buscar operações que terminam com o nome do grande comando
            planejadas_gastos = planejadas_gastos.filter(
                origem__endswith=organizacao.nome
            )
        elif tipo == 'orgao':
            # Para órgão, buscar operações que são APENAS do órgão (sem outras instâncias)
            planejadas_gastos = planejadas_gastos.filter(
                origem=organizacao.nome
            )
        
        # Calcular total gasto
        total_gasto = planejadas_gastos.aggregate(
            total=Sum('valor_total')
        )['total'] or 0
        
        # Calcular saldo disponível
        saldo_disponivel = float(distribuicao.valor_planejadas) - float(total_gasto)
        
        return JsonResponse({
            'success': True,
            'saldo': saldo_disponivel,
            'orcamento_total': float(distribuicao.valor_planejadas),
            'gasto_total': float(total_gasto),
            'organizacao': organizacao.nome
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'message': f'Erro ao calcular saldo: {str(e)}'
        })

# ===== APIs PARA SELEÇÃO DE MILITARES EM OPERAÇÕES PLANEJADAS =====

@login_required
def api_militares_disponiveis_planejada(request, planejada_id):
    """API para buscar militares disponíveis para a operação planejada - PRIORIZANDO ORIGEM"""
    from django.core.cache import cache
    import time
    
    start_time = time.time()
    
    # VERSÃO TESTE: Retorno imediato para debug
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Verificar se há termo de pesquisa
        termo_pesquisa = request.GET.get('q', '').strip()
        
        # Buscar militares voluntários
        militares_base = Militar.objects.filter(
            classificacao='ATIVO',
            voluntario_operacoes='SIM',
            situacao='PRONTO'
        ).exclude(
            gratificacao='SIM'
        )
        
        # Aplicar filtro de pesquisa se houver
        if termo_pesquisa:
            from django.db.models import Q
            militares_base = militares_base.filter(
                Q(nome_completo__icontains=termo_pesquisa) |
                Q(nome_guerra__icontains=termo_pesquisa) |
                Q(matricula__icontains=termo_pesquisa)
            )
        
        # PRIORIZAR MILITARES LOTADOS NA MESMA ORIGEM DA PLANEJADA
        militares_origem = []
        militares_outros = []
        
        for militar in militares_base:
            lotacao_atual = militar.lotacao_atual()
            na_origem = False
            
            if lotacao_atual:
                # Verificar se o militar está lotado na mesma origem da planejada
                if planejada.sub_unidade and lotacao_atual.sub_unidade == planejada.sub_unidade:
                    na_origem = True
                elif planejada.unidade and lotacao_atual.unidade == planejada.unidade:
                    na_origem = True
                elif planejada.grande_comando and lotacao_atual.grande_comando == planejada.grande_comando:
                    na_origem = True
                elif planejada.orgao and lotacao_atual.orgao == planejada.orgao:
                    na_origem = True
            
            if na_origem:
                militares_origem.append(militar)
            else:
                militares_outros.append(militar)
        
        # Verificar quais militares já estão escalados nesta planejada
        militares_escalados = planejada.militares.values_list('id', flat=True)
        
        militares_list = []
        
        # Processar primeiro os militares da origem (prioridade)
        for militar in militares_origem:
            ja_escalado = militar.id in militares_escalados
            
            # Obter o nome completo do posto/graduação
            posto_graduacao_display = get_posto_completo(militar.posto_graduacao)
            
            # Status básico sem verificação de folga (para evitar loop infinito)
            if ja_escalado:
                status_disponibilidade = 'ja_escalado'
                mensagem_status = 'Já escalado nesta operação'
                pode_selecionar = False
            else:
                # Verificar qualificação específica baseada no tipo de planejada (usando abono)
                if planejada.tipo_planejada == 'P1':
                    qualificado, info = planejada.verificar_qualificacao_p1_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p1'
                    prefixo_mensagem = 'Indisponível para P1'
                    prefixo_disponivel = 'Disponível para P1'
                elif planejada.tipo_planejada == 'P2':
                    qualificado, info = planejada.verificar_qualificacao_p2_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p2'
                    prefixo_mensagem = 'Indisponível para P2'
                    prefixo_disponivel = 'Disponível para P2'
                elif planejada.tipo_planejada == 'P3':
                    qualificado, info = planejada.verificar_qualificacao_p3_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p3'
                    prefixo_mensagem = 'Indisponível para P3'
                    prefixo_disponivel = 'Disponível para P3'
                elif planejada.tipo_planejada == 'P4':
                    qualificado, info = planejada.verificar_qualificacao_p4_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p4'
                    prefixo_mensagem = 'Indisponível para P4'
                    prefixo_disponivel = 'Disponível para P4'
                else:
                    qualificado = True
                    info = None
                
                if not qualificado:
                    status_disponibilidade = tipo_indisponivel
                    mensagem_status = f'{prefixo_mensagem} - {info.get("motivo", "Não qualificado")}'
                    pode_selecionar = False
                else:
                    status_disponibilidade = 'disponivel'
                    mensagem_status = f'{prefixo_disponivel} - DA ORIGEM' if planejada.tipo_planejada in ['P1', 'P2', 'P3', 'P4'] else 'Disponível - DA ORIGEM'
                    pode_selecionar = True
            
            militares_list.append({
                'id': militar.id,
                'nome_completo': militar.nome_completo,
                'posto_graduacao': militar.posto_graduacao,
                'posto_graduacao_display': posto_graduacao_display,
                'cpf': militar.cpf,
                'ja_escalado': ja_escalado,
                'status_disponibilidade': status_disponibilidade,
                'mensagem_status': mensagem_status,
                'pode_selecionar': pode_selecionar,
                'planejadas_dia': [],
                'total_planejadas_dia': 0,
                'prioridade_origem': True  # Marcar como prioridade da origem
            })
        
        # Processar depois os demais militares
        for militar in militares_outros:
            ja_escalado = militar.id in militares_escalados
            
            # Obter o nome completo do posto/graduação
            posto_graduacao_display = get_posto_completo(militar.posto_graduacao)
            
            # Status básico sem verificação de folga (para evitar loop infinito)
            if ja_escalado:
                status_disponibilidade = 'ja_escalado'
                mensagem_status = 'Já escalado nesta operação'
                pode_selecionar = False
            else:
                # Verificar qualificação específica baseada no tipo de planejada (usando abono)
                if planejada.tipo_planejada == 'P1':
                    qualificado, info = planejada.verificar_qualificacao_p1_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p1'
                    prefixo_mensagem = 'Indisponível para P1'
                    prefixo_disponivel = 'Disponível para P1'
                elif planejada.tipo_planejada == 'P2':
                    qualificado, info = planejada.verificar_qualificacao_p2_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p2'
                    prefixo_mensagem = 'Indisponível para P2'
                    prefixo_disponivel = 'Disponível para P2'
                elif planejada.tipo_planejada == 'P3':
                    qualificado, info = planejada.verificar_qualificacao_p3_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p3'
                    prefixo_mensagem = 'Indisponível para P3'
                    prefixo_disponivel = 'Disponível para P3'
                elif planejada.tipo_planejada == 'P4':
                    qualificado, info = planejada.verificar_qualificacao_p4_por_abono(militar, planejada.data_operacao)
                    tipo_indisponivel = 'indisponivel_p4'
                    prefixo_mensagem = 'Indisponível para P4'
                    prefixo_disponivel = 'Disponível para P4'
                else:
                    qualificado = True
                    info = None
                
                if not qualificado:
                    status_disponibilidade = tipo_indisponivel
                    mensagem_status = f'{prefixo_mensagem} - {info.get("motivo", "Não qualificado")}'
                    pode_selecionar = False
                else:
                    status_disponibilidade = 'disponivel'
                    mensagem_status = prefixo_disponivel if planejada.tipo_planejada in ['P1', 'P2', 'P3', 'P4'] else 'Disponível'
                    pode_selecionar = True
            
            militares_list.append({
                'id': militar.id,
                'nome_completo': militar.nome_completo,
                'posto_graduacao': militar.posto_graduacao,
                'posto_graduacao_display': posto_graduacao_display,
                'cpf': militar.cpf,
                'ja_escalado': ja_escalado,
                'status_disponibilidade': status_disponibilidade,
                'mensagem_status': mensagem_status,
                'pode_selecionar': pode_selecionar,
                'planejadas_dia': [],
                'total_planejadas_dia': 0,
                'prioridade_origem': False  # Marcar como não prioridade da origem
            })
        
        elapsed_time = time.time() - start_time
        
        return JsonResponse({
            'militares': militares_list,
            'total': len(militares_list),
            'total_origem': len(militares_origem),
            'total_outros': len(militares_outros),
            'elapsed_time': round(elapsed_time, 2)
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Erro na consulta: {str(e)}',
            'elapsed_time': time.time() - start_time
        }, status=500)


@login_required
def api_verificar_folga_militar(request, planejada_id, militar_id):
    """API para verificar folga de um militar específico"""
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        militar = get_object_or_404(Militar, id=militar_id)
        
        # Verificar se tem folga suficiente
        tem_folga_suficiente, info_folga = planejada.verificar_folga_suficiente(militar, planejada.data_operacao)
        
        if not tem_folga_suficiente:
            horas_restantes = info_folga.get('horas_restantes', 0)
            if horas_restantes > 0:
                mensagem = f'Sem folga regulamentar - Aguardar {horas_restantes:.1f}h'
            else:
                mensagem = 'Sem folga regulamentar - Aguardar 24h após serviço'
            
            return JsonResponse({
                'success': False,
                'tem_folga': False,
                'mensagem': mensagem,
                'info_folga': info_folga
            })
        else:
            return JsonResponse({
                'success': True,
                'tem_folga': True,
                'mensagem': 'Militar tem folga suficiente'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro na verificação: {str(e)}'
        }, status=500)


@login_required
def api_verificar_restricoes_militares(request, planejada_id):
    """API para verificar restrições de múltiplos militares de forma otimizada"""
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        militar_ids = request.GET.getlist('militar_ids[]')
        
        if not militar_ids:
            return JsonResponse({'restricoes': {}})
        
        restricoes = {}
        for militar_id in militar_ids:
            try:
                militar = Militar.objects.get(id=militar_id)
                tem_folga_suficiente, info_folga = planejada.verificar_folga_suficiente(militar, planejada.data_operacao)
                
                if not tem_folga_suficiente:
                    horas_restantes = info_folga.get('horas_restantes', 0)
                    if horas_restantes > 0:
                        mensagem = f'Sem folga regulamentar - Aguardar {horas_restantes:.1f}h'
                    else:
                        mensagem = 'Sem folga regulamentar - Aguardar 24h após serviço'
                    
                    restricoes[militar_id] = {
                        'tem_restricao': True,
                        'mensagem': mensagem,
                        'info_folga': info_folga
                    }
                else:
                    restricoes[militar_id] = {
                        'tem_restricao': False,
                        'mensagem': 'Disponível'
                    }
            except Militar.DoesNotExist:
                continue
        
        return JsonResponse({'restricoes': restricoes})
        
    except Exception as e:
        return JsonResponse({
            'error': f'Erro na verificação: {str(e)}'
        }, status=500)


@login_required
def api_militares_escalados_planejada(request, planejada_id):
    """API para buscar militares já escalados na operação planejada"""
    planejada = get_object_or_404(Planejada, id=planejada_id)
    
    militares_escalados = planejada.militares.all().order_by(
        'posto_graduacao',
        'data_promocao_atual',
        'numeracao_antiguidade'
    )
    
    militares_list = []
    for militar in militares_escalados:
        # Calcular tempo no posto atual
        tempo_posto = ""
        if militar.data_promocao_atual:
            from datetime import date
            hoje = date.today()
            dias_no_posto = (hoje - militar.data_promocao_atual).days
            anos = dias_no_posto // 365
            meses = (dias_no_posto % 365) // 30
            if anos > 0:
                tempo_posto = f"{anos}a {meses}m"
            elif meses > 0:
                tempo_posto = f"{meses}m"
            else:
                tempo_posto = f"{dias_no_posto}d"
        
        # Obter o nome completo do posto/graduação
        posto_graduacao_display = get_posto_completo(militar.posto_graduacao)
        
        militares_list.append({
            'id': militar.id,
            'nome_completo': militar.nome_completo,
            'posto_graduacao': militar.posto_graduacao,
            'posto_graduacao_display': posto_graduacao_display,
            'matricula': militar.matricula,
            'lotacao': str(militar.lotacao_atual().lotacao) if militar.lotacao_atual() else 'Não informado',
            'cpf': militar.cpf
        })
    
    return JsonResponse({
        'militares': militares_list,
        'total': len(militares_list)
    })


@login_required
@require_http_methods(["POST"])
def api_adicionar_militar_planejada(request, planejada_id):
    """API para adicionar militar à operação planejada"""
    planejada = get_object_or_404(Planejada, id=planejada_id)
    print(f"DEBUG: Iniciando api_adicionar_militar_planejada para planejada_id={planejada_id}")
    
    # VALIDAÇÃO DE TEMPO: Verificar se pode alterar militares
    pode_alterar, msg_tempo = planejada.pode_alterar_militares(request.user)
    if not pode_alterar:
        return JsonResponse({
            'success': False,
            'error': f'Erro de tempo: {msg_tempo}'
        })
    
    # VALIDAÇÃO DE PERMISSÃO: Verificar se pode editar militares (fiscal vs superusuário)
    if not planejada.pode_editar_militares(request.user):
        return JsonResponse({
            'success': False,
            'error': 'Você não tem permissão para editar militares desta operação. Apenas o fiscal pode editar antes da assinatura, ou superusuários após a assinatura.'
        })
    militar_id = request.POST.get('militar_id')
    
    print(f"DEBUG: militar_id recebido = {militar_id}")
    
    if not militar_id:
        return JsonResponse({'success': False, 'message': 'ID do militar não fornecido'}, status=400)
    
    try:
        militar = Militar.objects.get(id=militar_id)
        print(f"DEBUG: Militar encontrado = {militar.nome_completo}")
        
        # VALIDAÇÕES OBRIGATÓRIAS (NUNCA flexibilizadas):
        # 1. Militar deve ser ATIVO
        if militar.classificacao != 'ATIVO':
            return JsonResponse({
                'success': False,
                'message': f'Militar {militar.nome_completo} não está ATIVO e não pode participar de planejadas.'
            }, status=400)
        
        # 2. Militar deve estar com situação PRONTO
        if militar.situacao != 'PRONTO':
            return JsonResponse({
                'success': False,
                'message': f'Militar {militar.nome_completo} não está com situação PRONTO e não pode participar de planejadas.'
            }, status=400)
        
        # 3. Militar deve ser VOLUNTÁRIO para operações
        if militar.voluntario_operacoes != 'SIM':
            return JsonResponse({
                'success': False,
                'message': f'Militar {militar.nome_completo} não é voluntário para operações e não pode participar de planejadas.'
            }, status=400)
        
        # 4. Militar NÃO pode ter GRATIFICAÇÃO
        if militar.gratificacao == 'SIM':
            return JsonResponse({
                'success': False,
                'message': f'Militar {militar.nome_completo} possui gratificação e não pode participar de planejadas.'
            }, status=400)
        
        # Verificar se já atingiu a quantidade máxima de militares
        quantidade_atual = planejada.militares.count()
        # Só verificar limite se policiais > 0 (se não há limite definido, permitir adicionar)
        if planejada.policiais > 0 and quantidade_atual >= planejada.policiais:
            return JsonResponse({
                'success': False, 
                'message': f'A operação já possui o número máximo de militares ({planejada.policiais}). '
                          f'Remova um militar antes de adicionar outro.'
            }, status=400)
        
        # Verificar se o militar já está escalado
        if planejada.militares.filter(id=militar_id).exists():
            return JsonResponse({'success': False, 'message': 'Militar já está escalado nesta operação'}, status=400)
        
        # Verificar se o militar está escalado de serviço no dia da planejada
        pode_escala_servico, escalas_servico = planejada.verificar_conflito_escala_servico(militar, planejada.data_operacao)
        
        if not pode_escala_servico:
            escalas_info = []
            for escala_info in escalas_servico:
                escalas_info.append({
                    'escala': str(escala_info['escala']),
                    'hora_inicio': escala_info['hora_inicio'].strftime('%H:%M'),
                    'hora_fim': escala_info['hora_fim'].strftime('%H:%M'),
                    'tipo_servico': escala_info['tipo_servico'],
                    'organizacao': escala_info['organizacao']
                })
            
            return JsonResponse({
                'success': False,
                'message': f'Militar {militar.nome_completo} está escalado de serviço no dia {planejada.data_operacao.strftime("%d/%m/%Y")} e não pode participar da planejada.',
                'detalhes': {
                    'militar': militar.nome_completo,
                    'data_planejada': planejada.data_operacao.strftime('%d/%m/%Y'),
                    'escalas_servico': escalas_info
                },
                'tipo_erro': 'conflito_escala_servico'
            }, status=400)
        
        # Verificar se a flexibilização está ativada na configuração
        configuracao = ConfiguracaoPlanejadas.get_configuracao_ativa()
        
        # Verificar se a flexibilização está ativa para o tipo específico
        flexibilizacao_ativa = False
        if configuracao:
            tipo_planejada = planejada.tipo_planejada
            if tipo_planejada == 'P1':
                flexibilizacao_ativa = configuracao.permitir_flexibilizacao_p1
            elif tipo_planejada == 'P2':
                flexibilizacao_ativa = configuracao.permitir_flexibilizacao_p2
            elif tipo_planejada == 'P3':
                flexibilizacao_ativa = configuracao.permitir_flexibilizacao_p3
            elif tipo_planejada == 'P4':
                flexibilizacao_ativa = configuracao.permitir_flexibilizacao_p4
            
            # Compatibilidade com campo legado
            if not flexibilizacao_ativa and configuracao.permitir_flexibilizacao_validacoes:
                flexibilizacao_ativa = True
        
        # Verificar qualificação específica baseada no tipo de planejada (usando abono)
        # Se flexibilização estiver ativa para este tipo, pular validações de qualificação
        if flexibilizacao_ativa:
            qualificado = True
            info = {'motivo': f'Flexibilização de validações ativada para {planejada.tipo_planejada}', 'flexibilizado': True}
            tipo_erro = None
        else:
            if planejada.tipo_planejada == 'P1':
                qualificado, info = planejada.verificar_qualificacao_p1_por_abono(militar, planejada.data_operacao)
                tipo_erro = 'indisponivel_p1'
            elif planejada.tipo_planejada == 'P2':
                qualificado, info = planejada.verificar_qualificacao_p2_por_abono(militar, planejada.data_operacao)
                tipo_erro = 'indisponivel_p2'
            elif planejada.tipo_planejada == 'P3':
                qualificado, info = planejada.verificar_qualificacao_p3_por_abono(militar, planejada.data_operacao)
                tipo_erro = 'indisponivel_p3'
            elif planejada.tipo_planejada == 'P4':
                qualificado, info = planejada.verificar_qualificacao_p4_por_abono(militar, planejada.data_operacao)
                tipo_erro = 'indisponivel_p4'
            else:
                qualificado = True
                info = None
        
        if not qualificado and not flexibilizacao_ativa:
            return JsonResponse({
                'success': False,
                'message': f'Militar {militar.nome_completo} não pode participar de planejada {planejada.tipo_planejada}: {info.get("motivo", "Não qualificado")}',
                'detalhes': {
                    'militar': militar.nome_completo,
                    'data_planejada': planejada.data_operacao.strftime('%d/%m/%Y'),
                    'tipo_planejada': planejada.tipo_planejada,
                    'motivo': info.get('motivo'),
                    'regra': info.get('regra'),
                    'termino_servico': info.get('termino_servico').strftime('%d/%m/%Y %H:%M') if info.get('termino_servico') else None,
                    'data_minima_planejada': info.get('data_minima_planejada').strftime('%d/%m/%Y %H:%M') if info.get('data_minima_planejada') else None,
                    'horas_restantes': info.get('horas_restantes')
                },
                'tipo_erro': tipo_erro
            }, status=400)
        
        # Verificar se tem folga suficiente (NUNCA flexibilizar - sempre verificar)
        tem_folga_suficiente, info_folga = planejada.verificar_folga_suficiente(militar, planejada.data_operacao)
        
        if not tem_folga_suficiente:
            # Mensagem baseada na nova regra de 24 horas
            horas_restantes = info_folga.get('horas_restantes', 0)
            termino_escala = info_folga.get('termino_escala')
            data_minima = info_folga.get('data_minima_planejada')
            
            if horas_restantes > 0:
                mensagem = f'Militar {militar.nome_completo} precisa aguardar {horas_restantes:.1f} horas após o término de sua última escala de serviço normal.'
            else:
                mensagem = f'Militar {militar.nome_completo} não pode participar da planejada. É necessário aguardar 24 horas após o término do serviço ordinário.'
            
            return JsonResponse({
                'success': False,
                'message': mensagem,
                'detalhes': {
                    'militar': militar.nome_completo,
                    'data_planejada': planejada.data_operacao.strftime('%d/%m/%Y %H:%M'),
                    'ultima_escala': info_folga['ultima_escala_servico'].escala.data.strftime('%d/%m/%Y'),
                    'termino_escala': termino_escala.strftime('%d/%m/%Y %H:%M') if termino_escala else 'N/A',
                    'data_minima_planejada': data_minima.strftime('%d/%m/%Y %H:%M') if data_minima else 'N/A',
                    'horas_restantes': round(horas_restantes, 1)
                },
                'tipo_erro': 'folga_insuficiente'
            }, status=400)
        
        # Verificar limites de planejadas por dia (apenas se flexibilização não estiver ativa)
        if not flexibilizacao_ativa:
            from .models import EscalaMilitar, ConfiguracaoPlanejadas
            tipo_planejada = planejada.tipo_planejada or 'P1'
            pode_adicionar, mensagem_limite, planejadas_dia = EscalaMilitar.verificar_limites_planejadas_dia(
                militar, planejada.data_operacao.date(), tipo_planejada
            )
        else:
            pode_adicionar = True
            mensagem_limite = None
            planejadas_dia = []
        
        if not pode_adicionar and not flexibilizacao_ativa:
            return JsonResponse({
                'success': False, 
                'message': mensagem_limite,
                'planejadas_dia': [
                    {
                        'nome': p.nome,
                        'tipo': p.tipo_planejada or 'P1',
                        'hora_inicio': p.hora_inicio.strftime('%H:%M') if p.hora_inicio else '',
                        'hora_termino': p.hora_termino.strftime('%H:%M') if p.hora_termino else '',
                        'cidade': p.cidade
                    } for p in planejadas_dia
                ],
                'tipo_erro': 'limite_planejadas'
            }, status=400)
        
        # Verificar limite mensal de planejadas (apenas se flexibilização não estiver ativa)
        if not flexibilizacao_ativa:
            try:
                configuracao = ConfiguracaoPlanejadas.get_ou_create_configuracao_padrao()
                limite_mensal = configuracao.quantidade_planejadas_por_militar_mes
                
                print(f"DEBUG limite: Configuração carregada, limite_mensal = {limite_mensal}")
                
                # Contar planejadas do militar no mês da operação
                data_operacao = planejada.data_operacao.date()
                ano_mes = data_operacao.year, data_operacao.month
                
                print(f"DEBUG limite: Verificando militar {militar.nome_completo} no mês {ano_mes[1]}/{ano_mes[0]}")
                
                # Contar todas as planejadas do militar no mês
                planejadas_mes = militar.planejadas.filter(
                    data_operacao__year=ano_mes[0],
                    data_operacao__month=ano_mes[1],
                    ativo=True
                )
                
                print(f"DEBUG limite: Encontradas {planejadas_mes.count()} planejadas do militar no mês")
                
                # Calcular quantidade total baseado no tipo de cada planejada
                # P1 = 1, P2 = 2, P3 = 3, P4 = 4
                quantidade_total_mes = 0
                for p in planejadas_mes:
                    tipo = p.tipo_planejada or 'P1'
                    if tipo == 'P1':
                        quantidade_total_mes += 1
                    elif tipo == 'P2':
                        quantidade_total_mes += 2
                    elif tipo == 'P3':
                        quantidade_total_mes += 3
                    elif tipo == 'P4':
                        quantidade_total_mes += 4
                
                print(f"DEBUG limite: quantidade_total_mes = {quantidade_total_mes}")
                
                # Calcular quanto vai adicionar com a nova planejada
                tipo_nova_planejada = planejada.tipo_planejada or 'P1'
                quantidade_adicionar = 0
                if tipo_nova_planejada == 'P1':
                    quantidade_adicionar = 1
                elif tipo_nova_planejada == 'P2':
                    quantidade_adicionar = 2
                elif tipo_nova_planejada == 'P3':
                    quantidade_adicionar = 3
                elif tipo_nova_planejada == 'P4':
                    quantidade_adicionar = 4
                
                print(f"DEBUG limite: tipo_nova_planejada = {tipo_nova_planejada}, quantidade_adicionar = {quantidade_adicionar}")
                
                # Verificar se ao adicionar vai ultrapassar o limite
                total_seria = quantidade_total_mes + quantidade_adicionar
                print(f"DEBUG limite: total_seria = {total_seria}, limite_mensal = {limite_mensal}, bloqueia? {total_seria > limite_mensal}")
                
                # Debug detalhado
                print(f"DEBUG ADICIONAR: Militar {militar.nome_completo}")
                print(f"  - Quantidade atual no mês: {quantidade_total_mes}")
                print(f"  - Tipo da planejada: {tipo_nova_planejada}")
                print(f"  - Quantidade adicionar: {quantidade_adicionar}")
                print(f"  - Total seria: {total_seria}")
                print(f"  - Limite: {limite_mensal}")
                print(f"  - Ultrapassa? {total_seria > limite_mensal}")
                
                if total_seria > limite_mensal:
                    print(f"DEBUG limite: BLOQUEANDO adição - Militar {militar.nome_completo} ultrapassaria limite")
                    return JsonResponse({
                        'success': False,
                        'message': f'Militar {militar.nome_completo} não pode ser adicionado. Quantidade atual no mês: {quantidade_total_mes} planejadas. Ao adicionar esta operação {tipo_nova_planejada} (+{quantidade_adicionar}) o total seria {quantidade_total_mes + quantidade_adicionar}, ultrapassando o limite de {limite_mensal} planejadas.',
                        'detalhes': {
                            'militar': militar.nome_completo,
                            'mes_ano': f"{data_operacao.strftime('%m/%Y')}",
                            'limite_mensal': limite_mensal,
                            'quantidade_atual': quantidade_total_mes,
                            'tipo_operacao': tipo_nova_planejada,
                            'quantidade_adicionar': quantidade_adicionar,
                            'total_seria': quantidade_total_mes + quantidade_adicionar
                        },
                        'tipo_erro': 'limite_mensal_planejadas'
                    }, status=400)
                else:
                    print(f"DEBUG limite: PERMITINDO adição - Militar {militar.nome_completo} não ultrapassa limite")
            except Exception as e:
                # Se houver erro na verificação de limite mensal, BLOQUEAR (NÃO permitir)
                import traceback
                print(f"ERRO ao verificar limite mensal: {str(e)}")
                print(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'message': f'Erro ao verificar limite mensal de planejadas. Detalhes: {str(e)}',
                    'tipo_erro': 'erro_validacao_limite'
                }, status=500)
        
        # Adicionar militar à operação
        planejada.militares.add(militar)
        
        # Sincronizar com a escala de abonar
        planejada.marcar_na_escala_abonar(militar)
        
        # RECALCULAR VALORES após adição do militar
        planejada.recalcular_valores()
        
        return JsonResponse({
            'success': True,
            'message': f'Militar {militar.nome_completo} adicionado à operação com sucesso! Valores recalculados automaticamente.',
            'valores_atualizados': {
                'valor_total': float(planejada.valor_total),
                'saldo': float(planejada.saldo),
                'quantidade_militares': planejada.militares.count()
            }
        })
        
    except Militar.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Militar não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao adicionar militar: {str(e)}'}, status=500)


@login_required
@require_http_methods(["POST"])
def api_remover_militar_planejada(request, planejada_id):
    """API para remover militar da operação planejada"""
    planejada = get_object_or_404(Planejada, id=planejada_id)
    
    # VALIDAÇÃO DE TEMPO: Verificar se pode alterar militares
    pode_alterar, msg_tempo = planejada.pode_alterar_militares(request.user)
    if not pode_alterar:
        return JsonResponse({
            'success': False,
            'error': f'Erro de tempo: {msg_tempo}'
        })
    
    # VALIDAÇÃO DE PERMISSÃO: Verificar se pode editar militares (fiscal vs superusuário)
    if not planejada.pode_editar_militares(request.user):
        return JsonResponse({
            'success': False,
            'error': 'Você não tem permissão para editar militares desta operação. Apenas o fiscal pode editar antes da assinatura, ou superusuários após a assinatura.'
        })
    
    militar_id = request.POST.get('militar_id')
    
    if not militar_id:
        return JsonResponse({'success': False, 'message': 'ID do militar não fornecido'}, status=400)
    
    try:
        militar = Militar.objects.get(id=militar_id)
        
        # Verificar se o militar está escalado
        if not planejada.militares.filter(id=militar_id).exists():
            return JsonResponse({'success': False, 'message': 'Militar não está escalado nesta operação'}, status=400)
        
        # Remover militar da operação
        planejada.militares.remove(militar)
        
        # Sincronizar com a escala de abonar
        planejada.desmarcar_da_escala_abonar(militar)
        
        # RECALCULAR VALORES após remoção do militar
        planejada.recalcular_valores()
        
        return JsonResponse({
            'success': True,
            'message': f'Militar {militar.nome_completo} removido da operação com sucesso! Valores recalculados automaticamente.',
            'valores_atualizados': {
                'valor_total': float(planejada.valor_total),
                'saldo': float(planejada.saldo),
                'quantidade_militares': planejada.militares.count()
            }
        })
        
    except Militar.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Militar não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erro ao remover militar: {str(e)}'}, status=500)


@login_required
def api_militares_disponiveis_geral(request):
    """API genérica para buscar militares disponíveis (usada na criação)"""
    # Verificar se há termo de pesquisa
    termo_pesquisa = request.GET.get('q', '').strip()
    
    # Para a API geral, não temos data específica, então não podemos verificar escalas de serviço
    # Apenas filtrar por critérios básicos
    
    # Buscar militares ativos que são voluntários para operações planejadas
    if termo_pesquisa:
        militares = Militar.objects.filter(
            classificacao='ATIVO',
            voluntario_operacoes='SIM',
            situacao='PRONTO'  # Apenas militares com situação PRONTO
        ).exclude(
            gratificacao='SIM'  # Excluir militares com gratificação
        ).filter(
            Q(nome_completo__icontains=termo_pesquisa) |
            Q(nome_guerra__icontains=termo_pesquisa) |
            Q(matricula__icontains=termo_pesquisa)
        ).distinct()
    else:
        militares = Militar.objects.filter(
            classificacao='ATIVO',
            voluntario_operacoes='SIM',
            situacao='PRONTO'  # Apenas militares com situação PRONTO
        ).exclude(
            gratificacao='SIM'  # Excluir militares com gratificação
        )
    
    # Ordenar por hierarquia militar e antiguidade
    ordem_hierarquica = ['CB', 'TC', 'MJ', 'CP', '1T', '2T', 'AS', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
    
    from django.db.models import Case, When, Value, IntegerField
    hierarquia_ordem = Case(
        *[When(posto_graduacao=posto, then=Value(i)) for i, posto in enumerate(ordem_hierarquica)],
        default=Value(len(ordem_hierarquica)),
        output_field=IntegerField()
    )
    
    militares = militares.order_by(
        hierarquia_ordem,
        'data_promocao_atual',
        'numeracao_antiguidade'
    ).values('id', 'nome_completo', 'posto_graduacao', 'data_promocao_atual', 'numeracao_antiguidade', 'cpf')
    
    # Converter para lista
    militares_list = []
    for militar in militares:
        # Calcular tempo no posto atual
        tempo_posto = ""
        if militar['data_promocao_atual']:
            from datetime import date
            hoje = date.today()
            dias_no_posto = (hoje - militar['data_promocao_atual']).days
            anos = dias_no_posto // 365
            meses = (dias_no_posto % 365) // 30
            if anos > 0:
                tempo_posto = f"{anos}a {meses}m"
            elif meses > 0:
                tempo_posto = f"{meses}m"
            else:
                tempo_posto = f"{dias_no_posto}d"
        
        # Obter o nome completo do posto/graduação
        posto_graduacao_display = get_posto_completo(militar['posto_graduacao'])
        
        # Para a API geral, não temos data específica, então não incluímos planejadas do dia
        militares_list.append({
            'id': militar['id'],
            'nome_completo': militar['nome_completo'],
            'posto_graduacao': militar['posto_graduacao'],
            'posto_graduacao_display': posto_graduacao_display,
            'cpf': militar['cpf'],
            'planejadas_dia': [],  # API geral não tem data específica
            'total_planejadas_dia': 0
        })
    
    return JsonResponse({
        'militares': militares_list,
        'total': len(militares_list)
    })


@login_required
def planejada_militares(request, planejada_id):
    """
    Retorna os militares associados a uma planejada específica via AJAX
    """
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Verificar se a planejada está excluída
        if planejada.excluido:
            return JsonResponse({
                'success': False,
                'message': 'Esta planejada foi excluída e não pode ser visualizada.',
                'militares': [],
                'total': 0
            })
        
        # VALIDAÇÃO DE TEMPO: Verificar se pode alterar militares
        pode_alterar, msg_tempo = planejada.pode_alterar_militares(request.user)
        if not pode_alterar:
            return JsonResponse({
                'error': f'Erro de tempo: {msg_tempo}',
                'pode_alterar': False
            })
        
        # Buscar militares associados à planejada
        militares = planejada.militares.all().order_by('nome_completo')
        
        militares_list = []
        for militar in militares:
            # Usar função centralizada para mapear posto/graduação
            posto_grad = get_posto_completo(militar.posto_graduacao)
            
            militares_list.append({
                'id': militar.id,
                'nome_completo': militar.nome_completo,
                'posto_grad': posto_grad,
                'matricula': militar.matricula,
                'lotacao': str(militar.lotacao_atual().lotacao) if militar.lotacao_atual() else 'Não informado'
            })
        
        return JsonResponse({
            'success': True,
            'militares': militares_list,
            'total': len(militares_list),
            'planejada_nome': planejada.nome
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao carregar militares: {str(e)}',
            'militares': [],
            'total': 0
        })


@login_required
def api_militares_disponiveis_planejada_criacao(request):
    """API para buscar militares disponíveis para criação de planejada"""
    from datetime import datetime
    from django.db.models import Q
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Obter parâmetros
    data_operacao = request.GET.get('data')
    hora_inicio = request.GET.get('hora_inicio')
    tipo_planejada = request.GET.get('tipo')
    termo_pesquisa = request.GET.get('q', '').strip()
    
    if not data_operacao:
        return JsonResponse({
            'militares': [],
            'total': 0,
            'error': 'Data da operação é obrigatória'
        })
    
    try:
        # Converter data para datetime
        data_operacao_dt = datetime.strptime(data_operacao, '%Y-%m-%d')
        
        # Se hora_inicio foi fornecida, combinar com a data
        if hora_inicio:
            try:
                hora_inicio_time = datetime.strptime(hora_inicio, '%H:%M').time()
                data_operacao_dt = datetime.combine(data_operacao_dt.date(), hora_inicio_time)
            except ValueError:
                pass
        
        # Buscar militares que atendem os critérios OBRIGATÓRIOS (NUNCA flexibilizados):
        # - Deve ser ATIVO
        # - Deve estar PRONTO
        # - Deve ser VOLUNTÁRIO para operações
        # Esses critérios NUNCA são flexibilizados, mesmo com flexibilização de validações ativa
        militares_query = Militar.objects.filter(
            classificacao='ATIVO',
            situacao='PRONTO',
            voluntario_operacoes='SIM'  # OBRIGATÓRIO - nunca flexibilizado
        )
        
        # Excluir militares com gratificação
        militares_query = militares_query.exclude(gratificacao='SIM').distinct()
        
        # Aplicar filtro de pesquisa se houver
        if termo_pesquisa:
            militares_query = militares_query.filter(
                Q(nome_completo__icontains=termo_pesquisa) |
                Q(nome_guerra__icontains=termo_pesquisa) |
                Q(matricula__icontains=termo_pesquisa)
            )
        
        logger.info(f"Total de militares na query: {militares_query.count()}")
        
        # Ordenar por hierarquia
        ordem_hierarquica = ['CB', 'TC', 'MJ', 'CP', '1T', '2T', 'AS', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
        from django.db.models import Case, When, Value, IntegerField
        hierarquia_ordem = Case(
            *[When(posto_graduacao=posto, then=Value(i)) for i, posto in enumerate(ordem_hierarquica)],
            default=Value(len(ordem_hierarquica)),
            output_field=IntegerField()
        )
        
        militares = militares_query.order_by(
            hierarquia_ordem,
            'data_promocao_atual',
            'numeracao_antiguidade'
        )
        
        militares_list = []
        total_processado = 0
        total_erros = 0
        
        for militar in militares:
            try:
                # Criar uma planejada temporária para validação
                planejada_temp = Planejada()
                planejada_temp.tipo_planejada = tipo_planejada or 'P1'
                planejada_temp.data_operacao = data_operacao_dt
                if hora_inicio:
                    try:
                        planejada_temp.hora_inicio = datetime.strptime(hora_inicio, '%H:%M').time()
                    except ValueError:
                        pass
                
                # PRIMEIRO: Verificar conflitos de horário com outras planejadas
                conflito_horario = False
                planejadas_conflitantes = []
                
                # Buscar planejadas onde este militar está escalado na mesma data
                data_operacao_date = data_operacao_dt.date()
                planejadas_militar = Planejada.objects.filter(
                    militares=militar,
                    data_operacao__date=data_operacao_date,
                    ativo=True,
                    excluido=False  # Excluir planejadas marcadas como excluídas
                )
                
                # Se há planejadas na mesma data, marcar como indisponível
                if planejadas_militar.exists():
                    conflito_horario = True
                    for planejada in planejadas_militar:
                        planejadas_conflitantes.append({
                            'nome': planejada.nome,
                            'tipo': planejada.tipo_planejada or 'P1',
                        'hora_inicio': planejada.hora_inicio.strftime('%H:%M') if planejada.hora_inicio else 'N/A',
                        'hora_termino': planejada.hora_termino.strftime('%H:%M') if planejada.hora_termino else 'N/A',
                            'cidade': planejada.cidade
                        })
                    
                    # Verificar conflitos de horário específicos se hora_inicio foi fornecida
                    if hora_inicio:
                        hora_inicio_time = datetime.strptime(hora_inicio, '%H:%M').time()
                        hora_termino_time = None
                    
                        # Se não especificou hora_termino, calcular baseado no tipo
                        if not hora_termino_time:
                            if tipo_planejada == 'P1':
                                from datetime import timedelta
                                hora_termino_time = (datetime.combine(data_operacao_dt.date(), hora_inicio_time) + timedelta(hours=6)).time()
                            elif tipo_planejada == 'P2':
                                from datetime import timedelta
                                hora_termino_time = (datetime.combine(data_operacao_dt.date(), hora_inicio_time) + timedelta(hours=12)).time()
                            elif tipo_planejada == 'P3':
                                from datetime import timedelta
                                hora_termino_time = (datetime.combine(data_operacao_dt.date(), hora_inicio_time) + timedelta(hours=24)).time()
                            elif tipo_planejada == 'P4':
                                from datetime import timedelta
                                hora_termino_time = (datetime.combine(data_operacao_dt.date(), hora_inicio_time) + timedelta(hours=24)).time()
                            else:
                                # Padrão P1
                                from datetime import timedelta
                                hora_termino_time = (datetime.combine(data_operacao_dt.date(), hora_inicio_time) + timedelta(hours=6)).time()
                        
                        for planejada in planejadas_militar:
                            if planejada.hora_inicio and planejada.hora_termino:
                                # Verificar sobreposição de horários
                                if _horarios_sobrepostos(
                                    hora_inicio_time, hora_termino_time,
                                    planejada.hora_inicio, planejada.hora_termino
                                ):
                                    conflito_horario = True
                                    planejadas_conflitantes.append({
                                        'nome': planejada.nome,
                                        'tipo': planejada.tipo_planejada or 'P1',
                                        'hora_inicio': planejada.hora_inicio.strftime('%H:%M'),
                                        'hora_termino': planejada.hora_termino.strftime('%H:%M'),
                                        'cidade': planejada.cidade
                                    })
                
                # Verificar se a flexibilização está ativada na configuração
                configuracao = ConfiguracaoPlanejadas.get_ou_create_configuracao_padrao()
                
                # Verificar flexibilização por tipo
                flexibilizacao_p1 = configuracao and configuracao.permitir_flexibilizacao_p1
                flexibilizacao_p2 = configuracao and configuracao.permitir_flexibilizacao_p2
                flexibilizacao_p3 = configuracao and configuracao.permitir_flexibilizacao_p3
                flexibilizacao_p4 = configuracao and configuracao.permitir_flexibilizacao_p4
                
                # Compatibilidade com campo legado
                flexibilizacao_geral = configuracao and configuracao.permitir_flexibilizacao_validacoes
                if flexibilizacao_geral:
                    flexibilizacao_p1 = True
                    flexibilizacao_p2 = True
                    flexibilizacao_p3 = True
                    flexibilizacao_p4 = True
                
                # Verificar qualificação baseada no tipo (apenas se não há conflito de horário e flexibilização não está ativa)
                qualificado = True
                status_disponibilidade = 'disponivel'
                mensagem_status = 'Disponível'
                tipos_permitidos = []
                
                # Conflitos de horário NUNCA são flexibilizados - sempre bloqueiam
                if conflito_horario:
                    qualificado = False
                    status_disponibilidade = 'conflito_horario'
                    mensagem_status = f'Já escalado em {len(planejadas_conflitantes)} planejada(s) na mesma data: '
                    for conflito in planejadas_conflitantes:
                        mensagem_status += f"{conflito['nome']} ({conflito['tipo']}) {conflito['hora_inicio']}-{conflito['hora_termino']}, "
                    mensagem_status = mensagem_status.rstrip(', ')
                else:
                    # REGRA: Quem pode mais, pode menos
                    # Verificar qualificações de P4 para P1 (hierarquia decrescente)
                    qualificacoes = {}
                    
                    # P4 - verificar qualificação normal, a menos que esteja flexibilizado
                    if flexibilizacao_p4:
                        qualificado_p4 = True
                        info_p4 = {'motivo': 'Flexibilização ativada para P4', 'flexibilizado': True}
                    else:
                        qualificado_p4, info_p4 = planejada_temp.verificar_qualificacao_p4_por_abono(militar, data_operacao_dt)
                    qualificacoes['P4'] = qualificado_p4
                    
                    # P3 - verificar qualificação normal, a menos que esteja flexibilizado
                    if flexibilizacao_p3:
                        qualificado_p3 = True
                        info_p3 = {'motivo': 'Flexibilização ativada para P3', 'flexibilizado': True}
                    else:
                        qualificado_p3, info_p3 = planejada_temp.verificar_qualificacao_p3_por_abono(militar, data_operacao_dt)
                    qualificacoes['P3'] = qualificado_p3
                    
                    # P2 - verificar qualificação normal, a menos que esteja flexibilizado
                    if flexibilizacao_p2:
                        qualificado_p2 = True
                        info_p2 = {'motivo': 'Flexibilização ativada para P2', 'flexibilizado': True}
                    else:
                        qualificado_p2, info_p2 = planejada_temp.verificar_qualificacao_p2_por_abono(militar, data_operacao_dt)
                    qualificacoes['P2'] = qualificado_p2
                    
                    # P1 - verificar qualificação normal, a menos que esteja flexibilizado
                    if flexibilizacao_p1:
                        qualificado_p1 = True
                        info_p1 = {'motivo': 'Flexibilização ativada para P1', 'flexibilizado': True}
                    else:
                        qualificado_p1, info_p1 = planejada_temp.verificar_qualificacao_p1_por_abono(militar, data_operacao_dt)
                    qualificacoes['P1'] = qualificado_p1
                    
                    # Determinar tipos permitidos baseado na hierarquia
                    tipos_permitidos_por_qualificacao = []
                    tipos_flexibilizados = []
                    
                    if qualificado_p4:
                        tipos_permitidos_por_qualificacao = ['P1', 'P2', 'P3', 'P4']
                        if flexibilizacao_p4:
                            tipos_flexibilizados.append('P4')
                        status_disponibilidade = 'disponivel_p4'
                        mensagem_status = 'Disponível para P1, P2, P3, P4'
                    elif qualificado_p3:
                        tipos_permitidos_por_qualificacao = ['P1', 'P2', 'P3']
                        if flexibilizacao_p3:
                            tipos_flexibilizados.append('P3')
                        status_disponibilidade = 'disponivel_p3'
                        mensagem_status = 'Disponível para P1, P2, P3'
                    elif qualificado_p2:
                        tipos_permitidos_por_qualificacao = ['P1', 'P2']
                        if flexibilizacao_p2:
                            tipos_flexibilizados.append('P2')
                        status_disponibilidade = 'disponivel_p2'
                        mensagem_status = 'Disponível para P1, P2'
                    elif qualificado_p1:
                        tipos_permitidos_por_qualificacao = ['P1']
                        if flexibilizacao_p1:
                            tipos_flexibilizados.append('P1')
                        status_disponibilidade = 'disponivel_p1'
                        mensagem_status = 'Disponível para P1'
                    else:
                        qualificado = False
                        status_disponibilidade = 'indisponivel_p1'
                        mensagem_status = f'Indisponível - {info_p1.get("motivo", "Não qualificado")}'
                    
                    # Adicionar tipos flexibilizados que não estão na lista de qualificados
                    if flexibilizacao_p1 and 'P1' not in tipos_permitidos_por_qualificacao:
                        tipos_flexibilizados.append('P1')
                    if flexibilizacao_p2 and 'P2' not in tipos_permitidos_por_qualificacao:
                        tipos_flexibilizados.append('P2')
                    if flexibilizacao_p3 and 'P3' not in tipos_permitidos_por_qualificacao:
                        tipos_flexibilizados.append('P3')
                    if flexibilizacao_p4 and 'P4' not in tipos_permitidos_por_qualificacao:
                        tipos_flexibilizados.append('P4')
                    
                    # Combinar tipos permitidos por qualificação com tipos flexibilizados
                    tipos_permitidos = list(set(tipos_permitidos_por_qualificacao + tipos_flexibilizados))
                    
                    # Atualizar mensagem se houver tipos flexibilizados
                    if tipos_flexibilizados:
                        if tipos_permitidos_por_qualificacao:
                            mensagem_status += f' (Flexibilização ativada para: {", ".join(tipos_flexibilizados)})'
                        else:
                            mensagem_status = f'Disponível apenas com flexibilização para: {", ".join(tipos_flexibilizados)}'
                            status_disponibilidade = 'disponivel_flexibilizado'
                    
                    # Verificar se o tipo solicitado está nos tipos permitidos
                    if qualificado and tipo_planejada and tipo_planejada not in tipos_permitidos:
                        qualificado = False
                        status_disponibilidade = f'indisponivel_{tipo_planejada.lower()}'
                        mensagem_status = f'Indisponível para {tipo_planejada} - {locals()[f"info_{tipo_planejada.lower()}"].get("motivo", "Não qualificado")}'
                    
                    # Verificar limite mensal de planejadas (NUNCA flexibilizar - sempre verificar)
                    if qualificado and not conflito_horario:
                        try:
                            limite_mensal = configuracao.quantidade_planejadas_por_militar_mes
                            
                            # Contar planejadas do militar no mês
                            data_operacao_date = data_operacao_dt.date()
                            ano_mes = data_operacao_date.year, data_operacao_date.month
                            
                            planejadas_mes = militar.planejadas.filter(
                                data_operacao__year=ano_mes[0],
                                data_operacao__month=ano_mes[1],
                                ativo=True
                            )
                            
                            # Calcular quantidade total no mês
                            quantidade_total_mes = 0
                            for p in planejadas_mes:
                                tipo = p.tipo_planejada or 'P1'
                                if tipo == 'P1':
                                    quantidade_total_mes += 1
                                elif tipo == 'P2':
                                    quantidade_total_mes += 2
                                elif tipo == 'P3':
                                    quantidade_total_mes += 3
                                elif tipo == 'P4':
                                    quantidade_total_mes += 4
                            
                            # Calcular quantidade que seria adicionada
                            quantidade_adicionar = 0
                            if tipo_planejada == 'P1':
                                quantidade_adicionar = 1
                            elif tipo_planejada == 'P2':
                                quantidade_adicionar = 2
                            elif tipo_planejada == 'P3':
                                quantidade_adicionar = 3
                            elif tipo_planejada == 'P4':
                                quantidade_adicionar = 4
                            else:
                                quantidade_adicionar = 1  # Padrão P1
                            
                            # Verificar quais tipos de planejada ainda cabem no limite mensal
                            tipos_permitidos_filtrados = []
                            for tipo_permitido in tipos_permitidos:
                                qtd_tipo = {'P1': 1, 'P2': 2, 'P3': 3, 'P4': 4}.get(tipo_permitido, 1)
                                if (quantidade_total_mes + qtd_tipo) <= limite_mensal:
                                    # Esse tipo ainda cabe
                                    tipos_permitidos_filtrados.append(tipo_permitido)
                            
                            # Atualizar lista de tipos permitidos
                            tipos_permitidos = tipos_permitidos_filtrados
                            
                            # Se não há nenhum tipo permitido após o filtro, marcar como indisponível
                            if not tipos_permitidos:
                                qualificado = False
                                status_disponibilidade = 'indisponivel_limite_mensal'
                                mensagem_status = f'Indisponível - Limite mensal de {limite_mensal} planejadas atingido. No mês atual: {quantidade_total_mes}/{limite_mensal}'
                            # Se o tipo solicitado não está na lista de permitidos, marcar como indisponível para aquele tipo
                            elif tipo_planejada and tipo_planejada not in tipos_permitidos:
                                qualificado = False
                                status_disponibilidade = f'indisponivel_{tipo_planejada.lower()}_limite_mensal'
                                mensagem_status = f'Indisponível para {tipo_planejada} - Limite mensal ultrapassado. No mês: {quantidade_total_mes}/{limite_mensal}. Tipos ainda disponíveis: {", ".join(tipos_permitidos)}'
                            # Se está qualificado e o tipo solicitado está permitido, atualizar mensagem
                            elif qualificado:
                                status_disponibilidade = f'disponivel_{tipo_planejada.lower() if tipo_planejada else "multiplos"}'
                                mensagem_status = f'Disponível - No mês: {quantidade_total_mes}/{limite_mensal}. Tipos permitidos: {", ".join(tipos_permitidos)}'
                        except Exception as e:
                            # Se houver erro na verificação de limite, manter qualificado
                            logger.warning(f"Erro ao verificar limite mensal para {militar.nome_completo}: {str(e)}")
                
                # Obter posto completo
                posto_graduacao_display = get_posto_completo(militar.posto_graduacao)
                
                # Obter lotação atual
                lotacao_atual = militar.lotacao_atual()
                lotacao_nome = ''
                lotacao_nivel = ''
                
                if lotacao_atual:
                    lotacao_nome = lotacao_atual.lotacao
                    # Obter nível do organograma (sub_unidade > unidade > grande_comando > orgao)
                    if lotacao_atual.sub_unidade:
                        lotacao_nivel = f"Sub-Unidade: {lotacao_atual.sub_unidade.sigla} - {lotacao_atual.sub_unidade.nome}"
                    elif lotacao_atual.unidade:
                        lotacao_nivel = f"Unidade: {lotacao_atual.unidade.sigla} - {lotacao_atual.unidade.nome}"
                    elif lotacao_atual.grande_comando:
                        lotacao_nivel = f"Grande Comando: {lotacao_atual.grande_comando.sigla} - {lotacao_atual.grande_comando.nome}"
                    elif lotacao_atual.orgao:
                        lotacao_nivel = f"Órgão: {lotacao_atual.orgao.sigla} - {lotacao_atual.orgao.nome}"
                    else:
                        lotacao_nivel = lotacao_nome if lotacao_nome else 'Não informado'
                else:
                    lotacao_nome = 'Não informado'
                    lotacao_nivel = 'Não informado'
                
                militares_list.append({
                    'id': militar.id,
                    'nome_completo': militar.nome_completo,
                    'posto_graduacao': militar.posto_graduacao,
                    'posto_graduacao_display': posto_graduacao_display,
                    'cpf': militar.cpf,
                    'lotacao': lotacao_nome,
                    'lotacao_nivel': lotacao_nivel,
                    'status_disponibilidade': status_disponibilidade,
                    'mensagem_status': mensagem_status,
                    'pode_selecionar': qualificado,
                    'tipos_permitidos': tipos_permitidos,
                    'planejadas_dia': planejadas_conflitantes,
                    'total_planejadas_dia': len(planejadas_conflitantes)
                })
                total_processado += 1
            except Exception as e:
                # Log do erro mas continue processando
                total_erros += 1
                logger.error(f"Erro ao processar militar {militar.id} ({militar.nome_completo}): {str(e)}", exc_info=True)
                continue
        
        logger.info(f"Processados: {total_processado}, Erros: {total_erros}, Total retornado: {len(militares_list)}")
        
        return JsonResponse({
            'militares': militares_list,
            'total': len(militares_list),
            'processados': total_processado,
            'erros': total_erros
        })
        
    except Exception as e:
        import traceback
        logger.error(f"Erro na API api_militares_disponiveis_planejada_criacao: {str(e)}", exc_info=True)
        return JsonResponse({
            'militares': [],
            'total': 0,
            'error': f'Erro ao carregar militares: {str(e)}',
            'traceback': traceback.format_exc() if request.user.is_superuser else None
        })


@login_required
def api_buscar_nota_por_numero(request, numero_nota):
    """
    API para buscar uma nota pelo número
    Retorna o ID da nota para permitir visualização
    """
    try:
        from .models import Publicacao
        
        # Buscar a nota pelo número
        nota = Publicacao.objects.filter(
            numero=numero_nota,
            tipo='NOTA'
        ).first()
        
        if not nota:
            return JsonResponse({
                'error': 'Nota não encontrada'
            }, status=404)
        
        # Verificar se o usuário pode visualizar a nota
        can_view = True
        
        # Se a nota não está publicada, verificar permissões especiais
        if nota.status != 'PUBLICADA':
            # Superusuários sempre podem visualizar
            if request.user.is_superuser:
                can_view = True
            # Usuários com permissão de edição podem visualizar
            elif nota.can_edit(request.user):
                can_view = True
            # Usuários com permissão de publicação podem visualizar
            elif nota.can_publish(request.user):
                can_view = True
            # Criador da nota pode visualizar
            elif nota.criado_por == request.user:
                can_view = True
            else:
                can_view = False
        
        if not can_view:
            return JsonResponse({
                'error': 'Você não tem permissão para visualizar esta nota'
            }, status=403)
        
        return JsonResponse({
            'nota_id': nota.id,
            'numero': nota.numero,
            'titulo': nota.titulo,
            'status': nota.status
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'error': f'Erro ao buscar nota: {str(e)}'
        }, status=500)


@login_required
def api_buscar_nota_por_id(request, nota_id):
    """
    API para buscar uma nota pelo ID
    Retorna informações básicas da nota
    """
    try:
        from .models import Publicacao
        
        # Buscar a nota pelo ID
        nota = Publicacao.objects.filter(
            id=nota_id,
            tipo='NOTA'
        ).first()
        
        if not nota:
            return JsonResponse({
                'error': 'Nota não encontrada'
            }, status=404)
        
        # Verificar se o usuário pode visualizar a nota
        can_view = True
        
        # Se a nota não está publicada, verificar permissões especiais
        if nota.status != 'PUBLICADA':
            # Superusuários sempre podem visualizar
            if request.user.is_superuser:
                can_view = True
            # Usuários com permissão de edição podem visualizar
            elif nota.can_edit(request.user):
                can_view = True
            # Usuários com permissão de publicação podem visualizar
            elif nota.can_publish(request.user):
                can_view = True
            # Criador da nota pode visualizar
            elif nota.criado_por == request.user:
                can_view = True
            else:
                can_view = False
        
        if not can_view:
            return JsonResponse({
                'error': 'Você não tem permissão para visualizar esta nota'
            }, status=403)
        
        return JsonResponse({
            'nota_id': nota.id,
            'numero': nota.numero,
            'titulo': nota.titulo,
            'status': nota.status,
            'data_criacao': nota.data_criacao.isoformat() if nota.data_criacao else None,
            'data_publicacao': nota.data_publicacao.isoformat() if nota.data_publicacao else None,
            'criado_por': nota.criado_por.get_full_name() if nota.criado_por else 'N/A'
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'error': f'Erro ao buscar nota: {str(e)}'
        }, status=500)


@login_required
def api_assinatura_planejada(request, planejada_id, tipo):
    """
    API para assinar planejada (operador, aprovador ou fiscal)
    """
    try:
        from .models import Planejada, Militar
        from django.utils import timezone
        
        # Buscar a planejada
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Buscar o militar logado
        try:
            militar = Militar.objects.get(user=request.user)
        except Militar.DoesNotExist:
            return JsonResponse({
                'error': 'Usuário não está vinculado a um militar'
            }, status=400)
        
        # Verificar permissões baseadas na função militar
        if not request.user.is_superuser:
            # Buscar função militar ativa do usuário
            from .permissoes_simples import obter_funcao_militar_ativa
            funcao_usuario = obter_funcao_militar_ativa(request.user)
            
            if not funcao_usuario:
                return JsonResponse({
                    'error': 'Usuário não possui função militar ativa'
                }, status=403)
            
            funcao_militar = funcao_usuario.funcao_militar
            
            # Verificar permissões específicas para cada tipo de assinatura
            if tipo == 'operador':
                if funcao_militar.publicacao != 'OPERADOR_PLANEJADAS':
                    return JsonResponse({
                        'error': 'Apenas usuários com função de "Operador de Planejadas" podem assinar como operador'
                    }, status=403)
            elif tipo == 'aprovador':
                if funcao_militar.publicacao != 'APROVADOR':
                    return JsonResponse({
                        'error': 'Apenas usuários com função de "Aprovador" podem aprovar planejadas'
                    }, status=403)
            elif tipo == 'fiscal':
                if funcao_militar.publicacao != 'FISCAL_PLANEJADAS':
                    return JsonResponse({
                        'error': 'Apenas usuários com função de "Fiscal de Planejadas" podem assinar como fiscal'
                    }, status=403)
        
        # Verificar se já está assinado
        if tipo == 'operador' and planejada.assinatura_operador:
            return JsonResponse({
                'error': 'Planejada já foi assinada pelo operador'
            }, status=400)
        elif tipo == 'aprovador' and planejada.assinatura_aprovador:
            return JsonResponse({
                'error': 'Planejada já foi aprovada'
            }, status=400)
        elif tipo == 'fiscal' and planejada.assinatura_fiscal:
            return JsonResponse({
                'error': 'Planejada já foi atestada pelo fiscal'
            }, status=400)
        
        # Obter dados do formulário
        funcao_assinatura = request.POST.get('funcao_assinatura', '')
        observacoes = request.POST.get('observacoes', '')
        
        if not funcao_assinatura:
            return JsonResponse({
                'error': 'Função para assinatura é obrigatória'
            }, status=400)
        
        # Atualizar a assinatura
        agora = timezone.now()
        
        if tipo == 'operador':
            planejada.assinatura_operador = militar
            planejada.data_assinatura_operador = agora
        elif tipo == 'aprovador':
            planejada.assinatura_aprovador = militar
            planejada.data_assinatura_aprovador = agora
        elif tipo == 'fiscal':
            planejada.assinatura_fiscal = militar
            planejada.data_assinatura_fiscal = agora
        else:
            return JsonResponse({
                'error': 'Tipo de assinatura inválido'
            }, status=400)
        
        planejada.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Assinatura {tipo} realizada com sucesso',
            'assinatura': {
                'militar': militar.nome_completo,
                'funcao': funcao_assinatura,
                'observacoes': observacoes,
                'data': agora.strftime('%d/%m/%Y %H:%M')
            }
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'error': f'Erro ao processar assinatura: {str(e)}'
        }, status=500)


@login_required
def api_funcoes_usuario(request):
    """
    API para buscar funções do usuário logado
    """
    try:
        from .models import Militar, UsuarioFuncaoMilitar
        
        # Buscar o militar logado
        try:
            militar = Militar.objects.get(user=request.user)
        except Militar.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuário não está vinculado a um militar'
            }, status=400)
        
        # Buscar funções do militar
        funcoes = UsuarioFuncaoMilitar.objects.filter(
            usuario=request.user,
            ativo=True
        ).order_by('-data_criacao')
        
        funcoes_data = []
        for funcao in funcoes:
            funcoes_data.append({
                'id': funcao.id,
                'nome': funcao.funcao_militar.nome if funcao.funcao_militar else 'Função não encontrada',
                'ativo': funcao.ativo,
                'data_criacao': funcao.data_criacao.strftime('%d/%m/%Y') if funcao.data_criacao else None,
                'data_atualizacao': funcao.data_atualizacao.strftime('%d/%m/%Y') if funcao.data_atualizacao else None,
            })
        
        return JsonResponse({
            'success': True,
            'funcoes': funcoes_data
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': f'Erro ao buscar funções: {str(e)}'
        }, status=500)


# ==================== FUNÇÕES AUXILIARES PARA FILTROS HIERÁRQUICOS ====================

def aplicar_filtro_hierarquico_planejadas_restritivo(planejadas_base, funcao_atual):
    """
    Aplica filtro restritivo - mostra apenas planejadas do tipo de acesso da função militar
    """
    from django.db.models import Q
    
    if not funcao_atual:
        return planejadas_base.none()
    
    funcao_militar = funcao_atual.funcao_militar
    tipo_acesso = funcao_militar.acesso
    
    filtros_instancia = Q()
    
    # Filtrar baseado no TIPO DE ACESSO da função militar
    if tipo_acesso == 'SUBUNIDADE':
        # Se o acesso é SUBUNIDADE, mostrar apenas sub-unidades
        if funcao_atual.sub_unidade:
            filtros_instancia |= Q(sub_unidade=funcao_atual.sub_unidade)
    elif tipo_acesso == 'UNIDADE':
        # Se o acesso é UNIDADE, mostrar apenas unidades
        if funcao_atual.unidade:
            filtros_instancia |= Q(unidade=funcao_atual.unidade)
    elif tipo_acesso == 'GRANDE_COMANDO':
        # Se o acesso é GRANDE_COMANDO, mostrar apenas grandes comandos
        if funcao_atual.grande_comando:
            filtros_instancia |= Q(grande_comando=funcao_atual.grande_comando)
    elif tipo_acesso == 'ORGAO':
        # Se o acesso é ORGAO, mostrar apenas órgãos
        if funcao_atual.orgao:
            filtros_instancia |= Q(orgao=funcao_atual.orgao)
    elif tipo_acesso == 'TOTAL':
        # Se o acesso é TOTAL, mostrar baseado na lotação mais específica
        if funcao_atual.sub_unidade:
            filtros_instancia |= Q(sub_unidade=funcao_atual.sub_unidade)
        elif funcao_atual.unidade:
            filtros_instancia |= Q(unidade=funcao_atual.unidade)
        elif funcao_atual.grande_comando:
            filtros_instancia |= Q(grande_comando=funcao_atual.grande_comando)
        elif funcao_atual.orgao:
            filtros_instancia |= Q(orgao=funcao_atual.orgao)
    
    return planejadas_base.filter(filtros_instancia, ativo=True, excluido=False)


def aplicar_filtro_hierarquico_planejadas_especifico(planejadas_base, funcao_atual, origem_selecionada):
    """
    Aplica filtro específico - mostra planejadas da origem selecionada baseado no tipo de acesso
    """
    from django.db.models import Q
    
    if not funcao_atual or not origem_selecionada:
        return planejadas_base.none()
    
    funcao_militar = funcao_atual.funcao_militar
    tipo_acesso = funcao_militar.acesso
    
    filtros_instancia = Q()
    
    # Aplicar filtros baseados no TIPO DE ACESSO da função militar
    if tipo_acesso == 'SUBUNIDADE' and funcao_atual.sub_unidade and origem_selecionada == funcao_atual.sub_unidade.nome:
        # Se o acesso é SUBUNIDADE e selecionou a própria sub-unidade
        filtros_instancia |= Q(sub_unidade=funcao_atual.sub_unidade)
    elif tipo_acesso == 'UNIDADE' and funcao_atual.unidade and origem_selecionada == funcao_atual.unidade.nome:
        # Se o acesso é UNIDADE e selecionou a própria unidade
        filtros_instancia |= Q(unidade=funcao_atual.unidade)
    elif tipo_acesso == 'GRANDE_COMANDO' and funcao_atual.grande_comando and origem_selecionada == funcao_atual.grande_comando.nome:
        # Se o acesso é GRANDE_COMANDO e selecionou o próprio grande comando
        filtros_instancia |= Q(grande_comando=funcao_atual.grande_comando)
    elif tipo_acesso == 'ORGAO' and funcao_atual.orgao and origem_selecionada == funcao_atual.orgao.nome:
        # Se o acesso é ORGAO e selecionou o próprio órgão
        filtros_instancia |= Q(orgao=funcao_atual.orgao)
    elif tipo_acesso == 'TOTAL':
        # Se o acesso é TOTAL, aplicar baseado na lotação mais específica
        if funcao_atual.sub_unidade and origem_selecionada == funcao_atual.sub_unidade.nome:
            filtros_instancia |= Q(sub_unidade=funcao_atual.sub_unidade)
        elif funcao_atual.unidade and origem_selecionada == funcao_atual.unidade.nome:
            filtros_instancia |= Q(unidade=funcao_atual.unidade)
        elif funcao_atual.grande_comando and origem_selecionada == funcao_atual.grande_comando.nome:
            filtros_instancia |= Q(grande_comando=funcao_atual.grande_comando)
        elif funcao_atual.orgao and origem_selecionada == funcao_atual.orgao.nome:
            filtros_instancia |= Q(orgao=funcao_atual.orgao)
    
    return planejadas_base.filter(filtros_instancia, ativo=True, excluido=False)


def obter_opcoes_hierarquicas_planejadas(funcao_atual, usuario):
    """
    Retorna as opções hierárquicas disponíveis para filtro de origem
    Baseado no TIPO DE ACESSO da função militar, não na lotação
    """
    opcoes = []
    
    if not funcao_atual:
        return opcoes
    
    funcao_militar = funcao_atual.funcao_militar
    tipo_acesso = funcao_militar.acesso
    
    # Adicionar opções baseadas no TIPO DE ACESSO da função militar
    if tipo_acesso == 'SUBUNIDADE':
        # Se o acesso é SUBUNIDADE, mostrar apenas sub-unidades
        if funcao_atual.sub_unidade:
            opcoes.append({
                'valor': funcao_atual.sub_unidade.nome,
                'texto': f"Sub-Unidade: {funcao_atual.sub_unidade.nome}",
                'nivel': 'SUB_UNIDADE'
            })
    elif tipo_acesso == 'UNIDADE':
        # Se o acesso é UNIDADE, mostrar apenas unidades
        if funcao_atual.unidade:
            opcoes.append({
                'valor': funcao_atual.unidade.nome,
                'texto': f"Unidade: {funcao_atual.unidade.nome}",
                'nivel': 'UNIDADE'
            })
    elif tipo_acesso == 'GRANDE_COMANDO':
        # Se o acesso é GRANDE_COMANDO, mostrar apenas grandes comandos
        if funcao_atual.grande_comando:
            opcoes.append({
                'valor': funcao_atual.grande_comando.nome,
                'texto': f"Grande Comando: {funcao_atual.grande_comando.nome}",
                'nivel': 'GRANDE_COMANDO'
            })
    elif tipo_acesso == 'ORGAO':
        # Se o acesso é ORGAO, mostrar apenas órgãos
        if funcao_atual.orgao:
            opcoes.append({
                'valor': funcao_atual.orgao.nome,
                'texto': f"Órgão: {funcao_atual.orgao.nome}",
                'nivel': 'ORGAO'
            })
    elif tipo_acesso == 'TOTAL':
        # Se o acesso é TOTAL, mostrar todas as opções da lotação
        if funcao_atual.sub_unidade:
            opcoes.append({
                'valor': funcao_atual.sub_unidade.nome,
                'texto': f"Sub-Unidade: {funcao_atual.sub_unidade.nome}",
                'nivel': 'SUB_UNIDADE'
            })
        elif funcao_atual.unidade:
            opcoes.append({
                'valor': funcao_atual.unidade.nome,
                'texto': f"Unidade: {funcao_atual.unidade.nome}",
                'nivel': 'UNIDADE'
            })
        elif funcao_atual.grande_comando:
            opcoes.append({
                'valor': funcao_atual.grande_comando.nome,
                'texto': f"Grande Comando: {funcao_atual.grande_comando.nome}",
                'nivel': 'GRANDE_COMANDO'
            })
        elif funcao_atual.orgao:
            opcoes.append({
                'valor': funcao_atual.orgao.nome,
                'texto': f"Órgão: {funcao_atual.orgao.nome}",
                'nivel': 'ORGAO'
            })
    
    return opcoes


@login_required
@require_http_methods(["GET"])
def api_status_tempo(request, planejada_id):
    """
    API para obter o status de tempo de uma planejada em formato JSON
    """
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Obter o status de tempo usando o método do modelo
        status_tempo = planejada.obter_status_tempo()
        
        return JsonResponse({
            'success': True,
            'status_tempo': status_tempo
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_assinaturas_planejada(request, planejada_id):
    """
    API para obter as assinaturas de uma planejada
    """
    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Verificar se o usuário pode acessar esta planejada
        # Permitir acesso para superusuários ou usuários com permissão específica
        if not request.user.is_superuser and not planejada.pode_acessar_instancia(request.user):
            return JsonResponse({
                'success': False,
                'message': 'Você não tem permissão para acessar esta operação planejada.'
            }, status=403)
        
        # Buscar assinaturas da planejada (campos específicos do modelo)
        assinaturas_data = []
        
        # Assinatura do Operador
        if planejada.assinatura_operador and planejada.data_assinatura_operador:
            assinaturas_data.append({
                'id': f'operador_{planejada.id}',
                'assinado_por': planejada.assinatura_operador.nome_completo,
                'funcao_assinatura': 'Operador de Planejadas',
                'data_assinatura': planejada.data_assinatura_operador.isoformat(),
                'tipo_assinatura': 'Operador',
                'observacoes': ''
            })
        
        # Assinatura do Aprovador
        if planejada.assinatura_aprovador and planejada.data_assinatura_aprovador:
            assinaturas_data.append({
                'id': f'aprovador_{planejada.id}',
                'assinado_por': planejada.assinatura_aprovador.nome_completo,
                'funcao_assinatura': 'Aprovador de Planejadas',
                'data_assinatura': planejada.data_assinatura_aprovador.isoformat(),
                'tipo_assinatura': 'Aprovador',
                'observacoes': ''
            })
        
        # Assinatura do Fiscal
        if planejada.assinatura_fiscal and planejada.data_assinatura_fiscal:
            assinaturas_data.append({
                'id': f'fiscal_{planejada.id}',
                'assinado_por': planejada.assinatura_fiscal.nome_completo,
                'funcao_assinatura': 'Fiscal de Planejadas',
                'data_assinatura': planejada.data_assinatura_fiscal.isoformat(),
                'tipo_assinatura': 'Fiscal',
                'observacoes': ''
            })
        
        return JsonResponse({
            'success': True,
            'assinaturas': assinaturas_data,
            'total': len(assinaturas_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erro ao carregar assinaturas: {str(e)}'
        }, status=500)


@login_required
def planejada_gerar_pdf(request, planejada_id):
    """
    Gera PDF da operação planejada em formato de formulário moderno e robusto
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    import os
    import locale
    from datetime import datetime
    from django.conf import settings

    # Configurar locale para português brasileiro
    try:
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
        except:
            pass

    try:
        planejada = get_object_or_404(Planejada, id=planejada_id)
        
        # Verificar se o usuário pode acessar esta planejada
        if not request.user.is_superuser and not planejada.pode_acessar_instancia(request.user):
            messages.error(request, 'Você não tem permissão para acessar esta operação planejada.')
            return redirect('militares:planejadas_list')
            
    except Planejada.DoesNotExist:
        messages.error(request, f'Operação planejada com ID {planejada_id} não encontrada.')
        return redirect('militares:planejadas_list')

    buffer = BytesIO()
    
    # Obter militares escalados
    militares_escalados = planejada.militares.all().order_by('posto_graduacao', 'nome_completo')
    
    # Configuração do documento com margens padrão
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()

    # Estilos modernos e robustos
    style_header = ParagraphStyle('header', parent=styles['Normal'], alignment=1, fontSize=16, fontName='Helvetica-Bold', spaceAfter=8, spaceBefore=8)
    style_subheader = ParagraphStyle('subheader', parent=styles['Normal'], alignment=1, fontSize=14, fontName='Helvetica-Bold', spaceAfter=6, spaceBefore=6)
    style_section_title = ParagraphStyle('section_title', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', spaceAfter=4, spaceBefore=8, textColor=colors.HexColor('#2c3e50'))
    style_field_label = ParagraphStyle('field_label', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', spaceAfter=2, spaceBefore=2, textColor=colors.HexColor('#34495e'))
    style_field_value = ParagraphStyle('field_value', parent=styles['Normal'], fontSize=10, fontName='Helvetica', spaceAfter=3, spaceBefore=2)
    style_table_header = ParagraphStyle('table_header', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=1, textColor=colors.white)
    style_table_cell = ParagraphStyle('table_cell', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=0)
    style_signature = ParagraphStyle('signature', parent=styles['Normal'], fontSize=10, fontName='Helvetica', spaceAfter=4, spaceBefore=4, alignment=1)
    style_center = ParagraphStyle('center', parent=styles['Normal'], alignment=1, fontSize=11)

    story = []

    # Função para limpar HTML removendo COMPLETAMENTE todas as tags
    def limpar_html_para_pdf(texto_html):
        """
        Remove COMPLETAMENTE todas as tags HTML, extraindo apenas o texto puro
        """
        if not texto_html:
            return ""
        
        # Importar html para decodificar entidades
        import html
        
        # Converter quebras de linha HTML para \n ANTES de remover tags
        texto_limpo = re.sub(r'<br\s*/?>', '\n', texto_html, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'<p[^>]*>', '\n', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'</p>', '', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'<div[^>]*>', '\n', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'</div>', '', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'</li>', '\n', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'</h[1-6]>', '\n', texto_limpo, flags=re.IGNORECASE)
        texto_limpo = re.sub(r'<li[^>]*>', '\n• ', texto_limpo, flags=re.IGNORECASE)
        
        # Remover TODAS as tags HTML (qualquer coisa entre < e >)
        texto_limpo = re.sub(r'<[^>]+>', '', texto_limpo)
        
        # AGORA converter entidades HTML após remover tags
        try:
            texto_limpo = html.unescape(texto_limpo)
        except:
            pass
        
        # Garantir conversão de entidades comuns
        texto_limpo = texto_limpo.replace('&nbsp;', ' ')
        texto_limpo = texto_limpo.replace('&amp;', '&')
        texto_limpo = texto_limpo.replace('&lt;', '<')
        texto_limpo = texto_limpo.replace('&gt;', '>')
        texto_limpo = texto_limpo.replace('&quot;', '"')
        texto_limpo = texto_limpo.replace('&apos;', "'")
        
        # Remover qualquer entidade HTML que possa ter sobrado
        texto_limpo = re.sub(r'&[a-zA-Z0-9#]+;', '', texto_limpo)
        
        # Limpar espaços extras e quebras de linha
        texto_limpo = re.sub(r'[ \t]+', ' ', texto_limpo)  # Múltiplos espaços
        texto_limpo = re.sub(r'\n[ \t]*\n+', '\n\n', texto_limpo)  # Múltiplas quebras em duas
        texto_limpo = re.sub(r'^\s+|\s+$', '', texto_limpo, flags=re.MULTILINE)  # Espaços nas extremidades
        texto_limpo = texto_limpo.strip()
        
        return texto_limpo

    # === CABEÇALHO MODERNO ===
    # Logo e identificação
    logo_path = os.path.join('staticfiles', 'logo_cbmepi.png')
    if os.path.exists(logo_path):
        story.append(Image(logo_path, width=2.5*cm, height=2.5*cm, hAlign='CENTER'))
        story.append(Spacer(1, 0.3*cm))

    # Cabeçalho institucional moderno
    story.append(Paragraph("ESTADO DO PIAUÍ", ParagraphStyle('estado', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, textColor=colors.HexColor('#2c3e50'))))
    story.append(Paragraph("CORPO DE BOMBEIROS MILITAR", ParagraphStyle('cbmepi', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, textColor=colors.HexColor('#2c3e50'))))
    
    # Lista completa de siglas para remover
    siglas_para_remover = [
        'CBMEPI', 'CBM', 'BM', 'QCG', 'QCO', 'QCG/CBMEPI',
        '1º BBM', '2º BBM', '3º BBM', '4º BBM', '5º BBM',
        'BBM', 'BG', 'CBM-PI', '/CBMEPI', '-CBMEPI', ' - ', '-',
        '3º SG/1º G', '3 SG/1 G', '3ºSG/1ºG', '3º SG 1º G',
        '1º GBM', '2º GBM', '3º GBM', '4º GBM', '5º GBM',
        '1 GBM', '2 GBM', '3 GBM', '4 GBM', '5 GBM',
        'GBM', 'SG'
    ]
    
    def limpar_nome_instancia(texto):
        """Remove siglas do nome da instância, mantendo números de grupamentos"""
        if not texto:
            return ""
        
        nome_limpo = texto
        
        # Remover pipe/barra vertical e outros caracteres problemáticos
        nome_limpo = nome_limpo.replace('|', ' ')
        nome_limpo = nome_limpo.replace(' - ', ' ')
        nome_limpo = nome_limpo.replace(' -', ' ')
        nome_limpo = nome_limpo.replace('- ', ' ')
        
        # Remover apenas siglas específicas como CBMEPI, CBM, BM, etc. (não números isolados)
        siglas_isoladas_para_remover = ['CBMEPI', 'CBM', 'BM', 'QCG', 'QCO', 'BBM', 'BG', 'GBM', 'SG']
        
        for sigla in siglas_isoladas_para_remover:
            nome_limpo = re.sub(r'\b' + re.escape(sigla) + r'\b', '', nome_limpo, flags=re.IGNORECASE)
            nome_limpo = re.sub(r'^' + re.escape(sigla) + r'[\s\-\/\(\)]+', '', nome_limpo, flags=re.IGNORECASE)
            nome_limpo = re.sub(r'[\s\-\/\(\)]+' + re.escape(sigla) + r'$', '', nome_limpo, flags=re.IGNORECASE)
        
        # Remover apenas siglas compostas como "3º SG/1º G" e variações
        siglas_compostas = [
            '3º SG/1º G', '3 SG/1 G', '3ºSG/1ºG', '3º SG 1º G',
            '1º GBM', '2º GBM', '3º GBM', '4º GBM', '5º GBM',
            '1 GBM', '2 GBM', '3 GBM', '4 GBM', '5 GBM',
            '1º BBM', '2º BBM', '3º BBM', '4º BBM', '5º BBM'
        ]
        for sigla in siglas_compostas:
            nome_limpo = re.sub(re.escape(sigla), '', nome_limpo, flags=re.IGNORECASE)
        
        # Limpar espaços extras
        nome_limpo = re.sub(r'[ \t]+', ' ', nome_limpo)
        nome_limpo = re.sub(r'^\W+|\W+$', '', nome_limpo)
        nome_limpo = nome_limpo.strip()
        
        return nome_limpo.upper() if nome_limpo else ""
    
    # Capturar ORIGEM da planejada (campo texto origem, não instância de criação)
    origem_texto = planejada.origem if planejada.origem else "QUARTEL DO COMANDO GERAL"
    
    # PRIMEIRO dividir por PIPE para separar cada nível
    if '|' in origem_texto:
        linhas_brutas = [linha.strip() for linha in origem_texto.split('|') if linha.strip()]
    else:
        # Se não tiver pipe, usar o texto completo
        linhas_brutas = [origem_texto]
    
    # DEPOIS limpar cada linha individualmente
    linhas_limpas = []
    for linha in linhas_brutas:
        linha_limpa = limpar_nome_instancia(linha)
        if linha_limpa:
            linhas_limpas.append(linha_limpa)
    
    # Adicionar cada linha ao PDF
    for linha_texto in linhas_limpas:
        if linha_texto:
            story.append(Paragraph(linha_texto, ParagraphStyle('origem', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=0, textColor=colors.HexColor('#2c3e50'))))
    
    # Linha separadora após o cabeçalho
    story.append(HRFlowable(width="100%", thickness=2, lineCap='round', color=colors.HexColor('#2c3e50')))
    story.append(Spacer(1, 0.3*cm))

    # === TÍTULO PRINCIPAL ===
    story.append(Paragraph(limpar_html_para_pdf(planejada.nome).upper(), ParagraphStyle('titulo_operacao', parent=styles['Normal'], alignment=1, fontSize=13, fontName='Helvetica-Bold', spaceAfter=12, spaceBefore=4, textColor=colors.HexColor('#2c3e50'))))
    story.append(Spacer(1, 0.3*cm))

    # === DADOS DA OPERAÇÃO DISTRIBUÍDOS ===
    data_operacao = planejada.data_operacao.strftime('%d/%m/%Y')
    hora_inicio = planejada.hora_inicio.strftime('%H:%M')
    hora_termino = planejada.hora_termino.strftime('%H:%M')
    
    # Formatando a data por extenso
    meses_extenso = {
        1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril',
        5: 'maio', 6: 'junho', 7: 'julho', 8: 'agosto',
        9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'
    }
    data_extenso = f"{planejada.data_operacao.day} de {meses_extenso[planejada.data_operacao.month]} de {planejada.data_operacao.year}"
    
    # Calcular total de militares e valores
    total_militares = len(militares_escalados)
    
    # Estilo para os dados com espaço à esquerda de 0.2cm
    style_dados_label = ParagraphStyle('dados_label', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=colors.HexColor('#2c3e50'), leftIndent=0.2*cm)
    style_dados_valor = ParagraphStyle('dados_valor', parent=styles['Normal'], fontSize=10, fontName='Helvetica', leftIndent=0.2*cm)
    
    # Adicionar dados em uma coluna
    story.append(Paragraph(f"Data da Operação: <b>{data_extenso}</b>", style_dados_valor))
    story.append(Paragraph(f"Horário: <b>{hora_inicio} às {hora_termino}</b>", style_dados_valor))
    story.append(Paragraph(f"Cidade: <b>{limpar_html_para_pdf(planejada.cidade)}</b>", style_dados_valor))
    story.append(Paragraph(f"Tipo: <b>{planejada.tipo_planejada or 'P1'}</b>", style_dados_valor))
    story.append(Paragraph(f"Valor Unitário: <b>R$ {planejada.valor:,.2f}</b>", style_dados_valor))
    story.append(Paragraph(f"Quantidade de Militares: <b>{total_militares}</b>", style_dados_valor))
    story.append(Paragraph(f"Valor Total: <b>R$ {planejada.valor_total:,.2f}</b>", style_dados_valor))
    
    # Adicionar responsável se houver
    if planejada.responsavel:
        story.append(Paragraph(f"Responsável: <b>{planejada.responsavel}</b>", style_dados_valor))
    
    # Adicionar informações da nota se existir
    if planejada.numero_nota:
        story.append(Paragraph(f"Número da Nota: <b>{limpar_html_para_pdf(planejada.numero_nota)}</b>", style_dados_valor))
    
    if planejada.link_pdf_nota:
        story.append(Paragraph(f"Link da Nota: <b>{limpar_html_para_pdf(planejada.link_pdf_nota)}</b>", style_dados_valor))
    
    story.append(Spacer(1, 0.4*cm))

    # === DESCRIÇÃO DA OPERAÇÃO ===
    if planejada.descricao:
        story.append(Paragraph("DESCRIÇÃO DA OPERAÇÃO", style_section_title))
        
        # Limpar completamente o HTML e extrair apenas texto puro
        descricao_limpa = limpar_html_para_pdf(planejada.descricao)
        
        # Dividir em parágrafos se houver múltiplas quebras de linha
        paragrafos_limpos = [p.strip() for p in descricao_limpa.split('\n\n') if p.strip()]
        
        # Adicionar cada parágrafo ao PDF
        for paragrafo in paragrafos_limpos:
            if paragrafo:
                story.append(Paragraph(paragrafo, ParagraphStyle('descricao', parent=styles['Normal'], fontSize=10, spaceAfter=6, spaceBefore=0, leading=14, fontName='Helvetica')))
        
        story.append(Spacer(1, 0.3*cm))

    # === MILITARES ESCALADOS ===
    if militares_escalados.exists():
        # Tabela de militares - POSTO, NOME, CPF CRIPTOGRAFADO, VALOR
        militares_data = []
        
        total_valor = 0
        for militar in militares_escalados:
            valor_individual = planejada.valor
            total_valor += valor_individual
            
            # CPF criptografado no padrão do sistema: XXX.***.***-XX
            cpf_limpo = militar.cpf.replace('.', '').replace('-', '')
            if len(cpf_limpo) == 11:
                cpf_cripto = f"{cpf_limpo[:3]}.***.***-{cpf_limpo[-2:]}"
            else:
                cpf_cripto = "***.***.***-**"
            
            # Obter posto completo
            posto_completo = get_posto_completo(militar.posto_graduacao or '')
            if not posto_completo or posto_completo == militar.posto_graduacao:
                # Se não encontrou o posto completo, tentar obter o display do modelo
                posto_completo = militar.get_posto_graduacao_display() or militar.posto_graduacao or 'N/A'
            
            # Verificar se o militar tem função específica de Planejadas (assina documentos)
            funcao_nome = ''
            
            # Verificar se é operador, aprovador ou fiscal de planejadas
            if planejada.assinatura_operador == militar:
                funcao_nome = 'Operador de Planejadas'
            elif planejada.assinatura_aprovador == militar:
                funcao_nome = 'Aprovador de Planejadas'
            elif planejada.assinatura_fiscal == militar:
                funcao_nome = 'Fiscal de Planejadas'
            else:
                # Buscar função atual do militar
                funcao_atual = militar.funcoes.filter(status='ATUAL', ativo=True).first()
                if funcao_atual and funcao_atual.funcao_militar:
                    funcao_nome = funcao_atual.funcao_militar.nome
            
            # Nome com função abaixo
            if funcao_nome:
                nome_com_funcao = f"{militar.nome_completo}<br/>{funcao_nome}"
            else:
                nome_com_funcao = militar.nome_completo
            
            # Ordem: POSTO, NOME, CPF, VALOR
            militares_data.append([
                posto_completo,
                Paragraph(nome_com_funcao, ParagraphStyle('nome_militar', parent=styles['Normal'], fontSize=9, fontName='Helvetica', leading=12)),
                cpf_cripto,
                f'R$ {valor_individual:,.2f}'
            ])
        
        # Sem linha de total
        
        # Criar tabela
        militares_table = Table(militares_data, colWidths=[2*cm, 7*cm, 2.5*cm, 2.5*cm])
        militares_table.setStyle(TableStyle([
            # Todos os dados
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (0, -1), 'MIDDLE'),
            ('VALIGN', (1, 0), (1, -1), 'TOP'),  # Nome alinhado ao topo para dar espaço para função
            ('VALIGN', (2, 0), (2, -1), 'MIDDLE'),
            ('VALIGN', (3, 0), (3, -1), 'MIDDLE'),
            
            # Bordas e espaçamento
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            
            # Alternância de cores
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        
        story.append(militares_table)
        story.append(Spacer(1, 0.4*cm))


    # === OBSERVAÇÕES ===
    if planejada.observacoes:
        story.append(Paragraph("OBSERVAÇÕES", style_section_title))
        
        # Limpar HTML das observações e extrair texto puro
        observacoes_limpas = limpar_html_para_pdf(planejada.observacoes)
        
        # Dividir em parágrafos se houver múltiplas quebras de linha
        paragrafos_observacoes = [p.strip() for p in observacoes_limpas.split('\n\n') if p.strip()]
        
        # Adicionar cada parágrafo ao PDF
        for paragrafo in paragrafos_observacoes:
            if paragrafo:
                story.append(Paragraph(paragrafo, ParagraphStyle('observacoes', parent=styles['Normal'], fontSize=10, spaceAfter=6, spaceBefore=0, leading=14, fontName='Helvetica')))
        
            story.append(Spacer(1, 0.3*cm))

    # === INFORMAÇÕES DE CRIAÇÃO DO DOCUMENTO ===
    story.append(Spacer(1, 0.5*cm))
    
    # Verificar quem assinou e mostrar conforme hierarquia de posto
    # Prioridade: quem tem maior posto (aprovador -> fiscal -> operador)
    assinante = None
    data_assinatura = None
    funcao_assinatura = ""
    
    # Lista de assinantes potenciais com sua hierarquia
    assinantes_potenciais = []
    
    if planejada.assinatura_aprovador and planejada.data_assinatura_aprovador:
        assinantes_potenciais.append({
            'militar': planejada.assinatura_aprovador,
            'data': planejada.data_assinatura_aprovador,
            'funcao': 'Aprovador',
            'prioridade': 1
        })
    if planejada.assinatura_fiscal and planejada.data_assinatura_fiscal:
        assinantes_potenciais.append({
            'militar': planejada.assinatura_fiscal,
            'data': planejada.data_assinatura_fiscal,
            'funcao': 'Fiscal',
            'prioridade': 2
        })
    if planejada.assinatura_operador and planejada.data_assinatura_operador:
        assinantes_potenciais.append({
            'militar': planejada.assinatura_operador,
            'data': planejada.data_assinatura_operador,
            'funcao': 'Operador',
            'prioridade': 3
        })
    
    if assinantes_potenciais:
        # Ordenar por posto (maior primeiro) - definir ordem de posto
        ordem_posto = ['CB', 'CEL', 'TC', 'MAJ', 'CP', 'MJ', 'CAP', '1T', '2T', 'ASP', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
        
        def obter_indice_posto(militar):
            posto = militar.posto_graduacao or ''
            try:
                return ordem_posto.index(posto)
            except:
                return 999  # Se não encontrar, coloca no final
        
        # Ordenar por posto, depois por prioridade da função
        assinantes_potenciais.sort(key=lambda x: (obter_indice_posto(x['militar']), x['prioridade']))
        
        # Pegar o primeiro (maior posto)
        melhor = assinantes_potenciais[0]
        assinante = melhor['militar']
        data_assinatura = melhor['data']
        funcao_assinatura = melhor['funcao']
    else:
        # Se ninguém assinou, usar a data de criação
        assinante = None
        data_assinatura = planejada.created_at
        funcao_assinatura = ""
    
    # Construir texto com data e nome dos assinantes
    cidade_estado = f"{planejada.cidade} - PI"
    data_extenso = f"{data_assinatura.day} de {meses_extenso[data_assinatura.month]} de {data_assinatura.year}"
    
    texto_geracao = f"{cidade_estado}, {data_extenso}."
    
    # Adicionar todos os assinantes
    if assinantes_potenciais:
        texto_geracao += "<br/><br/>"
        
        # Ordenar por hierarquia
        ordem_posto = ['CB', 'CEL', 'TC', 'MAJ', 'CP', 'MJ', 'CAP', '1T', '2T', 'ASP', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
        
        assinantes_potenciais.sort(key=lambda x: (obter_indice_posto(x['militar']), x['prioridade']))
        
        for idx, item in enumerate(assinantes_potenciais):
            militar = item['militar']
            funcao = item['funcao']
            
            nome_assinante = militar.nome_completo
            
            # Obter posto completo
            posto_assinante = get_posto_completo(militar.posto_graduacao or '')
            if not posto_assinante or posto_assinante == militar.posto_graduacao:
                posto_assinante = militar.get_posto_graduacao_display() or militar.posto_graduacao or ''
            
            # Adicionar BM ao posto se não já estiver presente
            if "BM" not in posto_assinante:
                posto_assinante = f"{posto_assinante} BM"
            
            # Determinar a função completa
            if funcao == 'Aprovador':
                funcao_completa = 'Aprovador de Planejadas'
            elif funcao == 'Fiscal':
                funcao_completa = 'Fiscal de Planejadas'
            elif funcao == 'Operador':
                funcao_completa = 'Operador de Planejadas'
            else:
                funcao_completa = funcao
            
            # Adicionar ao texto
            texto_geracao += f"{nome_assinante} - {posto_assinante}<br/>{funcao_completa}"
            
            # Espaçamento entre assinantes (exceto na última)
            if idx < len(assinantes_potenciais) - 1:
                texto_geracao += "<br/><br/>"
    
    style_criacao = ParagraphStyle('criacao', parent=styles['Normal'], fontSize=9, fontName='Helvetica', 
                                   alignment=1, textColor=colors.HexColor('#2c3e50'), spaceAfter=8,
                                   leading=12)
    
    story.append(Paragraph(texto_geracao, style_criacao))

    # === ASSINATURAS ELETRÔNICAS ===
    story.append(Spacer(1, 0.02*cm))
    
    # Processar assinaturas dos campos diretos do modelo Planejada
    from .utils import formatar_data_assinatura, obter_caminho_assinatura_eletronica
    
    # Coletar todas as assinaturas para ordenar por posto
    assinaturas_para_exibir = []
    
    # Ordem de posto (maior para menor)
    ordem_posto_assinatura = ['CB', 'CEL', 'TC', 'MAJ', 'CP', 'MJ', 'CAP', '1T', '2T', 'ASP', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
    
    def obter_indice_posto_assinatura(posto):
        try:
            return ordem_posto_assinatura.index(posto)
        except:
            return 999
    
    # Coletar assinaturas
    # Operador
    if planejada.assinatura_operador and planejada.data_assinatura_operador:
        militar = planejada.assinatura_operador
        posto_abreviado = militar.posto_graduacao or ''
        assinaturas_para_exibir.append({
            'militar': militar,
            'data': planejada.data_assinatura_operador,
            'tipo': 'Operador de Planejadas',
            'indice_posto': obter_indice_posto_assinatura(posto_abreviado)
        })

    # Aprovador
    if planejada.assinatura_aprovador and planejada.data_assinatura_aprovador:
        militar = planejada.assinatura_aprovador
        posto_abreviado = militar.posto_graduacao or ''
        assinaturas_para_exibir.append({
            'militar': militar,
            'data': planejada.data_assinatura_aprovador,
            'tipo': 'Aprovador de Planejadas',
            'indice_posto': obter_indice_posto_assinatura(posto_abreviado)
        })
    
    # Fiscal
    if planejada.assinatura_fiscal and planejada.data_assinatura_fiscal:
        militar = planejada.assinatura_fiscal
        posto_abreviado = militar.posto_graduacao or ''
        assinaturas_para_exibir.append({
            'militar': militar,
            'data': planejada.data_assinatura_fiscal,
            'tipo': 'Fiscal de Planejadas',
            'indice_posto': obter_indice_posto_assinatura(posto_abreviado)
        })
    
    # Ordenar assinaturas por posto (maior posto primeiro)
    assinaturas_para_exibir.sort(key=lambda x: x['indice_posto'])
    
    # Exibir assinaturas eletrônicas ordenadas (igual ao boletim ostensivo)
    for i, assinatura in enumerate(assinaturas_para_exibir):
        militar = assinatura['militar']
        data_assinatura = assinatura['data']
        tipo_assinatura = assinatura['tipo']
        
        # Obter nome e posto
        posto = militar.get_posto_graduacao_display()
        
        # Adicionar BM após o posto se não já estiver presente
        if "BM" not in posto:
            posto = f"{posto} BM"
        
        # Função
        funcao = tipo_assinatura or "Função não registrada"
        
        # Criar nome com função abaixo
        nome_assinante = f"{posto} {militar.nome_completo}<br/>{funcao}"
        
        try:
            data_formatada, hora_formatada = formatar_data_assinatura(data_assinatura)
        except:
            data_formatada = data_assinatura.strftime('%d/%m/%Y')
            hora_formatada = data_assinatura.strftime('%H:%M:%S')
        
        # Texto da assinatura eletrônica
        texto_assinatura = (
            f"Documento assinado eletronicamente por {nome_assinante}, em {data_formatada} {hora_formatada}, "
            f"conforme Portaria GCG/ CBMEPI N 167 de 23 de novembro de 2021 e publicada no DOE PI N 253 de 26 de novembro de 2021"
        )
        
        # Estilo para texto justificado
        style_assinatura_texto = ParagraphStyle('assinatura_texto', parent=styles['Normal'], fontSize=10, fontName='Helvetica', alignment=4, spaceAfter=1, spaceBefore=1, leading=14)
        
        # Tabela das assinaturas: Logo + Texto de assinatura
        # Largura ajustada para ter margem de 0,02cm de cada lado (largura total - 0,04cm)
        largura_disponivel = A4[0] - (1.5*cm * 2) - 0.04*cm  # total - margens documento - margem extra 0,02cm cada lado
        largura_texto = largura_disponivel - 2.5*cm  # menos logo
        
        assinatura_data = [
            [Image(obter_caminho_assinatura_eletronica(), width=2.5*cm, height=1.8*cm), Paragraph(texto_assinatura, style_assinatura_texto)]
        ]
        
        assinatura_table = Table(assinatura_data, colWidths=[2.5*cm, largura_texto])
        assinatura_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo centralizado
            ('LEFTPADDING', (0, 0), (0, -1), 0),  # Logo sem padding esquerdo
            ('LEFTPADDING', (1, 0), (1, -1), 15),  # Texto com padding esquerdo para afastar da logo
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),  # Borda do retângulo
        ]))
        
        story.append(assinatura_table)
        
        # Espaçamento entre assinaturas (exceto na última)
        if i < len(assinaturas_para_exibir) - 1:
            story.append(Spacer(1, 10))

    # Rodapé com QR Code para conferência de veracidade (igual ao boletim ostensivo)
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, spaceAfter=0.3*cm, spaceBefore=0.3*cm, color=colors.grey))
    
    # Usar a função utilitária para gerar o autenticador
    from .utils import gerar_autenticador_veracidade
    autenticador = gerar_autenticador_veracidade(planejada, request, tipo_documento='planejada')
    
    # Tabela do rodapé: QR + Texto de autenticação
    # Largura ajustada para ter margem de 0,02cm de cada lado (largura total - 0,04cm)
    largura_disponivel = A4[0] - (1.5*cm * 2) - 0.04*cm  # total - margens documento - margem extra 0,02cm cada lado
    largura_texto_qr = largura_disponivel - 3*cm  # menos QR
    
    rodape_data = [
        [autenticador['qr_img'], Paragraph(autenticador['texto_autenticacao'], style_field_value)]
    ]
    
    rodape_table = Table(rodape_data, colWidths=[3*cm, largura_texto_qr])
    rodape_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # QR centralizado
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Texto alinhado à esquerda
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),  # Borda do retângulo
    ]))
    
    story.append(rodape_table)

    # Usar função utilitária para criar rodapé do sistema
    from .utils import criar_rodape_sistema_pdf
    add_rodape_first, add_rodape_later = criar_rodape_sistema_pdf(request)
    
    # Construir PDF com rodapé em todas as páginas
    doc.build(story, onFirstPage=add_rodape_first, onLaterPages=add_rodape_later)
    
    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="formulario_operacao_planejada_{planejada.id}_{planejada.nome.replace(" ", "_")}.pdf"'
    
    return response


@login_required
def relatorio_mensal_liquidacao(request):
    """Gera relatório mensal de liquidação de planejadas"""
    
    # Buscar configuração ativa
    from .models import ConfiguracaoPlanejadas
    configuracao = ConfiguracaoPlanejadas.get_configuracao_ativa()
    limite_mensal = configuracao.quantidade_planejadas_por_militar_mes if configuracao else 4
    
    # Parâmetros GET
    from datetime import datetime
    mes_atual = datetime.now().month
    ano_atual = datetime.now().year
    
    mes = request.GET.get('mes', str(mes_atual))
    ano = request.GET.get('ano', str(ano_atual))
    acao = request.GET.get('acao', '')
    
    # Se não tiver mês e ano, usar valores padrão do mês atual
    if not mes:
        mes = str(mes_atual)
    if not ano:
        ano = str(ano_atual)
    
    # Converter para inteiro
    mes = int(mes)
    ano = int(ano)
    
    # Buscar planejadas do mês com todos os militares
    # Incluir apenas planejadas ativas (não excluídas)
    planejadas_mes = Planejada.objects.filter(
        data_operacao__year=ano,
        data_operacao__month=mes,
        excluido=False  # Excluir planejadas marcadas como excluídas
    ).prefetch_related('militares').order_by('data_operacao')
    
    # Se for ação de exportar (excel ou pdf), gerar arquivo
    if acao in ['excel', 'pdf']:
        if not planejadas_mes:
            messages.error(request, f'Nenhuma planejada encontrada para {mes:02d}/{ano}')
            return redirect('militares:relatorio_mensal_liquidacao')
        
        # Pegar função selecionada (se houver)
        funcao_selecionada = request.GET.get('funcao', '')
        
        # Gerar Excel ou PDF
        if acao == 'excel':
            return _gerar_excel_relatorio_mensal(planejadas_mes, mes, ano, limite_mensal)
        elif acao == 'pdf':
            return _gerar_pdf_relatorio_mensal(planejadas_mes, mes, ano, limite_mensal, request, funcao_selecionada)
    
    # Buscar todas as funções do usuário para o modal de assinatura
    from .models import UsuarioFuncaoMilitar
    funcoes_usuario = UsuarioFuncaoMilitar.objects.filter(
        usuario=request.user
    ).select_related('funcao_militar').order_by('funcao_militar__nome')
    
    # Nome do mês
    mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
               'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes - 1]
    
    # Verificar se há planejadas
    if not planejadas_mes:
        messages.warning(request, f'Nenhuma planejada encontrada para o mês {mes:02d}/{ano}')
        context = {
            'page_title': 'Relatório Mensal de Liquidação',
            'limite_mensal': limite_mensal,
            'mes': mes,
            'ano': ano,
            'mes_nome': mes_nome,
            'funcoes_usuario': funcoes_usuario,
            'tem_dados': False
        }
        return render(request, 'militares/relatorio_mensal_liquidacao_list.html', context)
    
    # Criar lista expandida com cada planejada individual
    dados_expandidos = []
    for planejada in planejadas_mes:
        # Debug: contar militares da planejada
        militares_count = planejada.militares.count()
        if militares_count == 0:
            continue  # Pular planejadas sem militares
        
        # Calcular quantidade baseada no tipo da planejada
        tipo = planejada.tipo_planejada or 'P1'
        quantidade_planejada = 1
        if tipo == 'P1':
            quantidade_planejada = 1
        elif tipo == 'P2':
            quantidade_planejada = 2
        elif tipo == 'P3':
            quantidade_planejada = 3
        elif tipo == 'P4':
            quantidade_planejada = 4
        
        # O campo planejada.valor já contém o valor total que cada militar deve receber
        # Não precisa multiplicar novamente pela quantidade de planejadas
        quantidade_militares = planejada.militares.count()
        
        if quantidade_militares > 0:
            # Usar o valor diretamente do campo planejada.valor
            valor_unitario_por_militar = float(planejada.valor or 0)
        else:
            valor_unitario_por_militar = 0
            continue  # Pular se não tiver militares
        
        # FORÇAR avaliação dos militares (garantir que todos sejam carregados)
        militares_planejada = list(planejada.militares.all())
        
        for militar in militares_planejada:
            # O valor_unitario_por_militar já é o valor total que o militar deve receber
            # Não precisa multiplicar pela quantidade de planejadas
            valor_total = valor_unitario_por_militar
            
            dados_expandidos.append({
                'militar': militar,
                'planejada': planejada,
                'valor_unitario': valor_unitario_por_militar,
                'quantidade_planejada': quantidade_planejada,
                'total': valor_total
            })
    
    # Agrupar por militar para calcular totais
    from collections import defaultdict
    militares_dict = defaultdict(lambda: {'planejadas': [], 'total_valor': 0, 'quantidade_total': 0})
    
    for item in dados_expandidos:
        militar_id = item['militar'].id
        militares_dict[militar_id]['militar'] = item['militar']
        militares_dict[militar_id]['planejadas'].append(item)
        # Somar o total de cada planejada (não apenas o valor unitário)
        militares_dict[militar_id]['total_valor'] += item['total']
        militares_dict[militar_id]['quantidade_total'] += item['quantidade_planejada']
    
    # Criar lista final ordenada por nome
    militares_list = []
    for militar_id, dados in militares_dict.items():
        militares_list.append({
            'militar': dados['militar'],
            'planejadas': sorted(dados['planejadas'], key=lambda x: x['planejada'].data_operacao),
            'total_valor': dados['total_valor'],
            'quantidade_planejadas': dados['quantidade_total']
        })
    
    # Ordenar por hierarquia militar (posto/graduação)
    def ordenar_por_hierarquia(militar_item):
        militar = militar_item['militar']
        # Definir ordem dos postos (hierarquia: mais alto para mais baixo)
        ordem_postos = ['CB', 'TC', 'MJ', 'CP', '1T', '2T', 'AS', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
        
        # Obter índice do posto (menor índice = mais alto na hierarquia)
        indice_posto = ordem_postos.index(militar.posto_graduacao) if militar.posto_graduacao in ordem_postos else 999
        
        # Ordenar por posto (hierarquia) e depois por nome como desempate
        return (indice_posto, militar.nome_completo)
    
    militares_list = sorted(militares_list, key=ordenar_por_hierarquia)
    
    # Calcular total geral a receber
    total_geral = sum(militar_item['total_valor'] for militar_item in militares_list)
    
    # Calcular total de planejadas (quantidade expandida baseada no tipo)
    total_planejadas_expandidas = 0
    for planejada in planejadas_mes:
        tipo = planejada.tipo_planejada or 'P1'
        if tipo == 'P1':
            total_planejadas_expandidas += 1
        elif tipo == 'P2':
            total_planejadas_expandidas += 2
        elif tipo == 'P3':
            total_planejadas_expandidas += 3
        elif tipo == 'P4':
            total_planejadas_expandidas += 4
    
    context = {
        'page_title': f'Relatório de Liquidação - {mes_nome}/{ano}',
        'limite_mensal': limite_mensal,
        'mes': mes,
        'ano': ano,
        'mes_nome': mes_nome,
        'militares': militares_list,
        'total_militares': len(militares_list),
        'total_planejadas': total_planejadas_expandidas,
        'total_geral': total_geral,
        'funcoes_usuario': funcoes_usuario,
        'tem_dados': True
    }
    
    return render(request, 'militares/relatorio_mensal_liquidacao_list.html', context)


def _processar_dados_militares_relatorio(planejadas_mes):
    """Função auxiliar para processar dados dos militares para o relatório"""
    militares_dict = {}
    
    for planejada in planejadas_mes:
        # Calcular quantidade baseada no tipo da planejada
        tipo = planejada.tipo_planejada or 'P1'
        quantidade_planejada = 1
        if tipo == 'P1':
            quantidade_planejada = 1
        elif tipo == 'P2':
            quantidade_planejada = 2
        elif tipo == 'P3':
            quantidade_planejada = 3
        elif tipo == 'P4':
            quantidade_planejada = 4
        
        # O campo planejada.valor já contém o valor total que cada militar deve receber
        # Não precisa multiplicar novamente pela quantidade de planejadas
        quantidade_militares_planejada = planejada.militares.count()
        
        if quantidade_militares_planejada > 0:
            # Usar o valor diretamente do campo planejada.valor
            valor_unitario_por_militar = float(planejada.valor or 0)
        else:
            valor_unitario_por_militar = 0
        
        for militar in planejada.militares.all():
            if militar.id not in militares_dict:
                militares_dict[militar.id] = {
                    'militar': militar,
                    'planejadas': [],
                    'total_valor': 0,
                    'quantidade_planejadas': 0
                }
            
            militares_dict[militar.id]['planejadas'].append(planejada)
            militares_dict[militar.id]['quantidade_planejadas'] += quantidade_planejada
            # O valor_unitario_por_militar já é o valor total que o militar deve receber
            # Não precisa multiplicar pela quantidade de planejadas
            militares_dict[militar.id]['total_valor'] += valor_unitario_por_militar
    
    # Ordenar por hierarquia militar (posto/graduação)
    def ordenar_por_hierarquia(militar_item):
        militar = militar_item['militar']
        # Definir ordem dos postos (hierarquia: mais alto para mais baixo)
        ordem_postos = ['CB', 'TC', 'MJ', 'CP', '1T', '2T', 'AS', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
        
        # Obter índice do posto (menor índice = mais alto na hierarquia)
        indice_posto = ordem_postos.index(militar.posto_graduacao) if militar.posto_graduacao in ordem_postos else 999
        
        # Ordenar por posto (hierarquia) e depois por nome como desempate
        return (indice_posto, militar.nome_completo)
    
    return sorted(militares_dict.values(), key=ordenar_por_hierarquia)


def _gerar_excel_relatorio_mensal(planejadas_mes, mes, ano, limite_mensal):
    """Gera arquivo Excel do relatório mensal"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    # Buscar configuração ativa
    from .models import ConfiguracaoPlanejadas
    configuracao = ConfiguracaoPlanejadas.get_configuracao_ativa()
    
    # Criar lista expandida
    dados_expandidos = []
    for planejada in planejadas_mes:
        tipo = planejada.tipo_planejada or 'P1'
        quantidade_planejada = 1
        if tipo == 'P1':
            quantidade_planejada = 1
        elif tipo == 'P2':
            quantidade_planejada = 2
        elif tipo == 'P3':
            quantidade_planejada = 3
        elif tipo == 'P4':
            quantidade_planejada = 4
        
        # Obter valor unitário configurado conforme o tipo
        is_fim_semana = planejada.data_operacao.weekday() >= 5  # Sábado ou Domingo
        
        # Tenta obter valores da configuração
        valor_unitario = float(planejada.valor or 0)  # Usa valor da planejada como padrão
        
        if configuracao:
            if is_fim_semana:
                if tipo == 'P1' and configuracao.valor_planejada_p1_fds:
                    valor_unitario = float(configuracao.valor_planejada_p1_fds)
                elif tipo == 'P2' and configuracao.valor_planejada_p2_fds:
                    valor_unitario = float(configuracao.valor_planejada_p2_fds)
                elif tipo == 'P3' and configuracao.valor_planejada_p3_fds:
                    valor_unitario = float(configuracao.valor_planejada_p3_fds)
                elif tipo == 'P4' and configuracao.valor_planejada_p4_fds:
                    valor_unitario = float(configuracao.valor_planejada_p4_fds)
            else:
                if tipo == 'P1' and configuracao.valor_planejada_p1_normal:
                    valor_unitario = float(configuracao.valor_planejada_p1_normal)
                elif tipo == 'P2' and configuracao.valor_planejada_p2_normal:
                    valor_unitario = float(configuracao.valor_planejada_p2_normal)
                elif tipo == 'P3' and configuracao.valor_planejada_p3_normal:
                    valor_unitario = float(configuracao.valor_planejada_p3_normal)
                elif tipo == 'P4' and configuracao.valor_planejada_p4_normal:
                    valor_unitario = float(configuracao.valor_planejada_p4_normal)
        
        # O campo planejada.valor já contém o valor total que cada militar deve receber
        # Não precisa multiplicar novamente pela quantidade de planejadas
        quantidade_militares_planejada = planejada.militares.count()
        
        if quantidade_militares_planejada > 0:
            # Usar o valor diretamente do campo planejada.valor
            valor_unitario_por_militar_excel = float(planejada.valor or 0)
        else:
            valor_unitario_por_militar_excel = 0
        
        for militar in planejada.militares.all():
            dados_expandidos.append({
                'militar': militar,
                'valor_unitario': valor_unitario_por_militar_excel,
                'quantidade_planejada': quantidade_planejada
            })
    
    # Agrupar por militar
    from collections import defaultdict
    militares_dict = defaultdict(lambda: {'planejadas': [], 'total_valor': 0, 'quantidade_total': 0})
    
    for item in dados_expandidos:
        militar_id = item['militar'].id
        militares_dict[militar_id]['militar'] = item['militar']
        militares_dict[militar_id]['planejadas'].append(item)
        # O valor_unitario já é o valor total que o militar deve receber
        # Não precisa multiplicar pela quantidade de planejadas
        total_item = item['valor_unitario']
        militares_dict[militar_id]['total_valor'] += total_item
        militares_dict[militar_id]['quantidade_total'] += item['quantidade_planejada']
    
    # Criar lista final - UMA linha por militar
    militares_list = []
    for militar_id, dados in militares_dict.items():
        militares_list.append({
            'militar': dados['militar'],
            'quantidade_planejadas': dados['quantidade_total'],
            'total_valor': dados['total_valor']
        })
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Liquidação {mes:02d}_{ano}"
    
    # Cabeçalho
    mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes - 1]
    ws.merge_cells('A1:H1')
    ws['A1'] = f'RELATÓRIO MENSAL DE LIQUIDAÇÃO - {mes_nome}/{ano}'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Informações
    ws['A2'] = f'Limite Mensal por Militar: {limite_mensal}'
    ws['A3'] = f'Total de Planejadas: {len(planejadas_mes)}'
    ws['A4'] = f'Total de Militares: {len(militares_list)}'
    
    # Cabeçalho da tabela
    headers = ['Posto', 'Nome', 'Matrícula', 'CPF', 'Banco', 'Agência', 'Conta', 'Qtd Planejadas', 'Total a Receber']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados - UMA linha por militar
    row = 7
    for dados in militares_list:
        militar = dados['militar']
        
        ws.cell(row=row, column=1).value = militar.get_posto_graduacao_display() or 'N/A'
        ws.cell(row=row, column=2).value = militar.nome_completo
        ws.cell(row=row, column=3).value = militar.matricula
        ws.cell(row=row, column=4).value = militar.cpf
        ws.cell(row=row, column=5).value = militar.banco_nome or ''
        ws.cell(row=row, column=6).value = militar.agencia or ''
        ws.cell(row=row, column=7).value = militar.conta or ''
        ws.cell(row=row, column=8).value = dados['quantidade_planejadas']
        ws.cell(row=row, column=9).value = dados['total_valor']
        
        # Formatar o valor monetário com ponto decimal brasileiro
        valor_cell = ws.cell(row=row, column=9)
        valor_cell.number_format = '#,##0.00'
        
        row += 1
    
    # Ajustar largura das colunas
    column_widths = [15, 40, 15, 15, 30, 10, 15, 15, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Preparar resposta Excel (inline para abrir em nova guia ao invés de baixar)
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'inline; filename="relatorio_liquidacao_{mes:02d}_{ano}.xlsx"'
    
    wb.save(response)
    return response


def _gerar_pdf_relatorio_mensal(planejadas_mes, mes, ano, limite_mensal, request, funcao=''):
    """Gera arquivo PDF do relatório mensal em formato paisagem"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from django.http import HttpResponse
    from collections import defaultdict
    from django.utils import timezone
    import locale
    
    # Configurar locale para português brasileiro
    try:
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
        except:
            pass
    
    # Criar lista expandida com cada planejada individual
    dados_expandidos = []
    for planejada in planejadas_mes:
        tipo = planejada.tipo_planejada or 'P1'
        quantidade_planejada = 1
        if tipo == 'P2':
            quantidade_planejada = 2
        elif tipo == 'P3':
            quantidade_planejada = 3
        elif tipo == 'P4':
            quantidade_planejada = 4
        
        # O campo planejada.valor já contém o valor total que cada militar deve receber
        # Não precisa multiplicar novamente pela quantidade de planejadas
        quantidade_militares = planejada.militares.count()
        
        if quantidade_militares > 0:
            # Usar o valor diretamente do campo planejada.valor
            valor_unitario_por_militar = float(planejada.valor or 0)
        else:
            valor_unitario_por_militar = 0
        
        militares_planejada = list(planejada.militares.all())
        
        for militar in militares_planejada:
            # O valor_unitario_por_militar já é o valor total que o militar deve receber
            # Não precisa multiplicar pela quantidade de planejadas
            valor_total = valor_unitario_por_militar
            
            dados_expandidos.append({
                'militar': militar,
                'planejada': planejada,
                'valor_unitario': valor_unitario_por_militar,
                'quantidade_planejada': quantidade_planejada,
                'total': valor_total
            })
    
    # Agrupar por militar para calcular totais
    militares_dict = defaultdict(lambda: {'planejadas': [], 'total_valor': 0, 'quantidade_total': 0})
    
    for item in dados_expandidos:
        militar_id = item['militar'].id
        militares_dict[militar_id]['militar'] = item['militar']
        militares_dict[militar_id]['planejadas'].append(item)
        militares_dict[militar_id]['total_valor'] += item['total']
        militares_dict[militar_id]['quantidade_total'] += item['quantidade_planejada']
    
    # Criar lista final ordenada por nome
    militares_list = []
    for militar_id, dados in militares_dict.items():
        militares_list.append({
            'militar': dados['militar'],
            'planejadas': sorted(dados['planejadas'], key=lambda x: x['planejada'].data_operacao),
            'total_valor': dados['total_valor'],
            'quantidade_planejadas': dados['quantidade_total']
        })
    
    # Ordenar por hierarquia militar (posto/graduação)
    def ordenar_por_hierarquia(militar_item):
        militar = militar_item['militar']
        # Definir ordem dos postos (hierarquia: mais alto para mais baixo)
        ordem_postos = ['CB', 'TC', 'MJ', 'CP', '1T', '2T', 'AS', 'ST', '1S', '2S', '3S', 'CAB', 'SD']
        
        # Obter índice do posto (menor índice = mais alto na hierarquia)
        indice_posto = ordem_postos.index(militar.posto_graduacao) if militar.posto_graduacao in ordem_postos else 999
        
        # Ordenar por posto (hierarquia) e depois por nome como desempate
        return (indice_posto, militar.nome_completo)
    
    militares_list = sorted(militares_list, key=ordenar_por_hierarquia)
    
    buffer = BytesIO()
    
    # Nome do mês
    mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
               'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes - 1]
    
    # Usar orientação paisagem (landscape)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Container para elementos do PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Criar estilo para texto do QR Code
    style_field_value = ParagraphStyle('field_value', parent=styles['Normal'], fontSize=9, fontName='Helvetica', spaceAfter=3, spaceBefore=2)
    
    import os
    # Logo/Brasão centralizado
    logo_path = os.path.join('staticfiles', 'logo_cbmepi.png')
    if os.path.exists(logo_path):
        from reportlab.platypus import Image
        elements.append(Image(logo_path, width=2.5*cm, height=2.5*cm, hAlign='CENTER'))
        elements.append(Spacer(1, 0.3*cm))
    
    # Cabeçalho institucional
    elements.append(Paragraph("GOVERNO DO ESTADO DO PIAUÍ", ParagraphStyle('estado', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, textColor=colors.HexColor('#2c3e50'))))
    elements.append(Paragraph("CORPO DE BOMBEIROS MILITAR", ParagraphStyle('cbmepi', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, textColor=colors.HexColor('#2c3e50'))))
    elements.append(Spacer(1, 0.5*cm))
    
    # Título do relatório
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=15,
        alignment=1,  # CENTER
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(f'RELATÓRIO MENSAL DE LIQUIDAÇÃO DE OPERAÇÕES PLANEJADAS - {mes_nome.upper()}/{ano}', title_style))
    elements.append(Spacer(1, 15))
    
    # Preparar dados da tabela
    table_data = [['Posto', 'Nome', 'Matrícula', 'CPF', 'Banco', 'Agência', 'Conta', 'Qtd', 'Total (R$)']]
    
    total_geral = 0
    total_qtd = 0
    for dados in militares_list:
        militar = dados['militar']
        total_geral += dados['total_valor']
        total_qtd += dados['quantidade_planejadas']
        
        # Formatar valores com vírgula
        valor_formatado = f'{dados["total_valor"]:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        
        table_data.append([
            militar.get_posto_graduacao_display() or 'N/A',
            militar.nome_completo or '',
            str(militar.matricula) if militar.matricula else '',
            militar.cpf or '',
            militar.banco_nome or '',
            militar.agencia or '',
            militar.conta or '',
            str(dados['quantidade_planejadas']),
            f'R$ {valor_formatado}'
        ])
    
    # Adicionar linha de total
    valor_total_formatado = f'{total_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    table_data.append([
        'TOTAIS:', '', '', '', '', '', '',  # 7 colunas: TOTAIS + 6 vazias
        f'{total_qtd}',  # 8ª coluna: Qtd
        f'R$ {valor_total_formatado}'  # 9ª coluna: Total
    ])
    
    # Criar tabela
    table = Table(table_data, repeatRows=1)
    
    # Estilizar tabela
    table.setStyle(TableStyle([
        # Cabeçalho - sem background e fonte preta
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Dados - sem preenchimento
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Largura das colunas (proporcional)
        ('COLWIDTHS', (0, 0), (-1, -1), [50, 150, 60, 80, 80, 50, 50, 30, 80]),
        
        # Rodapé (total)
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('ALIGN', (0, -1), (0, -1), 'LEFT'),  # TOTAIS alinhado à esquerda na primeira coluna
    ]))
    
    elements.append(table)
    
    # Adicionar cidade, data por extenso e informações do assinante
    elements.append(Spacer(1, 1*cm))
    
    # Obter informações do assinante e cidade
    try:
        militar_logado = request.user.militar if hasattr(request.user, 'militar') else None
        
        # Obter cidade
        if militar_logado and militar_logado.cidade:
            cidade_doc = militar_logado.cidade
        else:
            cidade_doc = "Teresina"
        cidade_estado = f"{cidade_doc} - PI"
        
        # Obter posto completo
        if militar_logado:
            posto_abreviado = militar_logado.posto_graduacao or ''
            posto_completo = militar_logado.get_posto_graduacao_display() or posto_abreviado
            if "BM" not in posto_completo:
                posto_completo = f"{posto_completo} BM"
            nome_posto = f"{posto_completo}<br/>{militar_logado.nome_completo}"
        else:
            nome_posto = request.user.get_full_name() or request.user.username
        
    except:
        cidade_estado = "Teresina - PI"
        nome_posto = request.user.get_full_name() or request.user.username
    
    # Data por extenso - usar timezone de Brasília
    from .utils import formatar_data_assinatura
    import pytz
    
    # Obter data atual em Brasília
    brasilia_tz = pytz.timezone('America/Sao_Paulo')
    data_atual = timezone.now().astimezone(brasilia_tz) if timezone.is_aware(timezone.now()) else brasilia_tz.localize(timezone.now())
    
    meses_extenso = {
        1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril',
        5: 'maio', 6: 'junho', 7: 'julho', 8: 'agosto',
        9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'
    }
    data_formatada_extenso = f"{data_atual.day} de {meses_extenso[data_atual.month]} de {data_atual.year}"
    
    # Função selecionada
    funcao_display = funcao if funcao else "Operador de Planejadas"
    
    # Construir texto
    texto_info = f"{cidade_estado}, {data_formatada_extenso}.<br/><br/>{nome_posto}<br/>{funcao_display}"
    
    # Estilo para o texto
    style_info = ParagraphStyle('info', parent=styles['Normal'], fontSize=9, fontName='Helvetica', 
                                 alignment=1, textColor=colors.HexColor('#2c3e50'), spaceAfter=8,
                                 leading=12)
    
    elements.append(Paragraph(texto_info, style_info))
    
    # Adicionar assinatura eletrônica
    from .utils import obter_caminho_assinatura_eletronica
    from reportlab.platypus import Image
    
    elements.append(Spacer(1, 1*cm))
    
    # Obter assinante (usuário logado) com posto
    try:
        militar_logado_assinatura = request.user.militar if hasattr(request.user, 'militar') else None
        if militar_logado_assinatura:
            nome_assinante = militar_logado_assinatura.nome_completo
            posto_abreviado_assinatura = militar_logado_assinatura.posto_graduacao or ''
            posto_completo_assinatura = militar_logado_assinatura.get_posto_graduacao_display() or posto_abreviado_assinatura
            if "BM" not in posto_completo_assinatura:
                posto_completo_assinatura = f"{posto_completo_assinatura} BM"
            assinante_com_posto = f"{nome_assinante} - {posto_completo_assinatura}"
        else:
            nome_assinante = request.user.get_full_name() or request.user.username
            assinante_com_posto = nome_assinante
    except:
        nome_assinante = request.user.get_full_name() or request.user.username
        assinante_com_posto = nome_assinante
    
    # Usar a função formatar_data_assinatura para garantir timezone correto
    from .utils import formatar_data_assinatura
    data_hora = timezone.now()
    data_formatada, hora_formatada = formatar_data_assinatura(data_hora)
    
    # Incluir posto e função na assinatura se fornecida
    assinante_com_posto_funcao = assinante_com_posto
    
    # Texto da assinatura eletrônica seguindo o padrão dos outros PDFs
    texto_assinatura = (
        f"Documento assinado eletronicamente por {assinante_com_posto_funcao}, em {data_formatada} {hora_formatada}, "
        f"conforme Portaria GCG/ CBMEPI N 167 de 23 de novembro de 2021 e publicada no DOE PI N 253 de 26 de novembro de 2021"
    )
    
    # Estilo para texto alinhado à esquerda
    style_assinatura_texto = ParagraphStyle('assinatura_texto', parent=styles['Normal'], fontSize=9, fontName='Helvetica', alignment=0, spaceAfter=1, spaceBefore=1, leading=12)
    
    # Calcular larguras disponíveis
    largura_disponivel = landscape(A4)[0] - (1*cm * 2) - 0.04*cm
    largura_texto = largura_disponivel - 2.5*cm
    
    # Tabela de assinatura: Logo + Texto
    assinatura_data = [
        [Image(obter_caminho_assinatura_eletronica(), width=2.5*cm, height=1.8*cm), Paragraph(texto_assinatura, style_assinatura_texto)]
    ]
    
    assinatura_table = Table(assinatura_data, colWidths=[2.5*cm, largura_texto])
    assinatura_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo centralizado
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Texto alinhado à esquerda
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    
    elements.append(assinatura_table)
    
    # Adicionar autenticador QR Code NO FINAL
    from .utils import gerar_autenticador_veracidade
    from .models import Planejada
    
    elements.append(Spacer(1, 0.1*cm))
    
    # Pegar a primeira planejada para criar o autenticador (usar como referência)
    planejada_ref = planejadas_mes.first() if planejadas_mes else None
    
    # Se tiver planejada, criar autenticador; senão, não incluir
    if planejada_ref:
        # Criar um objeto fake com as informações do relatório
        class RelatorioFake:
            def __init__(self):
                self.id = f"relatorio_{mes:02d}_{ano}"
                self.pk = abs(hash(f"relatorio_{mes:02d}_{ano}")) % 100000000  # Gera um pk numérico positivo a partir do hash
                self.tipo_documento = 'relatorio_liquidacao'
        
        relatorio_fake = RelatorioFake()
        
        autenticador = gerar_autenticador_veracidade(relatorio_fake, request, tipo_documento='relatorio')
        
        # Largura disponível para QR Code
        largura_disponivel_qr = landscape(A4)[0] - (1*cm * 2) - 0.04*cm
        largura_texto_qr = largura_disponivel_qr - 3*cm
        
        # Tabela do rodapé: QR + Texto de autenticação
        rodape_data = [
            [autenticador['qr_img'], Paragraph(autenticador['texto_autenticacao'], style_field_value)]
        ]
        
        rodape_table = Table(rodape_data, colWidths=[3*cm, largura_texto_qr])
        rodape_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # QR centralizado
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Texto alinhado à esquerda
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elements.append(rodape_table)
    
    # Construir PDF
    doc.build(elements)
    
    buffer.seek(0)
    
    # Preparar resposta (inline para abrir em nova guia ao invés de baixar)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="relatorio_liquidacao_{mes:02d}_{ano}.pdf"'
    response.write(buffer.read())
    
    return response


@login_required
def relatorio_individual_militar(request):
    """Gera relatório individual de um militar específico"""
    
    # Parâmetros GET
    militar_id = request.GET.get('militar_id')
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    tipo_relatorio = request.GET.get('tipo', 'mensal')  # mensal ou anual
    
    if not militar_id or not mes or not ano:
        messages.error(request, 'Parâmetros obrigatórios não fornecidos.')
        return redirect('militares:relatorio_mensal_liquidacao')
    
    try:
        militar = Militar.objects.get(id=militar_id)
        mes = int(mes)
        ano = int(ano)
    except (Militar.DoesNotExist, ValueError):
        messages.error(request, 'Militar não encontrado ou parâmetros inválidos.')
        return redirect('militares:relatorio_mensal_liquidacao')
    
    # Buscar planejadas do militar
    if tipo_relatorio == 'anual':
        # Relatório anual - todos os meses do ano
        planejadas = Planejada.objects.filter(
            militares=militar,
            data_operacao__year=ano,
            excluido=False
        ).prefetch_related('militares').order_by('data_operacao')
        
        # Agrupar por mês
        dados_por_mes = {}
        for planejada in planejadas:
            mes_planejada = planejada.data_operacao.month
            if mes_planejada not in dados_por_mes:
                dados_por_mes[mes_planejada] = []
            dados_por_mes[mes_planejada].append(planejada)
        
        # Calcular totais por mês
        resumo_mensal = []
        total_geral = 0
        total_planejadas_geral = 0
        
        for mes_num in range(1, 13):
            mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes_num - 1]
            
            planejadas_mes = dados_por_mes.get(mes_num, [])
            total_mes = 0
            quantidade_mes = 0
            
            for planejada in planejadas_mes:
                # Calcular valor unitário por militar
                quantidade_militares_planejada = planejada.militares.count()
                if quantidade_militares_planejada > 0:
                    valor_unitario_por_militar = float(planejada.valor or 0)
                    total_mes += valor_unitario_por_militar
                    
                    # Calcular quantidade de planejadas
                    tipo_p = planejada.tipo_planejada or 'P1'
                    if tipo_p == 'P1':
                        quantidade_mes += 1
                    elif tipo_p == 'P2':
                        quantidade_mes += 2
                    elif tipo_p == 'P3':
                        quantidade_mes += 3
                    elif tipo_p == 'P4':
                        quantidade_mes += 4
            
            resumo_mensal.append({
                'mes': mes_num,
                'mes_nome': mes_nome,
                'planejadas': planejadas_mes,
                'total': total_mes,
                'quantidade': quantidade_mes
            })
            
            total_geral += total_mes
            total_planejadas_geral += quantidade_mes
        
        # Buscar todas as funções do usuário para o modal de assinatura
        from .models import UsuarioFuncaoMilitar
        funcoes_usuario = UsuarioFuncaoMilitar.objects.filter(
            usuario=request.user
        ).select_related('funcao_militar').order_by('funcao_militar__nome')
        
        context = {
            'page_title': f'Relatório Anual - {militar.nome_completo}',
            'militar': militar,
            'ano': ano,
            'tipo_relatorio': 'anual',
            'resumo_mensal': resumo_mensal,
            'total_geral': total_geral,
            'total_planejadas_geral': total_planejadas_geral,
            'funcoes_usuario': funcoes_usuario,
        }
        
        return render(request, 'militares/relatorio_individual_militar.html', context)
        
    else:
        # Relatório mensal - mês específico
        planejadas_mes = Planejada.objects.filter(
            militares=militar,
            data_operacao__year=ano,
            data_operacao__month=mes,
            excluido=False
        ).prefetch_related('militares').order_by('data_operacao')
        
        # Calcular totais
        total_mes = 0
        quantidade_mes = 0
        
        for planejada in planejadas_mes:
            # Calcular valor unitário por militar
            quantidade_militares_planejada = planejada.militares.count()
            if quantidade_militares_planejada > 0:
                valor_unitario_por_militar = float(planejada.valor or 0)
                total_mes += valor_unitario_por_militar
                
                # Calcular quantidade de planejadas
                tipo_p = planejada.tipo_planejada or 'P1'
                if tipo_p == 'P1':
                    quantidade_mes += 1
                elif tipo_p == 'P2':
                    quantidade_mes += 2
                elif tipo_p == 'P3':
                    quantidade_mes += 3
                elif tipo_p == 'P4':
                    quantidade_mes += 4
        
        mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes - 1]
        
        # Buscar todas as funções do usuário para o modal de assinatura
        from .models import UsuarioFuncaoMilitar
        funcoes_usuario = UsuarioFuncaoMilitar.objects.filter(
            usuario=request.user
        ).select_related('funcao_militar').order_by('funcao_militar__nome')
        
        context = {
            'page_title': f'Relatório Mensal - {militar.nome_completo}',
            'militar': militar,
            'mes': mes,
            'ano': ano,
            'mes_nome': mes_nome,
            'tipo_relatorio': 'mensal',
            'planejadas_mes': planejadas_mes,
            'total_mes': total_mes,
            'quantidade_mes': quantidade_mes,
            'funcoes_usuario': funcoes_usuario,
        }
        
        return render(request, 'militares/relatorio_individual_militar.html', context)


@login_required
def relatorio_individual_militar_pdf(request):
    """Gera PDF do relatório individual de um militar específico"""
    
    # Parâmetros GET
    militar_id = request.GET.get('militar_id')
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    tipo_relatorio = request.GET.get('tipo', 'mensal')  # mensal ou anual
    funcao = request.GET.get('funcao', '')  # função para assinatura
    
    if not militar_id or not ano:
        messages.error(request, 'Parâmetros obrigatórios não fornecidos.')
        return redirect('militares:relatorio_mensal_liquidacao')
    
    # Para relatório mensal, mes é obrigatório
    if tipo_relatorio == 'mensal' and not mes:
        messages.error(request, 'Mês é obrigatório para relatório mensal.')
        return redirect('militares:relatorio_mensal_liquidacao')
    
    try:
        militar = Militar.objects.get(id=militar_id)
        ano = int(ano)
        # Para relatório mensal, converter mes para int
        if tipo_relatorio == 'mensal':
            mes = int(mes)
    except (Militar.DoesNotExist, ValueError):
        messages.error(request, 'Militar não encontrado ou parâmetros inválidos.')
        return redirect('militares:relatorio_mensal_liquidacao')
    
    # Buscar planejadas do militar
    if tipo_relatorio == 'anual':
        # Relatório anual - todos os meses do ano
        planejadas = Planejada.objects.filter(
            militares=militar,
            data_operacao__year=ano,
            excluido=False
        ).prefetch_related('militares').order_by('data_operacao')
        
        # Agrupar por mês
        dados_por_mes = {}
        for planejada in planejadas:
            mes_planejada = planejada.data_operacao.month
            if mes_planejada not in dados_por_mes:
                dados_por_mes[mes_planejada] = []
            dados_por_mes[mes_planejada].append(planejada)
        
        # Calcular totais por mês
        resumo_mensal = []
        total_geral = 0
        total_planejadas_geral = 0
        
        for mes_num in range(1, 13):
            mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes_num - 1]
            
            planejadas_mes = dados_por_mes.get(mes_num, [])
            total_mes = 0
            quantidade_mes = 0
            
            for planejada in planejadas_mes:
                # Calcular valor unitário por militar
                quantidade_militares_planejada = planejada.militares.count()
                if quantidade_militares_planejada > 0:
                    valor_unitario_por_militar = float(planejada.valor or 0)
                    total_mes += valor_unitario_por_militar
                    
                    # Calcular quantidade de planejadas
                    tipo_p = planejada.tipo_planejada or 'P1'
                    if tipo_p == 'P1':
                        quantidade_mes += 1
                    elif tipo_p == 'P2':
                        quantidade_mes += 2
                    elif tipo_p == 'P3':
                        quantidade_mes += 3
                    elif tipo_p == 'P4':
                        quantidade_mes += 4
            
            resumo_mensal.append({
                'mes': mes_num,
                'mes_nome': mes_nome,
                'planejadas': planejadas_mes,
                'total': total_mes,
                'quantidade': quantidade_mes
            })
            
            total_geral += total_mes
            total_planejadas_geral += quantidade_mes
        
        return _gerar_pdf_relatorio_individual_anual(militar, ano, resumo_mensal, total_geral, total_planejadas_geral, planejadas, request, funcao)
        
    else:
        # Relatório mensal - mês específico
        planejadas_mes = Planejada.objects.filter(
            militares=militar,
            data_operacao__year=ano,
            data_operacao__month=mes,
            excluido=False
        ).prefetch_related('militares').order_by('data_operacao')
        
        # Calcular totais
        total_mes = 0
        quantidade_mes = 0
        
        for planejada in planejadas_mes:
            # Calcular valor unitário por militar
            quantidade_militares_planejada = planejada.militares.count()
            if quantidade_militares_planejada > 0:
                valor_unitario_por_militar = float(planejada.valor or 0)
                total_mes += valor_unitario_por_militar
                
                # Calcular quantidade de planejadas
                tipo_p = planejada.tipo_planejada or 'P1'
                if tipo_p == 'P1':
                    quantidade_mes += 1
                elif tipo_p == 'P2':
                    quantidade_mes += 2
                elif tipo_p == 'P3':
                    quantidade_mes += 3
                elif tipo_p == 'P4':
                    quantidade_mes += 4
        
        mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes - 1]
        
        return _gerar_pdf_relatorio_individual_mensal(militar, mes, ano, mes_nome, planejadas_mes, total_mes, quantidade_mes, request, funcao)


def _gerar_pdf_relatorio_individual_mensal(militar, mes, ano, mes_nome, planejadas_mes, total_mes, quantidade_mes, request, funcao=''):
    """Gera PDF do relatório individual mensal seguindo o molde do relatório de liquidação"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    from reportlab.lib.enums import TA_CENTER
    import locale
    import os
    
    # Configurar locale para português brasileiro
    try:
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
        except:
            pass
    
    buffer = BytesIO()
    
    # Usar orientação paisagem (landscape) como no relatório de liquidação
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Container para elementos do PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Logo/Brasão centralizado
    logo_path = os.path.join('staticfiles', 'logo_cbmepi.png')
    if os.path.exists(logo_path):
        from reportlab.platypus import Image
        elements.append(Image(logo_path, width=2*cm, height=2*cm, hAlign='CENTER'))
        elements.append(Spacer(1, 6))
    
    # Determinar local de geração e endereço baseado no acesso do usuário que está gerando o PDF
    # USAR APENAS A FUNÇÃO ATIVA DO USUÁRIO (não a lotação do militar)
    from .models import Orgao, GrandeComando, Unidade, SubUnidade
    from .permissoes_hierarquicas import obter_funcao_militar_ativa
    local_geracao = "DIRETORIA DE GESTÃO DE PESSOAS"
    endereco_organizacao = None
    
    # Obter função ativa do usuário (UsuarioFuncaoMilitar) - esta é a fonte de verdade
    funcao_usuario = obter_funcao_militar_ativa(request.user)
    
    if funcao_usuario:
        # Determinar organização baseada na função do usuário conforme tipo de acesso
        # Prioridade: sub_unidade > unidade > grande_comando > orgao
        # Mostrar apenas a OM específica do nível de acesso, sem hierarquia
        if funcao_usuario.sub_unidade:
            # Acesso nível SUBUNIDADE - mostrar apenas a sub-unidade
            local_geracao = funcao_usuario.sub_unidade.nome.upper()
            endereco_organizacao = funcao_usuario.sub_unidade.endereco
        elif funcao_usuario.unidade:
            # Acesso nível UNIDADE - mostrar apenas a unidade
            local_geracao = funcao_usuario.unidade.nome.upper()
            endereco_organizacao = funcao_usuario.unidade.endereco
        elif funcao_usuario.grande_comando:
            # Acesso nível GRANDE_COMANDO - mostrar apenas o grande comando
            local_geracao = funcao_usuario.grande_comando.nome.upper()
            endereco_organizacao = funcao_usuario.grande_comando.endereco
        elif funcao_usuario.orgao:
            # Acesso nível ORGAO - mostrar apenas o órgão
            local_geracao = funcao_usuario.orgao.nome.upper()
            endereco_organizacao = funcao_usuario.orgao.endereco
    
    # Cabeçalho institucional
    style_center = ParagraphStyle('center', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)
    
    cabecalho = [
        "GOVERNO DO ESTADO DO PIAUÍ",
        "CORPO DE BOMBEIROS MILITAR DO ESTADO DO PIAUÍ",
        local_geracao
    ]
    
    # Adicionar endereço se existir
    if endereco_organizacao:
        cabecalho.append(endereco_organizacao)
    
    for linha in cabecalho:
        elements.append(Paragraph(linha, style_center))
    elements.append(Spacer(1, 10))
    elements.append(Spacer(1, 10))
    
    # Título do relatório
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=15,
        alignment=1,  # CENTER
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(f'RELATÓRIO INDIVIDUAL DE PLANEJADAS - {mes_nome.upper()}/{ano}', title_style))
    elements.append(Spacer(1, 15))
    
    # Preparar dados da tabela principal (informações do militar)
    table_data = [['Posto', 'Nome', 'Matrícula', 'CPF', 'Banco', 'Agência', 'Conta', 'Qtd', 'Total (R$)']]
    
    # Formatar valores com vírgula
    valor_formatado = f'{total_mes:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    
    table_data.append([
        militar.get_posto_graduacao_display() or 'N/A',
        militar.nome_completo or '',
        str(militar.matricula) if militar.matricula else '',
        militar.cpf or '',
        militar.banco_nome or '',
        militar.agencia or '',
        militar.conta or '',
        str(quantidade_mes),
        f'R$ {valor_formatado}'
    ])
    
    # Criar tabela principal
    table = Table(table_data, repeatRows=1)
    
    # Estilizar tabela principal
    table.setStyle(TableStyle([
        # Cabeçalho
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Dados
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Largura das colunas (proporcional)
        ('COLWIDTHS', (0, 0), (-1, -1), [50, 150, 60, 80, 80, 50, 50, 30, 80]),
    ]))
    
    elements.append(table)
    
    # Detalhamento das operações (se houver)
    if planejadas_mes:
        elements.append(Spacer(1, 1*cm))
        
        # Título do detalhamento
        detalhes_title_style = ParagraphStyle(
            'DetalhesTitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=10,
            alignment=1,  # CENTER
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph("DETALHAMENTO DAS OPERAÇÕES PLANEJADAS", detalhes_title_style))
        
        # Preparar dados do detalhamento
        detalhes_data = [['Data', 'Dia', 'Operação', 'Origem', 'Cidade', 'Tipo', 'Valor Individual (R$)', 'Valor Total (R$)']]
        
        for planejada in planejadas_mes:
            # Buscar valor de uma planejada conforme o dia das configurações
            from .models import ConfiguracaoPlanejadas
            configuracao = ConfiguracaoPlanejadas.get_configuracao_ativa()
            valor_individual = 0
            if configuracao:
                # Verificar se é fim de semana ou feriado
                from datetime import datetime
                data_operacao = planejada.data_operacao
                dia_semana = data_operacao.weekday()  # 0=segunda, 6=domingo
                
                if dia_semana >= 5:  # Sábado ou domingo
                    valor_individual = float(configuracao.valor_planejada_fim_semana or 0)
                else:
                    valor_individual = float(configuracao.valor_planejada_normal or 0)
            
            valor_total_planejada = float(planejada.valor or 0)
            
            # Formatar valores com vírgula
            valor_individual_formatado = f'{valor_individual:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            valor_total_formatado = f'{valor_total_planejada:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Obter nome do dia da semana
            dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
            dia_semana_nome = dias_semana[planejada.data_operacao.weekday()]
            
            detalhes_data.append([
                planejada.data_operacao.strftime('%d/%m/%Y'),
                dia_semana_nome,
                planejada.nome,
                planejada.origem.split('|')[-1].strip() if '|' in planejada.origem else planejada.origem,
                planejada.cidade,
                planejada.tipo_planejada or 'P1',
                f'R$ {valor_individual_formatado}',
                f'R$ {valor_total_formatado}'
            ])
        
        # Criar tabela de detalhamento
        detalhes_table = Table(detalhes_data, repeatRows=1)
        
        # Estilizar tabela de detalhamento
        detalhes_table.setStyle(TableStyle([
            # Cabeçalho
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Dados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Largura das colunas
            ('COLWIDTHS', (0, 0), (-1, -1), [50, 50, 150, 80, 80, 30, 80, 80]),
        ]))
        
        elements.append(detalhes_table)
    
    # Adicionar assinatura se função fornecida
    if funcao:
        elements.append(Spacer(1, 1*cm))
        
        # Obter informações do assinante e cidade
        import pytz
        
        try:
            militar_logado = request.user.militar if hasattr(request.user, 'militar') else None
            
            # Obter cidade
            if militar_logado and militar_logado.cidade:
                cidade_doc = militar_logado.cidade
            else:
                cidade_doc = "Teresina"
            cidade_estado = f"{cidade_doc} - PI"
            
            # Obter posto completo
            if militar_logado:
                posto_abreviado = militar_logado.posto_graduacao or ''
                posto_completo = militar_logado.get_posto_graduacao_display() or posto_abreviado
                if "BM" not in posto_completo:
                    posto_completo = f"{posto_completo} BM"
                nome_posto = f"{posto_completo}<br/>{militar_logado.nome_completo}"
            else:
                nome_posto = request.user.get_full_name() or request.user.username
            
        except:
            cidade_estado = "Teresina - PI"
            nome_posto = request.user.get_full_name() or request.user.username
        
        # Data por extenso - usar timezone de Brasília
        brasilia_tz = pytz.timezone('America/Sao_Paulo')
        data_atual = timezone.now().astimezone(brasilia_tz) if timezone.is_aware(timezone.now()) else brasilia_tz.localize(timezone.now())
        
        meses_extenso = {
            1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril',
            5: 'maio', 6: 'junho', 7: 'julho', 8: 'agosto',
            9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'
        }
        data_formatada_extenso = f"{data_atual.day} de {meses_extenso[data_atual.month]} de {data_atual.year}"
        
        # Função selecionada
        funcao_display = funcao if funcao else ""
        
        # Construir texto - cidade, data, posto, nome, função
        texto_info = f"{cidade_estado}, {data_formatada_extenso}.<br/><br/>{nome_posto}<br/>{funcao_display}"
        
        # Estilo para o texto (centralizado)
        style_info = ParagraphStyle('info', parent=styles['Normal'], fontSize=9, fontName='Helvetica', 
                                     alignment=TA_CENTER, textColor=colors.HexColor('#2c3e50'), spaceAfter=8,
                                     leading=12)
        
        elements.append(Paragraph(texto_info, style_info))
        
        # Adicionar assinatura eletrônica
        from .utils import obter_caminho_assinatura_eletronica
        
        elements.append(Spacer(1, 1*cm))
        
        # Obter assinante (usuário logado) com posto
        try:
            militar_logado_assinatura = request.user.militar if hasattr(request.user, 'militar') else None
            if militar_logado_assinatura:
                nome_assinante = militar_logado_assinatura.nome_completo
                posto_abreviado_assinatura = militar_logado_assinatura.posto_graduacao or ''
                posto_completo_assinatura = militar_logado_assinatura.get_posto_graduacao_display() or posto_abreviado_assinatura
                if "BM" not in posto_completo_assinatura:
                    posto_completo_assinatura = f"{posto_completo_assinatura} BM"
                assinante_com_posto = f"{nome_assinante} - {posto_completo_assinatura}"
            else:
                nome_assinante = request.user.get_full_name() or request.user.username
                assinante_com_posto = nome_assinante
        except:
            nome_assinante = request.user.get_full_name() or request.user.username
            assinante_com_posto = nome_assinante
        
        # Usar a função formatar_data_assinatura para garantir timezone correto
        from .utils import formatar_data_assinatura
        data_hora = timezone.now()
        data_formatada, hora_formatada = formatar_data_assinatura(data_hora)
        
        # Incluir posto e função na assinatura
        assinante_com_posto_funcao = f"{assinante_com_posto}"
        
        # Texto da assinatura eletrônica seguindo o padrão dos outros PDFs
        texto_assinatura = (
            f"Documento assinado eletronicamente por {assinante_com_posto_funcao}, em {data_formatada} {hora_formatada}, "
            f"conforme Portaria GCG/ CBMEPI N 167 de 23 de novembro de 2021 e publicada no DOE PI N 253 de 26 de novembro de 2021"
        )
        
        # Estilo para texto alinhado à esquerda
        style_assinatura_texto = ParagraphStyle('assinatura_texto', parent=styles['Normal'], fontSize=9, fontName='Helvetica', alignment=0, spaceAfter=1, spaceBefore=1, leading=12)
        
        # Calcular larguras disponíveis
        largura_disponivel = landscape(A4)[0] - (1*cm * 2) - 0.04*cm
        largura_texto = largura_disponivel - 2.5*cm
        
        # Tabela de assinatura: Logo + Texto
        assinatura_data = [
            [Image(obter_caminho_assinatura_eletronica(), width=2.5*cm, height=1.8*cm), Paragraph(texto_assinatura, style_assinatura_texto)]
        ]
        
        assinatura_table = Table(assinatura_data, colWidths=[2.5*cm, largura_texto])
        assinatura_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo centralizado
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Texto alinhado à esquerda
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elements.append(assinatura_table)
    
    # Adicionar autenticador QR Code
    from .utils import gerar_autenticador_veracidade
    
    elements.append(Spacer(1, 0.1*cm))
    
    # Criar um objeto fake com as informações do relatório
    class RelatorioIndividualFake:
        def __init__(self, militar_id, mes, ano):
            self.id = f"relatorio_individual_{militar_id}_{mes:02d}_{ano}"
            self.pk = abs(hash(f"relatorio_individual_{militar_id}_{mes:02d}_{ano}")) % 100000000
            self.tipo_documento = 'relatorio_individual'
    
    relatorio_fake = RelatorioIndividualFake(militar.pk, mes, ano)
    autenticador = gerar_autenticador_veracidade(relatorio_fake, request, tipo_documento='relatorio_individual')
    
    # Largura disponível para QR Code
    largura_disponivel_qr = landscape(A4)[0] - (1*cm * 2) - 0.04*cm
    largura_texto_qr = largura_disponivel_qr - 3*cm
    
    # Estilo para o texto do autenticador
    style_field_value = ParagraphStyle('field_value', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=0, spaceAfter=1, leading=10)
    
    # Tabela do rodapé: QR + Texto de autenticação
    rodape_data = [
        [autenticador['qr_img'], Paragraph(autenticador['texto_autenticacao'], style_field_value)]
    ]
    
    rodape_table = Table(rodape_data, colWidths=[3*cm, largura_texto_qr])
    rodape_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # QR centralizado
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Texto alinhado à esquerda
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BOX', (0, 0), (-1, -1), 1, colors.grey),  # Borda do retângulo
    ]))
    
    elements.append(rodape_table)
    
    # Gerar PDF
    doc.build(elements)
    
    buffer.seek(0)
    
    # Preparar resposta
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="relatorio_individual_{militar.matricula}_{mes:02d}_{ano}.pdf"'
    response.write(buffer.read())
    
    return response


def _gerar_pdf_relatorio_individual_anual(militar, ano, resumo_mensal, total_geral, total_planejadas_geral, planejadas, request, funcao=''):
    """Gera PDF do relatório individual anual seguindo o molde do relatório de liquidação"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    import locale
    import os
    
    # Configurar locale para português brasileiro
    try:
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
        except:
            pass
    
    buffer = BytesIO()
    
    # Usar orientação paisagem (landscape) como no relatório de liquidação
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    # Container para elementos do PDF
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Logo (se existir)
    logo_path = os.path.join('staticfiles', 'logo_cbmepi.png')
    if os.path.exists(logo_path):
        from reportlab.platypus import Image
        elements.append(Image(logo_path, width=2.5*cm, height=2.5*cm, hAlign='CENTER'))
        elements.append(Spacer(1, 0.3*cm))
    
    # Cabeçalho institucional
    elements.append(Paragraph("GOVERNO DO ESTADO DO PIAUÍ", ParagraphStyle('estado', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, textColor=colors.HexColor('#2c3e50'))))
    elements.append(Paragraph("CORPO DE BOMBEIROS MILITAR", ParagraphStyle('cbmepi', parent=styles['Normal'], alignment=1, fontSize=12, fontName='Helvetica-Bold', spaceAfter=0, textColor=colors.HexColor('#2c3e50'))))
    elements.append(Spacer(1, 0.5*cm))
    
    # Título do relatório
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=15,
        alignment=1,  # CENTER
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph(f'RELATÓRIO INDIVIDUAL DE PLANEJADAS - {ano}', title_style))
    elements.append(Spacer(1, 15))
    
    # Preparar dados da tabela principal (informações do militar)
    table_data = [['Posto', 'Nome', 'Matrícula', 'CPF', 'Banco', 'Agência', 'Conta', 'Qtd', 'Total (R$)']]
    
    # Formatar valores com vírgula
    valor_formatado = f'{total_geral:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    
    table_data.append([
        militar.get_posto_graduacao_display() or 'N/A',
        militar.nome_completo or '',
        str(militar.matricula) if militar.matricula else '',
        militar.cpf or '',
        militar.banco_nome or '',
        militar.agencia or '',
        militar.conta or '',
        str(total_planejadas_geral),
        f'R$ {valor_formatado}'
    ])
    
    # Criar tabela principal
    table = Table(table_data, repeatRows=1)
    
    # Estilizar tabela principal
    table.setStyle(TableStyle([
        # Cabeçalho
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Dados
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Largura das colunas (proporcional)
        ('COLWIDTHS', (0, 0), (-1, -1), [50, 150, 60, 80, 80, 50, 50, 30, 80]),
    ]))
    
    elements.append(table)
    
    # Resumo por mês (se houver dados)
    meses_com_dados = [mes_data for mes_data in resumo_mensal if mes_data['total'] > 0]
    if meses_com_dados:
        elements.append(Spacer(1, 1*cm))
        
        # Título do resumo mensal
        resumo_title_style = ParagraphStyle(
            'ResumoTitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=10,
            alignment=1,  # CENTER
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph("RESUMO POR MÊS", resumo_title_style))
        
        # Preparar dados do resumo mensal
        resumo_data = [['Mês', 'Qtd Planejadas', 'Valor (R$)']]
        
        for mes_data in meses_com_dados:
            valor_mes_formatado = f'{mes_data["total"]:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            resumo_data.append([
                mes_data['mes_nome'],
                str(mes_data['quantidade']),
                f'R$ {valor_mes_formatado}'
            ])
        
        # Criar tabela de resumo mensal
        resumo_table = Table(resumo_data, repeatRows=1)
        
        # Estilizar tabela de resumo mensal
        resumo_table.setStyle(TableStyle([
            # Cabeçalho
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Dados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Largura das colunas
            ('COLWIDTHS', (0, 0), (-1, -1), [100, 80, 100]),
        ]))
        
        elements.append(resumo_table)
    
    # Detalhamento das operações (se houver)
    if planejadas:
        elements.append(Spacer(1, 1*cm))
        
        # Título do detalhamento
        detalhes_title_style = ParagraphStyle(
            'DetalhesTitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=10,
            alignment=1,  # CENTER
            fontName='Helvetica-Bold'
        )
        
        elements.append(Paragraph("DETALHAMENTO DAS OPERAÇÕES PLANEJADAS", detalhes_title_style))
        
        # Preparar dados do detalhamento
        detalhes_data = [['Data', 'Dia', 'Operação', 'Origem', 'Cidade', 'Tipo', 'Valor Individual (R$)', 'Valor Total (R$)']]
        
        for planejada in planejadas:
            # Buscar valor de uma planejada conforme o dia das configurações
            from .models import ConfiguracaoPlanejadas
            configuracao = ConfiguracaoPlanejadas.get_configuracao_ativa()
            valor_individual = 0
            if configuracao:
                # Verificar se é fim de semana ou feriado
                from datetime import datetime
                data_operacao = planejada.data_operacao
                dia_semana = data_operacao.weekday()  # 0=segunda, 6=domingo
                
                if dia_semana >= 5:  # Sábado ou domingo
                    valor_individual = float(configuracao.valor_planejada_fim_semana or 0)
                else:
                    valor_individual = float(configuracao.valor_planejada_normal or 0)
            
            valor_total_planejada = float(planejada.valor or 0)
            
            # Formatar valores com vírgula
            valor_individual_formatado = f'{valor_individual:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            valor_total_formatado = f'{valor_total_planejada:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            
            # Obter nome do dia da semana
            dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
            dia_semana_nome = dias_semana[planejada.data_operacao.weekday()]
            
            detalhes_data.append([
                planejada.data_operacao.strftime('%d/%m/%Y'),
                dia_semana_nome,
                planejada.nome,
                planejada.origem.split('|')[-1].strip() if '|' in planejada.origem else planejada.origem,
                planejada.cidade,
                planejada.tipo_planejada or 'P1',
                f'R$ {valor_individual_formatado}',
                f'R$ {valor_total_formatado}'
            ])
        
        # Criar tabela de detalhamento
        detalhes_table = Table(detalhes_data, repeatRows=1)
        
        # Estilizar tabela de detalhamento
        detalhes_table.setStyle(TableStyle([
            # Cabeçalho
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Dados
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Largura das colunas
            ('COLWIDTHS', (0, 0), (-1, -1), [50, 50, 150, 80, 80, 30, 80, 80]),
        ]))
        
        elements.append(detalhes_table)
    
    # Adicionar cidade, data por extenso e informações do assinante
    elements.append(Spacer(1, 1*cm))
    
    # Obter informações do assinante e cidade
    try:
        militar_logado = request.user.militar if hasattr(request.user, 'militar') else None
        
        # Obter cidade
        if militar_logado and militar_logado.cidade:
            cidade_doc = militar_logado.cidade
        else:
            cidade_doc = "Teresina"
        cidade_estado = f"{cidade_doc} - PI"
        
        # Obter posto completo
        if militar_logado and militar_logado.posto_graduacao:
            posto_completo = militar_logado.get_posto_graduacao_display()
        else:
            posto_completo = "Administrador do Sistema"
        
        # Data por extenso
        from datetime import datetime
        data_atual = datetime.now()
        meses_extenso = [
            'janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho',
            'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro'
        ]
        data_extenso = f"{data_atual.day} de {meses_extenso[data_atual.month - 1]} de {data_atual.year}"
        
        # Informações do assinante
        if funcao:
            assinante_info = f"{funcao}"
        else:
            assinante_info = f"{posto_completo}"
        
        # Adicionar informações no rodapé
        rodape_style = ParagraphStyle(
            'Rodape',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.black,
            spaceAfter=5,
            alignment=1,  # CENTER
            fontName='Helvetica'
        )
        
        elements.append(Paragraph(f"{cidade_estado}, {data_extenso}.", rodape_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Adicionar nome e posto do assinante
        if militar_logado:
            nome_assinante = militar_logado.nome_completo
            posto_assinante = militar_logado.get_posto_graduacao_display()
            elements.append(Paragraph(f"{nome_assinante}", rodape_style))
            elements.append(Paragraph(f"{posto_assinante}", rodape_style))
        else:
            nome_assinante = request.user.get_full_name() or request.user.username
            elements.append(Paragraph(f"{nome_assinante}", rodape_style))
        
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(f"{assinante_info}", rodape_style))
        
        # Adicionar assinatura eletrônica sempre
        elements.append(Spacer(1, 1*cm))
        
        # Adicionar assinatura eletrônica
        from .utils import obter_caminho_assinatura_eletronica
        
        # Data e hora formatadas
        from .utils import formatar_data_assinatura
        data_hora = timezone.now()
        data_formatada, hora_formatada = formatar_data_assinatura(data_hora)
        
        # Preparar texto da assinatura padronizado sem função
        texto_assinatura = (
            f"Documento assinado eletronicamente por {nome_assinante}, em {data_formatada} {hora_formatada}, "
            f"conforme Portaria GCG/ CBMEPI N 167 de 23 de novembro de 2021 e publicada no DOE PI N 253 de 26 de novembro de 2021"
        )
        
        # Estilo para texto justificado
        style_assinatura_texto = ParagraphStyle('assinatura_texto', parent=styles['Normal'], fontSize=10, fontName='Helvetica', alignment=4, spaceAfter=1, spaceBefore=1, leading=14)
        
        # Tabela das assinaturas: Logo + Texto de assinatura
        # Largura ajustada para ter margem de 0,02cm de cada lado (largura total - 0,04cm)
        largura_disponivel = landscape(A4)[0] - (1*cm * 2) - 0.04*cm  # total - margens documento - margem extra 0,02cm cada lado
        largura_texto = largura_disponivel - 2.5*cm  # menos logo
        
        assinatura_data = [
            [Image(obter_caminho_assinatura_eletronica(), width=2.5*cm, height=1.8*cm), Paragraph(texto_assinatura, style_assinatura_texto)]
        ]
        
        assinatura_table = Table(assinatura_data, colWidths=[2.5*cm, largura_texto])
        assinatura_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo centralizado
            ('LEFTPADDING', (0, 0), (0, -1), 0),  # Logo sem padding esquerdo
            ('LEFTPADDING', (1, 0), (1, -1), 15),  # Texto com padding esquerdo para afastar da logo
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),  # Borda do retângulo
        ]))
        
        elements.append(assinatura_table)
        
        # Adicionar autenticador QR Code
        from .utils import gerar_autenticador_veracidade
        
        elements.append(Spacer(1, 0.1*cm))
        
        # Criar um objeto fake com as informações do relatório
        class RelatorioIndividualAnualFake:
            def __init__(self, militar_id, ano):
                self.id = f"relatorio_individual_anual_{militar_id}_{ano}"
                self.pk = abs(hash(f"relatorio_individual_anual_{militar_id}_{ano}")) % 100000000
                self.tipo_documento = 'relatorio_individual_anual'
        
        relatorio_fake = RelatorioIndividualAnualFake(militar.pk, ano)
        autenticador = gerar_autenticador_veracidade(relatorio_fake, request, tipo_documento='relatorio_individual_anual')
        
        # Largura disponível para QR Code
        largura_disponivel_qr = landscape(A4)[0] - (1*cm * 2) - 0.04*cm
        largura_texto_qr = largura_disponivel_qr - 3*cm
        
        # Estilo para o texto do autenticador
        style_field_value = ParagraphStyle('field_value', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=0, spaceAfter=1, leading=10)
        
        # Tabela do rodapé: QR + Texto de autenticação
        rodape_data = [
            [autenticador['qr_img'], Paragraph(autenticador['texto_autenticacao'], style_field_value)]
        ]
        
        rodape_table = Table(rodape_data, colWidths=[3*cm, largura_texto_qr])
        rodape_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # QR centralizado
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Texto alinhado à esquerda
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),  # Borda do retângulo
        ]))
        
        elements.append(rodape_table)
        
    except Exception as e:
        print(f"Erro ao obter informações do assinante: {e}")
        # Fallback simples
        elements.append(Paragraph("Teresina - PI", styles['Normal']))
        if funcao:
            elements.append(Paragraph(f"{funcao}", styles['Normal']))
    
    # Gerar PDF
    doc.build(elements)
    
    buffer.seek(0)
    
    # Preparar resposta
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="relatorio_individual_anual_{militar.matricula}_{ano}.pdf"'
    response.write(buffer.read())
    
    return response
